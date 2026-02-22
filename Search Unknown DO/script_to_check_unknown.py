#!/usr/bin/env python3
"""
Search Unknown DO Script (Odoo 18)

Reads Unknown Setsco.xlsx and for each row:
1. Searches stock.move.line via Odoo RPC with:
   - picking_id.partner_id.name = column A (Company Name)
   - product_id.barcode = column E (Item Code)
   - quantity = column I (Quantity)
   - date > 31 January 2026
2. If match found: writes stock.move.line.picking_id.name to column L (DO Number)
   - Uses first record if multiple matches
3. If not found: skips (no change)

Excel columns (0-based, header row 1):
  A(0): Company Name    B(1): Customer Reference   C(2): Invoice Date
  D(3): Invoice Number  E(4): Item Code             F(5): Key By
  G(6): PO #            H(7): Project               I(8): Quantity
  J(9): Salesperson     K(10): Serial Number       L(11): Do Number
"""

import sys
import logging
from pathlib import Path
from typing import Optional, Any

import xmlrpc.client

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
if not logger.handlers:
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    logger.addHandler(h)

script_dir = Path(__file__).resolve().parent
log_file = script_dir / "check_unknown_do.log"
fh = logging.FileHandler(log_file, mode="a", encoding="utf-8")
fh.setLevel(logging.DEBUG)
fh.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
logger.addHandler(fh)

# Odoo connection
ODOO_URL = 'http://localhost:8099'
ODOO_DB = 'lingjack-test4'
ODOO_USERNAME = 'dataimport'
ODOO_PASSWORD = 'Admin@123456'

EXCEL_FILE = script_dir / "Unknown Setsco.xlsx"
OUTPUT_FILE = script_dir / "Unknown Setsco.xlsx"  # overwrite in place

# Column indices (0-based)
COL_COMPANY_NAME = 0  # A: Company Name -> picking_id.partner_id.name
COL_ITEM_CODE = 4     # E: Item Code -> product_id.barcode
COL_QUANTITY = 8      # I: Quantity
COL_DO_NUMBER = 11    # L: Do Number (output)

# Date filter: stock.move.line.date must be after 31 January 2026
MIN_DATE = "2026-01-31"


def _normalize_str(val: Any) -> Optional[str]:
    """Normalize cell value to string."""
    if val is None:
        return None
    if isinstance(val, float) and val != val:  # NaN
        return None
    s = str(val).strip()
    return s if s else None


def _normalize_item_code(val: Any) -> Optional[str]:
    """Normalize item code to text (avoid Excel scientific notation e.g. 2.08016E+12)."""
    if val is None:
        return None
    if isinstance(val, float):
        if val != val:  # NaN
            return None
        if val == int(val):
            return format(int(val), "d")
        return format(val, ".0f")
    if isinstance(val, int):
        return str(val)
    s = str(val).strip()
    return s if s else None


def _normalize_quantity(val: Any) -> Optional[float]:
    """Parse quantity from Excel cell."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


class OdooDoSearch:
    def __init__(self, url: str, db: str, username: str, password: str):
        self.db = db
        self.username = username
        self.password = password
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
        self.uid = common.authenticate(db, username, password, {})
        if not self.uid:
            raise Exception(f"Authentication failed for {username} on {db}")
        self.models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")
        logger.info("Connected to Odoo %s as %s", db, username)

    def search_move_line_by_barcode_qty_date(
        self,
        barcode: str,
        quantity: float,
        min_date: str,
        partner_name: Optional[str] = None,
    ) -> Optional[dict]:
        """
        Search stock.move.line by product barcode, quantity, date, and partner.
        Returns first match with picking_id.name, or None if not found.
        """
        if not barcode or quantity is None:
            return None

        domain = [
            ("product_id.barcode", "=", barcode),
            ("quantity", "=", quantity),
            ("date", ">", min_date),
        ]
        if partner_name and str(partner_name).strip():
            domain.append(("picking_id.partner_id.name", "=", str(partner_name).strip()))

        try:
            ids = self.models.execute_kw(
                self.db, self.uid, self.password,
                "stock.move.line", "search",
                [domain],
                {"limit": 1, "order": "date asc"},
            )
        except Exception as e:
            logger.warning("Search failed for barcode=%s qty=%s: %s", barcode, quantity, e)
            return None

        if not ids:
            return None

        # Read picking_id to get picking name
        lines = self.models.execute_kw(
            self.db, self.uid, self.password,
            "stock.move.line", "read",
            [ids],
            {"fields": ["picking_id"]},
        )
        if not lines or not lines[0].get("picking_id"):
            return None

        picking_id = lines[0]["picking_id"][0]
        picking_name = lines[0]["picking_id"][1]

        return {"picking_id": picking_id, "picking_name": picking_name}

    def search_move_line_fallback_default_code(
        self,
        default_code: str,
        quantity: float,
        min_date: str,
        partner_name: Optional[str] = None,
    ) -> Optional[dict]:
        """
        Fallback: search by product_id.default_code if barcode search fails.
        """
        if not default_code or quantity is None:
            return None

        domain = [
            ("product_id.default_code", "=", default_code),
            ("quantity", "=", quantity),
            ("date", ">", min_date),
        ]
        if partner_name and str(partner_name).strip():
            domain.append(("picking_id.partner_id.name", "=", str(partner_name).strip()))

        try:
            ids = self.models.execute_kw(
                self.db, self.uid, self.password,
                "stock.move.line", "search",
                [domain],
                {"limit": 1, "order": "date asc"},
            )
        except Exception as e:
            logger.warning("Search (default_code) failed for code=%s qty=%s: %s", default_code, quantity, e)
            return None

        if not ids:
            return None

        lines = self.models.execute_kw(
            self.db, self.uid, self.password,
            "stock.move.line", "read",
            [ids],
            {"fields": ["picking_id"]},
        )
        if not lines or not lines[0].get("picking_id"):
            return None

        picking_name = lines[0]["picking_id"][1]
        return {"picking_id": lines[0]["picking_id"][0], "picking_name": picking_name}


def main():
    if not HAS_OPENPYXL:
        logger.error("openpyxl required. Install with: pip install openpyxl")
        sys.exit(1)

    if not EXCEL_FILE.exists():
        logger.error("Excel file not found: %s", EXCEL_FILE)
        sys.exit(1)

    logger.info("Loading %s", EXCEL_FILE)
    wb = load_workbook(str(EXCEL_FILE), read_only=False)
    ws = wb.active

    odoorpc = OdooDoSearch(ODOO_URL, ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD)

    updated = 0
    skipped = 0
    not_found = 0

    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        company_name_cell = ws.cell(row=row_idx, column=COL_COMPANY_NAME + 1)  # 1-based
        item_code_cell = ws.cell(row=row_idx, column=COL_ITEM_CODE + 1)
        quantity_cell = ws.cell(row=row_idx, column=COL_QUANTITY + 1)
        do_number_cell = ws.cell(row=row_idx, column=COL_DO_NUMBER + 1)

        partner_name = _normalize_str(company_name_cell.value)
        item_code = _normalize_item_code(item_code_cell.value)
        quantity = _normalize_quantity(quantity_cell.value)

        if not item_code or quantity is None:
            logger.debug("Row %d: skip (missing item_code or quantity)", row_idx)
            skipped += 1
            continue

        # Skip if DO Number already filled
        existing_do = do_number_cell.value
        if existing_do and str(existing_do).strip():
            logger.debug("Row %d: skip (DO already set: %s)", row_idx, existing_do)
            skipped += 1
            continue

        # Search stock.move.line
        result = odoorpc.search_move_line_by_barcode_qty_date(
            barcode=item_code,
            quantity=quantity,
            min_date=MIN_DATE,
            partner_name=partner_name,
        )

        if not result:
            # Fallback: try product_id.default_code (Item Code might be default_code)
            result = odoorpc.search_move_line_fallback_default_code(
                default_code=item_code,
                quantity=quantity,
                min_date=MIN_DATE,
                partner_name=partner_name,
            )

        if result:
            picking_name = result["picking_name"]
            do_number_cell.value = picking_name
            updated += 1
            logger.info("Row %d: found DO %s (item=%s qty=%s)", row_idx, picking_name, item_code, quantity)
        else:
            not_found += 1
            logger.debug("Row %d: no match (item=%s qty=%s)", row_idx, item_code, quantity)

    wb.save(OUTPUT_FILE)
    logger.info("Done. Updated=%d, Not found=%d, Skipped=%d. Saved to %s", updated, not_found, skipped, OUTPUT_FILE)


if __name__ == "__main__":
    main()
