#!/usr/bin/env python3
"""
Control Tag – Expand Serial Ranges into New Sheet

Reads control-tag.xlsx (first sheet) and creates a new sheet "Update control tag"
where column B (Serial Range) is expanded so that every serial in a range gets
one row. E.g. "RW7507001-RW7508000" becomes 1000 rows (RW7507001, RW7507002, ... RW7508000).

Optional: use --user-excel (e.g. user_control_tag.xlsx) to exclude "already used" serials.
Column D in that file contains ranges (e.g. RW7507001-RW7508000). Serials in those ranges
are removed from the update sheet (not added); only serials not in the user file are included.

Usage:
  python control-tag-expand-ranges.py [--excel path] [--output path] [--user-excel path]
  --excel       : Source control-tag Excel (default: control-tag.xlsx in script dir).
  --output      : Output Excel path. If not set, overwrites the source file (adds new sheet).
  --user-excel   : Exclude these serials (already used). Column D in range form. Default: user_control_tag.xlsx.
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Any, List, Optional, Set

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

script_dir = Path(__file__).resolve().parent
DEFAULT_EXCEL = script_dir / "control-tag.xlsx"
DEFAULT_USER_EXCEL = script_dir / "user_control_tag.xlsx"


def generate_serial_range(start: str, end: str) -> List[str]:
    """Generate list of serials from start to end (e.g. RW7507001..RW7508000)."""
    if not start or not end:
        return []
    start = start.strip()
    end = end.strip()
    try:
        start_match = re.match(r"^(.+?)(\d+)$", start)
        end_match = re.match(r"^(.+?)(\d+)$", end)
        if start_match and end_match:
            start_prefix, start_num_str = start_match.groups()
            end_prefix, end_num_str = end_match.groups()
            if start_prefix != end_prefix:
                try:
                    start_num = int(start)
                    end_num = int(end)
                except ValueError:
                    return []
                num_length = max(len(str(start_num)), len(str(end_num)))
                return [str(n).zfill(num_length) for n in range(start_num, end_num + 1)]
            start_num = int(start_num_str)
            end_num = int(end_num_str)
            num_length = len(start_num_str)
            if start_num > end_num:
                return []
            return [f"{start_prefix}{n:0{num_length}d}" for n in range(start_num, end_num + 1)]
        return []
    except (ValueError, AttributeError):
        return []


def parse_serial_range(cell: Any) -> List[str]:
    """
    Parse 'Serial Range' cell into a list of serial numbers.
    Supports: 'RW7507001-RW7508000', 'RW7149421 - RW7150000', 'RW7148988'.
    """
    if cell is None or (hasattr(pd, "isna") and pd.isna(cell)):
        return []
    s = str(cell).strip()
    if not s:
        return []
    if "-" in s:
        parts = [p.strip() for p in s.split("-", 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            return generate_serial_range(parts[0], parts[1])
    return [s] if s else []


def load_exclude_serials_from_user_excel(user_excel_path: Path) -> Set[str]:
    """
    Read user_control_tag.xlsx (first sheet), column D (index 3).
    Column D contains ranges (e.g. RW7507001-RW7508000). Expand and return the set of
    serial numbers that are "already used" – these will be excluded from the update sheet.
    First row is treated as header and skipped.
    """
    df = pd.read_excel(user_excel_path, sheet_name=0, header=None)
    col_d = df.iloc[1:, 3]
    exclude: Set[str] = set()
    for cell in col_d:
        for serial in parse_serial_range(cell):
            if serial:
                exclude.add(serial)
    return exclude


def expand_df(
    df: pd.DataFrame,
    exclude_serials: Optional[Set[str]] = None,
) -> pd.DataFrame:
    """
    Expand dataframe so that each row with a range in 'Serial Range' becomes
    one row per serial. If exclude_serials is set, serials in that set are not added.
    """
    out_rows: List[dict] = []
    for _, row in df.iterrows():
        com_no = row.get("Com No")
        serial_range_cell = row.get("Serial Range")
        no = row.get("No.")
        unit = row.get("Unit", "")
        serials = parse_serial_range(serial_range_cell)
        if not serials:
            val = serial_range_cell
            if val is None or (hasattr(pd, "isna") and pd.isna(val)):
                val = ""
            else:
                val = str(val).strip() if not isinstance(val, str) else val
            if exclude_serials is None or val not in exclude_serials:
                out_rows.append({"Com No": com_no, "Serial Range": val, "No.": no, "Unit": unit, "All": 1})
        else:
            for serial in serials:
                if exclude_serials is None or serial not in exclude_serials:
                    out_rows.append({"Com No": com_no, "Serial Range": serial, "No.": no, "Unit": unit, "All": 1})
    return pd.DataFrame(out_rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Expand control-tag Serial Range into Update control tag sheet.")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Source control-tag Excel path.")
    parser.add_argument("--output", type=Path, default=None, help="Output Excel path (default: same as source).")
    parser.add_argument(
        "--user-excel",
        type=Path,
        default=None,
        help="Exclude serials in this file (already used). Column D in range form. Default: user_control_tag.xlsx in script dir if present.",
    )
    args = parser.parse_args()

    excel_path = args.excel.resolve()
    if not excel_path.exists():
        print(f"Error: Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    exclude_serials: Optional[Set[str]] = None
    user_excel_path = args.user_excel.resolve() if args.user_excel else DEFAULT_USER_EXCEL
    if user_excel_path.exists():
        exclude_serials = load_exclude_serials_from_user_excel(user_excel_path)
        print(f"Excluding {len(exclude_serials)} already-used serials from column D in {user_excel_path.name}")
    elif args.user_excel:
        print(f"Warning: User Excel not found: {user_excel_path}, including all serials.", file=sys.stderr)

    # Read first sheet like control-tag.py: no header, then assign column names and skip first row
    df = pd.read_excel(excel_path, sheet_name=0, header=None)
    if df.shape[0] < 2:
        print("Error: No data rows in Excel.", file=sys.stderr)
        sys.exit(1)

    df.columns = ["Com No", "Serial Range", "No.", "Unit"]
    # Keep header row for reference; data starts at index 1
    header_row = df.iloc[0].tolist()
    data_df = df.iloc[1:].copy()

    expanded = expand_df(data_df, exclude_serials=exclude_serials)

    out_path = args.output.resolve() if args.output else excel_path
    if load_workbook is None:
        # Fallback: write only the new sheet to a new file with pandas
        expanded.to_excel(out_path, sheet_name="Update control tag", index=False)
        print(f"Wrote sheet 'Update control tag' to {out_path} (pandas only; openpyxl not installed).")
        return

    # Load workbook, add or replace "Update control tag" sheet, save
    wb = load_workbook(excel_path)
    if "Update control tag" in wb.sheetnames:
        del wb["Update control tag"]
    ws = wb.create_sheet("Update control tag", index=1)  # after first sheet

    # Row 1: header (first sheet columns + All)
    for col, value in enumerate(header_row, start=1):
        ws.cell(row=1, column=col, value=value)
    ws.cell(row=1, column=5, value="All")

    # Rows 2+: expanded data (column All = 1)
    for excel_row, (_, row) in enumerate(expanded.iterrows(), start=2):
        ws.cell(row=excel_row, column=1, value=row.get("Com No"))
        ws.cell(row=excel_row, column=2, value=row.get("Serial Range"))
        ws.cell(row=excel_row, column=3, value=row.get("No."))
        ws.cell(row=excel_row, column=4, value=row.get("Unit"))
        ws.cell(row=excel_row, column=5, value=1)

    wb.save(out_path)
    print(f"Sheet 'Update control tag' written to {out_path} ({len(expanded)} rows).")


if __name__ == "__main__":
    main()
