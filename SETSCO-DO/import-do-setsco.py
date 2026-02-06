#!/usr/bin/env python3
"""
SETSCO Delivery Order Import Script (Odoo 18)

Reads DeliveryOrderSetco.xlsx and:
1. Links setsco.serial.number to stock.move.line (by picking name + product).
2. Links setsco.serial.number to account.move (invoice by old_move).
3. Sets setsco.serial.number state to 'delivered'.

Excel columns (0-based):
  - Col 0 (A): old_move_id -> account.move.old_move (invoice lookup)
  - Col 1 (B): stock.picking name
  - Col 2: Delivery Date (ignored)
  - Col 3: item_code -> product.product.default_code
  - Col 4: quantity
  - Col 5: setsco remarks -> setsco.serial.number name(s), comma-separated

Edge cases:
  - If move_line not found but setsco serial exists and state != 'warehouse': set state to 'warehouse'.
  - If move_line not found and setsco serial not found: create setsco.serial.number and set state to 'warehouse'.
"""

import sys
import logging
import re
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Tuple, Any, Dict

import xmlrpc.client

# Optional: pandas for Excel; fallback to openpyxl only
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
if not logger.handlers:
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    logger.addHandler(h)

script_dir = Path(__file__).resolve().parent
log_file = script_dir / "import_do_setsco.log"
fh = logging.FileHandler(log_file, mode="a", encoding="utf-8")
fh.setLevel(logging.DEBUG)
fh.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
logger.addHandler(fh)

# Config: use BOM config if available
bom_dir = script_dir.parent.parent / "BOM"
config_path = bom_dir / "config.py"
if bom_dir.exists() and config_path.exists() and str(bom_dir) not in sys.path:
    sys.path.insert(0, str(bom_dir))

# Odoo connection (override via env or config if needed)
ODOO_URL = 'http://localhost:8099'
ODOO_DB = 'lingjack-test4'
ODOO_USERNAME = 'dataimport'
ODOO_PASSWORD = 'Admin@123456'

# ODOO_URL = 'https://lingjack.odoo.com/'
# ODOO_DB = 'alitecpteltd-lingjack-main-21976694'
# ODOO_USERNAME = 'dataimport'
# ODOO_PASSWORD = 'Admin@123456'
DRY_RUN = False

EXCEL_FILE = script_dir / "DeliveryOrderSetco.xlsx"
ERROR_LIST_FILE = script_dir / "do-setsco-error-list.xlsx"

# Error list sheet names (must match existing Excel if present)
SHEET_DO_FOUND = "DO Found"
SHEET_MOVE_LINE_NOT_FOUND = "Move Line not found"
SHEET_SETCO_NOT_CREATED = "Setco not created"
SHEET_PICKING_NOT_CREATED = "Picking not created"

# Column indices (0-based)
COL_OLD_MOVE_ID = 0   # A: account.move.old_move
COL_PICKING_NAME = 1  # B: stock.picking name
COL_DELIVERY_DATE = 2 # ignored
COL_ITEM_CODE = 3     # product.product.default_code
COL_QUANTITY = 4      # quantity
COL_SETSCO_REMARKS = 5  # setsco.serial.number name(s)


def _normalize_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, float):
        if val != val:  # NaN
            return None
        val = int(val)
    s = str(val).strip()
    return s if s else None


def _normalize_item_code(val: Any) -> Optional[str]:
    """Normalize item code to text so it never displays as scientific notation (e.g. 2.08016E+12).
    Excel may store long numbers as float; we convert to full digit string.
    """
    if val is None:
        return None
    if isinstance(val, float):
        if val != val:  # NaN
            return None
        # Avoid scientific notation: format as integer string (no decimals)
        if val == int(val):
            return format(int(val), "d")
        return format(val, ".0f")
    if isinstance(val, int):
        return str(val)
    s = str(val).strip()
    return s if s else None


def _expand_one_range(segment: str) -> List[str]:
    """Expand a single segment to a list of serial names.
    - 'A2011247 - A2011261' -> A2011247, A2011248, ..., A2011261 (same prefix, zero-padded).
    - 'A229092' -> ['A229092'] (single).
    """
    segment = (segment or "").strip()
    if not segment:
        return []
    if " - " not in segment:
        return [segment]
    parts = segment.split(" - ", 1)
    if len(parts) != 2:
        return [segment]
    left, right = parts[0].strip(), parts[1].strip()
    # Parse "PREFIX" + "DIGITS" (e.g. A2011247 -> prefix A, num 2011247)
    match_left = re.match(r"^([A-Za-z]*)(\d+)$", left)
    match_right = re.match(r"^([A-Za-z]*)(\d+)$", right)
    if not match_left or not match_right:
        return [segment]
    prefix_l, str_l = match_left.group(1), match_left.group(2)
    prefix_r, str_r = match_right.group(1), match_right.group(2)
    if prefix_l != prefix_r:
        return [segment]
    num_l, num_r = int(str_l), int(str_r)
    width = max(len(str_l), len(str_r))
    if num_l > num_r:
        num_l, num_r = num_r, num_l
    return [f"{prefix_l}{n:0{width}d}" for n in range(num_l, num_r + 1)]


def _parse_serial_names(remarks: Any) -> List[str]:
    """Parse 'setsco remarks' (column F) into a list of serial names.
    - Comma ',' separates items: "A229092,A229103" -> exactly those two.
    - "A2011247 - A2011261" within a segment means range (A2011247 to A2011261).
    - Mixed: "A2013776 - A2013835,A2013716 - A2013775" -> first range + second range.
    """
    s = _normalize_str(remarks)
    if not s:
        return []
    result = []
    for segment in s.split(","):
        result.extend(_expand_one_range(segment))
    return result


def load_excel(path: Path, header_row: int = 0) -> List[dict]:
    """Load Excel into list of row dicts. Keys: old_move_id, picking_name, item_code, quantity, setsco_names (list).
    If header_row=0, first row is treated as header (pandas) or skipped (openpyxl).
    """
    if not path.exists():
        raise FileNotFoundError(f"Excel not found: {path}")
    rows = []
    if HAS_PANDAS:
        df = pd.read_excel(path, sheet_name=0, header=0)
        # Normalize column access by index if no headers
        for _, row in df.iterrows():
            old_move = row.iloc[COL_OLD_MOVE_ID] if len(row) > COL_OLD_MOVE_ID else None
            picking_name = row.iloc[COL_PICKING_NAME] if len(row) > COL_PICKING_NAME else None
            item_code = row.iloc[COL_ITEM_CODE] if len(row) > COL_ITEM_CODE else None
            qty = row.iloc[COL_QUANTITY] if len(row) > COL_QUANTITY else None
            remarks = row.iloc[COL_SETSCO_REMARKS] if len(row) > COL_SETSCO_REMARKS else None
            old_move_str = _normalize_str(old_move)
            picking_str = _normalize_str(picking_name)
            item_str = _normalize_item_code(item_code)
            try:
                qty_val = float(qty) if qty is not None and str(qty).strip() else 0
            except (TypeError, ValueError):
                qty_val = 0
            names = _parse_serial_names(remarks)
            if picking_str is None and item_str is None and not names:
                continue
            delivery_date = row.iloc[COL_DELIVERY_DATE] if len(row) > COL_DELIVERY_DATE else None
            rows.append({
                "old_move_id": old_move_str,
                "picking_name": picking_str,
                "delivery_date": delivery_date,
                "item_code": item_str,
                "quantity": qty_val,
                "setsco_names": names,
                "raw_row": [old_move, picking_name, delivery_date, item_str, qty, remarks],
            })
    elif HAS_OPENPYXL:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 1:
                continue
            if not row:
                continue
            old_move = row[COL_OLD_MOVE_ID] if len(row) > COL_OLD_MOVE_ID else None
            picking_name = row[COL_PICKING_NAME] if len(row) > COL_PICKING_NAME else None
            item_code = row[COL_ITEM_CODE] if len(row) > COL_ITEM_CODE else None
            qty = row[COL_QUANTITY] if len(row) > COL_QUANTITY else None
            remarks = row[COL_SETSCO_REMARKS] if len(row) > COL_SETSCO_REMARKS else None
            old_move_str = _normalize_str(old_move)
            picking_str = _normalize_str(picking_name)
            item_str = _normalize_item_code(item_code)
            try:
                qty_val = float(qty) if qty is not None and str(qty).strip() else 0
            except (TypeError, ValueError):
                qty_val = 0
            names = _parse_serial_names(remarks)
            if picking_str is None and item_str is None and not names:
                continue
            delivery_date = row[COL_DELIVERY_DATE] if len(row) > COL_DELIVERY_DATE else None
            rows.append({
                "old_move_id": old_move_str,
                "picking_name": picking_str,
                "delivery_date": delivery_date,
                "item_code": item_str,
                "quantity": qty_val,
                "setsco_names": names,
                "raw_row": [old_move, picking_name, delivery_date, item_str, qty, remarks],
            })
        wb.close()
    else:
        raise RuntimeError("Install pandas or openpyxl to read Excel.")
    return rows


def _ensure_error_sheets(wb: Any) -> None:
    """Ensure workbook has the 4 error sheets with headers (openpyxl)."""
    headers = ["Old Move ID", "Picking Name", "Delivery Date", "Item Code", "Quantity", "Setsco Remarks", "Reason"]
    sheet_names = [SHEET_DO_FOUND, SHEET_MOVE_LINE_NOT_FOUND, SHEET_SETCO_NOT_CREATED, SHEET_PICKING_NOT_CREATED]
    for name in sheet_names:
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws.append(headers)
        else:
            ws = wb[name]
            if ws.max_row == 0:
                ws.append(headers)


def write_error_excel(collector: Dict[str, List[dict]], path: Path) -> None:
    """Append collected rows to do-setsco-error-list.xlsx (4 sheets)."""
    if not HAS_OPENPYXL:
        logger.warning("openpyxl not available; cannot write error list %s", path)
        return
    from openpyxl import load_workbook, Workbook
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.remove(wb.active)
    _ensure_error_sheets(wb)
    for sheet_name, rows in collector.items():
        if not rows:
            continue
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        if ws.max_row == 0:
            ws.append(["Old Move ID", "Picking Name", "Delivery Date", "Item Code", "Quantity", "Setsco Remarks", "Reason"])
        for item in rows:
            raw = item.get("raw_row") or []
            reason = item.get("reason", "")
            row_values = list(raw)[:6]
            while len(row_values) < 6:
                row_values.append("")
            # Item Code (column D, index 3): always text so Excel does not show scientific notation
            if len(row_values) > 3:
                row_values[3] = _normalize_item_code(row_values[3]) or ""
            row_values.append(reason)
            ws.append(row_values)
            # Set Item Code cell format to text so it displays as full digits
            if len(row_values) > 3:
                _cell = ws.cell(row=ws.max_row, column=4)
                _cell.number_format = "@"
                if row_values[3] != "":
                    _cell.value = row_values[3]
    wb.save(path)
    logger.info("Wrote error list to %s", path)


class OdooDoSetscoImporter:
    def __init__(self, url: str, db: str, username: str, password: str):
        self.db = db
        self.username = username
        self.password = password
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
        self.uid = common.authenticate(db, username, password, {})
        if not self.uid:
            raise Exception(f"Authentication failed for {username} on {db}")
        self.models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")
        logger.info("Connected to Odoo %s as %s", db, username)

    def _search(self, model: str, domain: list, limit: Optional[int] = None) -> List[int]:
        kwargs = {}
        if limit is not None:
            kwargs["limit"] = limit
        return self.models.execute_kw(self.db, self.uid, self.password, model, "search", [domain], kwargs)

    def _read(self, model: str, ids: List[int], fields: List[str]) -> List[dict]:
        if not ids:
            return []
        return self.models.execute_kw(self.db, self.uid, self.password, model, "read", [ids], {"fields": fields})

    def _write(self, model: str, ids: List[int], vals: dict) -> bool:
        vals = {k: v for k, v in vals.items() if v is not None}
        return self.models.execute_kw(self.db, self.uid, self.password, model, "write", [ids, vals])

    def _create(self, model: str, vals: dict) -> int:
        vals = {k: v for k, v in vals.items() if v is not None}
        return self.models.execute_kw(self.db, self.uid, self.password, model, "create", [vals])

    def _call(self, model: str, method: str, ids: List[int], *args, **kwargs) -> Any:
        return self.models.execute_kw(self.db, self.uid, self.password, model, method, [ids] + list(args), kwargs)

    def find_invoice_by_old_move(self, old_move_val: Optional[str]) -> Optional[int]:
        if not old_move_val:
            return None
        try:
            old_id = int(float(str(old_move_val).strip()))
        except (ValueError, TypeError):
            return None
        ids = self._search("account.move", [
            ("old_move", "=", old_id),
            ("move_type", "in", ["out_invoice", "out_refund"]),
        ], limit=1)
        return ids[0] if ids else None

    def find_picking_by_name(self, name: Optional[str]) -> Optional[int]:
        if not name:
            return None
        name = str(name).strip()
        ids = self._search("stock.picking", [("name", "=", name)], limit=1)
        return ids[0] if ids else None

    def find_product_by_default_code(self, code: Optional[str]) -> Optional[int]:
        if not code:
            return None
        code = str(code).strip()
        if isinstance(code, float):
            code = str(int(code))
        ids = self._search("product.product", [("default_code", "=", code)], limit=1)
        return ids[0] if ids else None

    def find_move_lines(self, picking_name: str, product_id: int) -> List[int]:
        """Stock move lines for this picking and product (outgoing)."""
        picking_ids = self._search("stock.picking", [("name", "=", picking_name)], limit=1)
        if not picking_ids:
            return []
        ids = self._search("stock.move.line", [
            ("picking_id", "=", picking_ids[0]),
            ("product_id", "=", product_id),
        ], limit=None)
        return ids

    def find_setsco_serial_by_name(self, name: str) -> Optional[int]:
        name = str(name).strip()
        ids = self._search("setsco.serial.number", [("name", "=", name)], limit=1)
        return ids[0] if ids else None

    def get_product_setsco_category(self, product_id: int) -> Optional[int]:
        """Get setsco_category_id from product.template (product.product has related)."""
        prods = self._read("product.product", [product_id], ["product_tmpl_id"])
        if not prods or not prods[0].get("product_tmpl_id"):
            return None
        tmpl_id = prods[0]["product_tmpl_id"][0]
        tmpls = self._read("product.template", [tmpl_id], ["setsco_category_id"])
        if not tmpls or not tmpls[0].get("setsco_category_id"):
            return None
        return tmpls[0]["setsco_category_id"][0]

    def get_stock_location_stock(self) -> Optional[int]:
        ids = self._search("stock.location", [("usage", "=", "internal"), ("name", "=", "Stock")], limit=1)
        return ids[0] if ids else None

    def process_row(self, row: dict, error_collector: Dict[str, List[dict]]) -> Tuple[int, int, int]:
        """Process one Excel row. Returns (linked_count, warehouse_count, created_count).
        Appends to error_collector for DO Found, Move Line not found, Setco not created as needed.
        """
        old_move_val = row.get("old_move_id")
        picking_name = row.get("picking_name")
        item_code = row.get("item_code")
        setsco_names = row.get("setsco_names") or []
        raw_row = row.get("raw_row") or []

        if not setsco_names:
            return (0, 0, 0)

        # Picking check: when column B (picking_name) is present, picking must exist (handled in run() before calling process_row)
        picking_id = self.find_picking_by_name(picking_name) if picking_name else None
        if picking_name and not picking_id:
            error_collector[SHEET_PICKING_NOT_CREATED].append({
                "raw_row": raw_row,
                "reason": "Picking not found in Odoo",
            })
            return (0, 0, 0)

        if picking_name and picking_id:
            error_collector[SHEET_DO_FOUND].append({"raw_row": raw_row, "reason": ""})

        invoice_id = self.find_invoice_by_old_move(old_move_val)
        product_id = self.find_product_by_default_code(item_code) if item_code else None
        move_line_ids = []
        if picking_name and product_id:
            move_line_ids = self.find_move_lines(picking_name, product_id)

        move_line_id = move_line_ids[0] if move_line_ids else None
        picking_id_from_line = None
        if move_line_ids:
            lines = self._read("stock.move.line", [move_line_ids[0]], ["picking_id"])
            if lines and lines[0].get("picking_id"):
                picking_id_from_line = lines[0]["picking_id"][0]
        picking_id = picking_id or picking_id_from_line

        if picking_name and product_id and not move_line_id:
            error_collector[SHEET_MOVE_LINE_NOT_FOUND].append({
                "raw_row": raw_row,
                "reason": "No move line for this picking + product",
            })

        linked, to_warehouse, created = 0, 0, 0
        stock_location_id = self.get_stock_location_stock() if (not move_line_id) else None
        setsco_category_id = None
        if product_id:
            setsco_category_id = self.get_product_setsco_category(product_id)
        setco_not_created_reasons: List[str] = []

        for name in setsco_names:
            serial_id = self.find_setsco_serial_by_name(name)
            if move_line_id:
                if serial_id:
                    if not DRY_RUN:
                        self._write("setsco.serial.number", [serial_id], {
                            "move_line_id": move_line_id,
                            "delivery_picking_id": picking_id,
                            "delivery_move_line_id": move_line_id,
                            "state": "delivered",
                            "delivery_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "invoice_id": invoice_id,
                        })
                    linked += 1
                    logger.info("Linked serial %s -> move_line %s, delivered, invoice %s", name, move_line_id, invoice_id)
                else:
                    if not setsco_category_id and product_id:
                        setsco_category_id = self.get_product_setsco_category(product_id)
                    if not setsco_category_id:
                        setco_not_created_reasons.append("No setsco_category for product %s (serial %s)" % (item_code or "?", name))
                        logger.warning("No setsco_category for product %s; skip create for serial %s", item_code, name)
                        continue
                    if not DRY_RUN:
                        new_id = self._create("setsco.serial.number", {
                            "name": name,
                            "setsco_category_id": setsco_category_id,
                            "product_id": product_id,
                            "state": "warehouse",
                            "location_id": stock_location_id,
                        })
                        self._write("setsco.serial.number", [new_id], {
                            "move_line_id": move_line_id,
                            "delivery_picking_id": picking_id,
                            "delivery_move_line_id": move_line_id,
                            "state": "delivered",
                            "delivery_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "invoice_id": invoice_id,
                        })
                    created += 1
                    logger.info("Created serial %s, linked to move_line %s, delivered", name, move_line_id)
            else:
                if serial_id:
                    recs = self._read("setsco.serial.number", [serial_id], ["state"])
                    if recs and recs[0].get("state") != "warehouse":
                        if not DRY_RUN:
                            self._write("setsco.serial.number", [serial_id], {"state": "warehouse"})
                        to_warehouse += 1
                        logger.info("Serial %s: no move line; set state to warehouse", name)
                else:
                    if not setsco_category_id and product_id:
                        setsco_category_id = self.get_product_setsco_category(product_id)
                    if not setsco_category_id:
                        setco_not_created_reasons.append("No setsco_category for product %s (serial %s)" % (item_code or "?", name))
                        logger.warning("No setsco_category for product %s; skip create for serial %s", item_code, name)
                        continue
                    if not DRY_RUN:
                        self._create("setsco.serial.number", {
                            "name": name,
                            "setsco_category_id": setsco_category_id,
                            "product_id": product_id or False,
                            "state": "warehouse",
                            "location_id": stock_location_id,
                        })
                    created += 1
                    logger.info("Created serial %s (no move line), state=warehouse", name)

        if setco_not_created_reasons:
            error_collector[SHEET_SETCO_NOT_CREATED].append({
                "raw_row": raw_row,
                "reason": "; ".join(setco_not_created_reasons),
            })

        return (linked, to_warehouse, created)

    def run(self, excel_path: Path, error_list_path: Optional[Path] = None) -> dict:
        rows = load_excel(excel_path)
        logger.info("Loaded %d rows from %s", len(rows), excel_path)
        error_collector = {
            SHEET_DO_FOUND: [],
            SHEET_MOVE_LINE_NOT_FOUND: [],
            SHEET_SETCO_NOT_CREATED: [],
            SHEET_PICKING_NOT_CREATED: [],
        }
        total_linked, total_warehouse, total_created = 0, 0, 0
        for i, row in enumerate(rows):
            try:
                a, b, c = self.process_row(row, error_collector)
                total_linked += a
                total_warehouse += b
                total_created += c
            except Exception as e:
                logger.exception("Row %s error: %s", i + 1, e)
        if error_list_path is None:
            error_list_path = ERROR_LIST_FILE
        write_error_excel(error_collector, error_list_path)
        return {"linked": total_linked, "to_warehouse": total_warehouse, "created": total_created}


def main():
    if DRY_RUN:
        logger.info("DRY RUN - no writes will be performed")
    if not EXCEL_FILE.exists():
        logger.error("Excel file not found: %s", EXCEL_FILE)
        sys.exit(1)
    importer = OdooDoSetscoImporter(ODOO_URL, ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD)
    result = importer.run(EXCEL_FILE, ERROR_LIST_FILE)
    logger.info("Done. Linked: %d, To warehouse: %d, Created: %d", result["linked"], result["to_warehouse"], result["created"])
    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
