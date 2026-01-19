#!/usr/bin/env python3
"""
Odoo 18 MRP Production Import Script

Reads manufacturing orders from an Excel file (output.xlsx) and
creates mrp.production records with components via XML-RPC.

Rows are grouped by ProductWorkOrderID (PWO ID).
First row with Product value = main MO, subsequent rows = components.
"""

import sys
import logging
from typing import Dict, List, Optional
from datetime import datetime, date
import os
from collections import defaultdict

import xmlrpc.client
from openpyxl import load_workbook

# Set up logging to both console and file
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Remove existing handlers to avoid duplicates
logger.handlers = []

# Console handler
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)

# File handler for errors and detailed logs
log_file_path = os.path.join(os.path.dirname(__file__), 'import_mrp_errors.log')
file_handler = logging.FileHandler(log_file_path, mode='a', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_formatter = logging.Formatter(
    "%(asctime)s - %(levelname)s - %(name)s - %(message)s",
    datefmt='%Y-%m-%d %H:%M:%S'
)
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

logger.info("=" * 60)
logger.info("MRP Production Import Script Started")
logger.info(f"Log file: {log_file_path}")
logger.info("=" * 60)


class OdooMRPProductionImporter:
    """Import MRP Production Orders from Excel to Odoo 18"""

    def __init__(self, url: str, db: str, username: str, password: str):
        """
        Initialize Odoo connection

        Args:
            url: Odoo server URL (e.g., 'http://localhost:8069')
            db: Database name
            username: Odoo username
            password: Odoo password
        """
        self.url = url
        self.db = db
        self.username = username
        self.password = password

        # Authenticate
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
        self.uid = common.authenticate(db, username, password, {})
        if not self.uid:
            raise Exception(
                f"Authentication failed for user '{username}' "
                f"on database '{db}'."
            )

        self.models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")
        logger.info("Connected to Odoo DB '%s' as '%s'", db, username)

    # ---------------- Generic helpers ------------------

    def _search(self, model: str, domain: list, limit: int = 1) -> List[int]:
        """Search records in Odoo"""
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'search',
            [domain],
            {'limit': limit}
        )

    def _create(self, model: str, vals: dict) -> int:
        """Create a record in Odoo"""
        # Filter out None values - XML-RPC cannot marshal None
        filtered_vals = {k: v for k, v in vals.items() if v is not None}
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'create',
            [filtered_vals]
        )

    def _read(self, model: str, ids: List[int], fields: List[str]) -> List[dict]:
        """Read records from Odoo"""
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'read',
            [ids],
            {'fields': fields}
        )

    # ---------------- Lookups ------------------

    def find_product_by_default_code(self, default_code: str) -> Optional[int]:
        """Find product.product by default_code"""
        if not default_code:
            return None
        default_code = str(default_code).strip()
        if not default_code:
            return None
        product_ids = self._search('product.product', [('default_code', '=', default_code)], limit=1)
        return product_ids[0] if product_ids else None

    def create_product(self, reference: str, name: str = None, auto_create: bool = True) -> Optional[int]:
        """
        Create a product in Odoo with import_newly_created flag
        
        Args:
            reference: Product reference code (default_code)
            name: Product name (if None, uses reference as name)
            auto_create: If True, creates product if not found
            
        Returns:
            Product ID or None if creation failed
        """
        if not reference or not str(reference).strip():
            return None
        
        reference = str(reference).strip()
        
        # Check if product already exists
        product_id = self.find_product_by_default_code(reference)
        if product_id:
            return product_id
        
        if not auto_create:
            return None
        
        # Prepare product name
        if not name or not str(name).strip():
            name = reference
        
        name = str(name).strip()
        
        # Get default product category (All / Product category)
        try:
            # Try to find "All" category first
            categ_ids = self._search('product.category', [('name', '=', 'All')], limit=1)
            if not categ_ids:
                # Try to find any category
                categ_ids = self._search('product.category', [], limit=1)
            categ_id = categ_ids[0] if categ_ids else None
        except Exception as e:
            logger.warning(f"Could not find product category: {e}")
            categ_id = None
        
        # Get default UOM (Units)
        uom_id = self.find_uom_by_name('Units')
        
        # Create product template first
        product_vals = {
            'name': name,
            'default_code': reference,
            'type': 'consu',  # Manufacturing products are typically 'product' type
            'import_newly_created': True,
        }
        
        # Only add categ_id if we found one
        if categ_id:
            product_vals['categ_id'] = categ_id
        
        # Only add UOM if we found one
        if uom_id:
            product_vals['uom_id'] = uom_id
            product_vals['uom_po_id'] = uom_id
        
        logger.debug(f"Creating product with values: {product_vals}")
        
        try:
            # Create product template
            template_id = self._create('product.template', product_vals)
            
            # Get the product variant (product.product)
            product_ids = self._search('product.product', [('product_tmpl_id', '=', template_id)], limit=1)
            
            if product_ids:
                product_id = product_ids[0]
                logger.info(f"Created product: {name} (Reference: {reference}, ID: {product_id})")
                return product_id
            else:
                logger.error(f"Failed to create product variant for {reference}")
                return None
                
        except xmlrpc.client.Fault as e:
            # Odoo-specific error
            error_msg = str(e)
            logger.error(f"Odoo error creating product '{reference}': {error_msg}")
            return None
        except Exception as e:
            # General error
            logger.error(f"Error creating product '{reference}': {str(e)}", exc_info=True)
            return None

    def find_or_create_product(self, default_code: str, name: str = None, stats: dict = None) -> Optional[int]:
        """
        Find product by default_code, or create if not found
        
        Args:
            default_code: Product reference code
            name: Product name (optional, for creation)
            stats: Optional stats dict to track created products
            
        Returns:
            Product ID or None
        """
        product_id = self.find_product_by_default_code(default_code)
        if product_id:
            return product_id
        
        # Product not found, create it
        created_id = self.create_product(default_code, name=name, auto_create=True)
        if created_id and stats is not None:
            stats['created_products'] = stats.get('created_products', 0) + 1
        return created_id

    def find_uom_by_name(self, name: str) -> Optional[int]:
        """Find uom.uom by name (default: Units)"""
        if not name:
            name = 'Units'
        name = str(name).strip()
        if not name:
            name = 'Units'
        uom_ids = self._search('uom.uom', [('name', '=', name)], limit=1)
        if uom_ids:
            return uom_ids[0]
        # Fallback to Units
        uom_ids = self._search('uom.uom', [('name', '=', 'Units')], limit=1)
        return uom_ids[0] if uom_ids else None

    def find_or_create_opening_lot(self, product_id: int, product_code: str, dry_run: bool = False) -> Optional[int]:
        """
        Find or create an "OPENING-X" lot for a product
        
        Args:
            product_id: Product ID
            product_code: Product default_code (for lot name)
            dry_run: If True, only log without creating
            
        Returns:
            Lot ID or None
        """
        if not product_id or not product_code:
            return None
        
        lot_name = f"OPENING-{product_code}"
        
        # Search for existing lot
        lot_ids = self._search('stock.lot', [
            ('product_id', '=', product_id),
            ('name', '=', lot_name)
        ], limit=1)
        
        if lot_ids:
            return lot_ids[0]
        
        # Lot not found, create it
        if dry_run:
            logger.info(f"[DRY RUN] Would create lot '{lot_name}' for product ID {product_id}")
            return None
        
        try:
            lot_vals = {
                'product_id': product_id,
                'name': lot_name,
            }
            lot_id = self._create('stock.lot', lot_vals)
            logger.info(f"Created lot '{lot_name}' (ID: {lot_id}) for product ID {product_id}")
            return lot_id
        except Exception as e:
            logger.error(f"Error creating lot '{lot_name}' for product ID {product_id}: {e}", exc_info=True)
            return None

    def create_stock_adjustment(self, product_id: int, lot_id: int, quantity: float = 10000.0,
                                location_id: Optional[int] = None, dry_run: bool = False) -> bool:
        """
        Create a stock adjustment (inventory adjustment) to set quantity for a lot
        
        Args:
            product_id: Product ID
            lot_id: Lot ID
            quantity: Quantity to set (default: 10000)
            location_id: Location ID (default: stock location)
            dry_run: If True, only log without creating
            
        Returns:
            True if successful, False otherwise
        """
        if not product_id or not lot_id:
            return False
        
        if dry_run:
            logger.info(
                f"[DRY RUN] Would create stock adjustment: Product ID {product_id}, "
                f"Lot ID {lot_id}, Quantity {quantity}"
            )
            return False
        
        try:
            # Get default stock location if not provided
            if not location_id:
                # Try to find stock location
                location_ids = self._search('stock.location', [
                    ('usage', '=', 'internal'),
                    ('active', '=', True)
                ], limit=1)
                if location_ids:
                    location_id = location_ids[0]
                else:
                    logger.warning("Could not find stock location for adjustment")
                    return False
            
            # Find or create stock quant
            quant_ids = self._search('stock.quant', [
                ('product_id', '=', product_id),
                ('lot_id', '=', lot_id),
                ('location_id', '=', location_id)
            ], limit=1)
            
            if quant_ids:
                # Update existing quant
                self.models.execute_kw(
                    self.db, self.uid, self.password,
                    'stock.quant', 'write',
                    [[quant_ids[0]], {'inventory_quantity': quantity}]
                )
                logger.info(
                    f"Updated stock quant: Product ID {product_id}, Lot ID {lot_id}, "
                    f"Location ID {location_id}, Quantity {quantity}"
                )
            else:
                # Create new quant
                quant_vals = {
                    'product_id': product_id,
                    'lot_id': lot_id,
                    'location_id': location_id,
                    'inventory_quantity': quantity,
                }
                quant_id = self._create('stock.quant', quant_vals)
                logger.info(
                    f"Created stock quant: Product ID {product_id}, Lot ID {lot_id}, "
                    f"Location ID {location_id}, Quantity {quantity} (Quant ID: {quant_id})"
                )
            
            return True
            
        except Exception as e:
            logger.error(
                f"Error creating stock adjustment for Product ID {product_id}, "
                f"Lot ID {lot_id}: {e}",
                exc_info=True
            )
            return False

    def reset_opening_lot_quantities(self, lot_ids: List[int], dry_run: bool = False):
        """
        Reset all OPENING-X lot quantities to 0 after manufacturing is done
        
        Args:
            lot_ids: List of lot IDs to reset
            dry_run: If True, only log without updating
        """
        if not lot_ids:
            return
        
        logger.info(f"Resetting {len(lot_ids)} OPENING-X lot quantities to 0")
        
        for lot_id in lot_ids:
            try:
                # Get lot info
                lot_data = self._read('stock.lot', [lot_id], ['product_id', 'name'])
                if not lot_data:
                    continue
                
                product_id = lot_data[0]['product_id'][0]
                lot_name = lot_data[0]['name']
                
                # Find all quants for this lot
                quant_ids = self._search('stock.quant', [
                    ('product_id', '=', product_id),
                    ('lot_id', '=', lot_id)
                ])
                
                if not quant_ids:
                    continue
                
                if not dry_run:
                    # Set all quants to 0
                    self.models.execute_kw(
                        self.db, self.uid, self.password,
                        'stock.quant', 'write',
                        [quant_ids, {'inventory_quantity': 0.0}]
                    )
                    logger.info(f"Reset lot '{lot_name}' (ID: {lot_id}) quantity to 0")
                else:
                    logger.info(f"[DRY RUN] Would reset lot '{lot_name}' (ID: {lot_id}) quantity to 0")
                    
            except Exception as e:
                logger.error(f"Error resetting lot ID {lot_id}: {e}", exc_info=True)

    # ---------------- Parsing helpers ------------------

    def _to_datetime_str(self, value) -> Optional[str]:
        """Convert Excel date/datetime to ISO string (YYYY-MM-DD HH:MM:SS)"""
        if not value:
            return None
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d 00:00:00')
        if isinstance(value, str):
            value = value.strip()
            try:
                # Try various date formats
                for fmt in ['%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                    try:
                        dt = datetime.strptime(value, fmt)
                        return dt.strftime('%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        continue
                logger.warning(f"Could not parse date format: {value}")
                return value
            except Exception as e:
                logger.warning(f"Error parsing date '{value}': {e}")
                return value
        return None

    def _map_state(self, state: str) -> tuple:
        """
        Map Excel state to Odoo mrp.production state
        
        Returns:
            tuple: (state, should_mark_done) - state to set and whether to call button_mark_done
        """
        if not state:
            return ('draft', False)
        state = str(state).strip().lower()
        mapping = {
            'complete': ('confirmed', True),  # Will use button_mark_done
            'completed': ('confirmed', True),  # Will use button_mark_done
            'done': ('confirmed', True),  # Will use button_mark_done
            'cancel': ('cancel', False),
            'cancelled': ('cancel', False),
            'draft': ('draft', False),
            'confirmed': ('confirmed', False),
            'progress': ('progress', False),
            'to_close': ('to_close', False),
        }
        return mapping.get(state, ('draft', False))

    # ---------------- Excel parsing ------------------

    def parse_excel(self, excel_path: str, sheet_name: Optional[str] = None, header_row: int = 1) -> Dict[str, Dict]:
        """
        Parse Excel file and group by PWO ID

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name (default: active sheet)
            header_row: Row number of header (1-based)

        Returns:
            Dict keyed by PWO ID, containing MO data and components
        """
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        start_row = header_row + 1
        max_row = ws.max_row

        # Group by PWO ID: {pwo_id: {'mo_data': {...}, 'components': [...]}}
        mo_data = defaultdict(lambda: {'mo_data': None, 'components': []})

        for row_idx in range(start_row, max_row + 1):
            row_cells = [cell.value for cell in ws[row_idx]]

            if len(row_cells) < 15:
                continue

            # Extract values by column index (0-based)
            # Column A (0): ProductWorkOrderID (PWO ID)
            # Column B (1): Display Name (PWO Number)
            # Column C (2): Start (date_start)
            # Column D (3): End (date_finished)
            # Column E (4): Product (product_id - main product, None for components)
            # Column F (5): Quantity To Produce (product_qty)
            # Column G (6): State
            # Column H (7): Components/Product/Internal Reference (component default_code)
            # Column I (8): Components/Product/Name (component name)
            # Column J (9): Components/Quantity To Consume (component qty)
            # Column K (10): sale.order.id (old_so_id)
            # Column L (11): sale.order.reference (old_so_number)
            # Column M (12): sw.order.id (old_swo_id)
            # Column N (13): sw.order.reference (old_swo_number)
            # Column O (14): ProductWorkOrderID (duplicate)

            pwo_id = row_cells[0] if len(row_cells) > 0 else None
            pwo_number = row_cells[1] if len(row_cells) > 1 else None
            start_date = row_cells[2] if len(row_cells) > 2 else None
            end_date = row_cells[3] if len(row_cells) > 3 else None
            product_code = row_cells[4] if len(row_cells) > 4 else None
            product_qty = row_cells[5] if len(row_cells) > 5 else None
            state = row_cells[6] if len(row_cells) > 6 else None
            component_code = row_cells[7] if len(row_cells) > 7 else None
            component_name = row_cells[8] if len(row_cells) > 8 else None
            component_qty = row_cells[9] if len(row_cells) > 9 else None
            old_so_id = row_cells[10] if len(row_cells) > 10 else None
            old_so_number = row_cells[11] if len(row_cells) > 11 else None
            old_swo_id = row_cells[12] if len(row_cells) > 12 else None
            old_swo_number = row_cells[13] if len(row_cells) > 13 else None

            if not pwo_id:
                continue

            pwo_id_str = str(pwo_id).strip()

            # If Product column has value, this is the main MO row
            if product_code:
                mo_data[pwo_id_str]['mo_data'] = {
                    'pwo_id': pwo_id_str,
                    'pwo_number': str(pwo_number).strip() if pwo_number else None,
                    'start_date': start_date,
                    'end_date': end_date,
                    'product_code': str(product_code).strip() if product_code else None,
                    'product_qty': float(product_qty) if product_qty else 0.0,
                    'state': str(state).strip() if state else None,
                    'old_so_id': str(old_so_id).strip() if old_so_id else None,
                    'old_so_number': str(old_so_number).strip() if old_so_number else None,
                    'old_swo_id': str(old_swo_id).strip() if old_swo_id else None,
                    'old_swo_number': str(old_swo_number).strip() if old_swo_number else None,
                    'row_index': row_idx,
                }
            # Otherwise, this is a component row
            elif component_code:
                mo_data[pwo_id_str]['components'].append({
                    'component_code': str(component_code).strip() if component_code else None,
                    'component_name': str(component_name).strip() if component_name else None,
                    'component_qty': float(component_qty) if component_qty else 0.0,
                    'row_index': row_idx,
                })

        wb.close()
        logger.info("Parsed %d MRP Production orders from Excel", len(mo_data))
        return dict(mo_data)

    # ---------------- Import Logic ------------------

    def import_mrp_productions(self, excel_path: str, sheet_name: Optional[str] = None, dry_run: bool = True):
        """
        Import MRP production orders from Excel

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name (default: active sheet)
            dry_run: If True, only log operations without creating records
        """
        mo_data = self.parse_excel(excel_path, sheet_name)

        if not mo_data:
            logger.warning("No MRP Production orders found in Excel file")
            return

        stats = {
            'total_mo': len(mo_data),
            'created_mo': 0,
            'created_components': 0,
            'created_products': 0,
            'created_lots': 0,
            'errors': [],
        }

        # Track all OPENING-X lots created during import
        opening_lot_ids = []

        if dry_run:
            logger.info("DRY RUN MODE - No records will be created in Odoo")

        for pwo_id, data in mo_data.items():
            try:
                mo_info = data.get('mo_data')
                if not mo_info:
                    error_msg = f"PWO ID {pwo_id}: No main MO data found (missing Product column)"
                    logger.error(error_msg)
                    stats['errors'].append(error_msg)
                    continue

                # Find or create product by default_code
                product_id = self.find_or_create_product(mo_info['product_code'], stats=stats)
                if not product_id:
                    error_msg = f"PWO ID {pwo_id} (Row {mo_info['row_index']}): Failed to find or create product for '{mo_info['product_code']}'"
                    logger.error(error_msg)
                    stats['errors'].append(error_msg)
                    continue

                # Get product UOM
                product_data = self._read('product.product', [product_id], ['uom_id'])
                uom_id = product_data[0]['uom_id'][0] if product_data and product_data[0].get('uom_id') else None

                # Map state - returns (state, should_mark_done)
                mapped_state, should_mark_done = self._map_state(mo_info['state'])

                # Prepare MO values
                mo_vals = {
                    'product_id': product_id,
                    'product_qty': mo_info['product_qty'],
                    'qty_producing': mo_info['product_qty'] if should_mark_done else None,
                    'product_uom_id': uom_id,
                    'date_start': self._to_datetime_str(mo_info['start_date']),
                    'date_finished': self._to_datetime_str(mo_info['end_date']),
                    # 'state': mapped_state,
                    'old_pwo_id': mo_info['pwo_id'],
                    'old_pwo_number': mo_info['pwo_number'],
                    'old_so_id': mo_info['old_so_id'],
                    'old_so_number': mo_info['old_so_number'],
                    'old_swo_id': mo_info['old_swo_id'],
                    'bom_id': False,
                    'old_swo_number': mo_info['old_swo_number'],
                }

                # Prepare components (move_raw_ids)
                components = data.get('components', [])
                move_raw_vals = []
                for comp in components:
                    # Find or create component product
                    comp_product_id = self.find_or_create_product(
                        comp['component_code'],
                        name=comp.get('component_name'),
                        stats=stats
                    )
                    if not comp_product_id:
                        logger.warning(
                            f"PWO ID {pwo_id} (Row {comp['row_index']}): Failed to find or create component product for '{comp['component_code']}'"
                        )
                        continue

                    # Get component UOM and tracking info
                    comp_data = self._read('product.product', [comp_product_id], ['uom_id', 'tracking'])
                    comp_uom_id = comp_data[0]['uom_id'][0] if comp_data and comp_data[0].get('uom_id') else None
                    tracking = comp_data[0].get('tracking', 'none') if comp_data else 'none'
                    
                    # Prepare move raw values
                    move_raw_val = {
                        'product_id': comp_product_id,
                        # 'product_uom_id': comp_uom_id,
                        'product_uom_qty': comp['component_qty'],
                        'quantity': comp['component_qty'],
                        'picked': True,
                    }
                    
                    # If product requires lot tracking, find or create "OPENING-X" lot
                    # Note: lot_id must be set on move_line_ids, not directly on the move
                    if tracking in ('lot', 'serial'):
                        lot_id = self.find_or_create_opening_lot(
                            comp_product_id,
                            comp['component_code'],
                            dry_run=dry_run
                        )
                        if lot_id:
                            # Track the lot for later reset
                            if lot_id not in opening_lot_ids:
                                opening_lot_ids.append(lot_id)
                                stats['created_lots'] += 1
                            
                            # Create stock adjustment with 10k quantity
                            if not dry_run:
                                self.create_stock_adjustment(
                                    comp_product_id,
                                    lot_id,
                                    quantity=10000.0,
                                    dry_run=dry_run
                                )
                            
                            # Set lot_id on move_line_ids (stock.move.line), not on the move itself
                            move_raw_val['move_line_ids'] = [(0, 0, {
                                'product_id': comp_product_id,
                                'product_uom_id': comp_uom_id,
                                'qty_done': comp['component_qty'],
                                'lot_id': lot_id,
                            })]
                        else:
                            logger.warning(
                                f"PWO ID {pwo_id} (Row {comp['row_index']}): "
                                f"Failed to find or create OPENING lot for product '{comp['component_code']}' "
                                f"that requires {tracking} tracking"
                            )
                    
                    move_raw_vals.append((0, 0, move_raw_val))

                if move_raw_vals:
                    mo_vals['move_raw_ids'] = move_raw_vals

                if not dry_run:
                    # Create MO
                    mo_id = self._create('mrp.production', mo_vals)
                    stats['created_mo'] += 1
                    stats['created_components'] += len(move_raw_vals)
                    logger.info(
                        f"PWO ID {pwo_id} (Row {mo_info['row_index']}): Created MO '{mo_info['pwo_number']}' "
                        f"(ID: {mo_id}) with {len(move_raw_vals)} components"
                    )
                    
                    # If state should be "done", use button_mark_done instead of setting state directly
                    if should_mark_done:
                        try:
                            # Call button_mark_done method
                            self.models.execute_kw(
                                self.db, self.uid, self.password,
                                'mrp.production',
                                'button_set_done',
                                [[mo_id]]
                            )
                            logger.info(
                                f"PWO ID {pwo_id} (Row {mo_info['row_index']}): Marked MO '{mo_info['pwo_number']}' "
                                f"(ID: {mo_id}) as done using button_mark_done"
                            )
                        except Exception as e:
                            error_msg = f"PWO ID {pwo_id} (Row {mo_info['row_index']}): Failed to mark MO as done: {e}"
                            logger.error(error_msg, exc_info=True)
                            stats['errors'].append(error_msg)
                else:
                    logger.info(
                        f"[DRY RUN] PWO ID {pwo_id} (Row {mo_info['row_index']}): Would create MO "
                        f"'{mo_info['pwo_number']}' with {len(move_raw_vals)} components"
                    )
                    if should_mark_done:
                        logger.info(
                            f"[DRY RUN] PWO ID {pwo_id} (Row {mo_info['row_index']}): Would mark MO "
                            f"'{mo_info['pwo_number']}' as done using button_mark_done"
                        )
                    stats['created_mo'] += 1
                    stats['created_components'] += len(move_raw_vals)
                self.models.execute_kw(
                        self.db, self.uid, self.password,
                        'mrp.production',
                        'swap_old_name',
                        [[mo_id]]
                    )
            except Exception as e:
                error_msg = f"PWO ID {pwo_id}: Error processing MO: {e}"
                logger.error(error_msg, exc_info=True)
                stats['errors'].append(error_msg)

        # After all MOs are processed and marked as done, reset OPENING-X lot quantities to 0
        if opening_lot_ids and not dry_run:
            logger.info("=" * 60)
            logger.info("Resetting OPENING-X lot quantities to 0 after manufacturing completion")
            logger.info("=" * 60)
            self.reset_opening_lot_quantities(opening_lot_ids, dry_run=dry_run)
        elif opening_lot_ids and dry_run:
            logger.info(
                f"[DRY RUN] Would reset {len(opening_lot_ids)} OPENING-X lot quantities to 0 "
                f"after manufacturing completion"
            )

        # Summary
        logger.info("=" * 60)
        logger.info("Import Summary:")
        logger.info("  Total MOs processed: %d", stats['total_mo'])
        logger.info("  MOs created: %d", stats['created_mo'])
        logger.info("  Components created: %d", stats['created_components'])
        logger.info("  Products created: %d", stats.get('created_products', 0))
        logger.info("  OPENING lots created: %d", stats.get('created_lots', 0))
        logger.info("  Errors: %d", len(stats['errors']))
        logger.info("=" * 60)

        if stats['errors']:
            logger.error("=" * 60)
            logger.error("ERRORS ENCOUNTERED DURING IMPORT:")
            logger.error("=" * 60)
            for idx, err in enumerate(stats['errors'], 1):
                logger.error(f"Error #{idx}: {err}")
            logger.error("=" * 60)
            logger.error(f"Total errors: {len(stats['errors'])}")
            logger.error("=" * 60)

            # Also log to console (first 10)
            logger.info("\nErrors (first 10):")
            for err in stats['errors'][:10]:
                logger.info(f"  - {err}")
            if len(stats['errors']) > 10:
                logger.info(f"  ... and {len(stats['errors']) - 10} more (see log file for all errors)")

        logger.info("=" * 60)
        logger.info("MRP Production Import Script Completed")
        logger.info(f"Full log saved to: {log_file_path}")
        logger.info("=" * 60)


def main():
    """
    CLI entry point.

    Configuration is taken from the central config.py in the BOM directory.
    """
    # Load configuration from central config.py in BOM directory
    import os
    from pathlib import Path
    
    # Get the BOM directory (parent of this script's directory)
    script_dir = Path(__file__).parent
    bom_dir = script_dir.parent
    config_path = bom_dir / 'config.py'
    
    # Add BOM directory to path to import config
    if str(bom_dir) not in sys.path:
        sys.path.insert(0, str(bom_dir))
    
    try:
        import config  # type: ignore

        ODOO_URL = getattr(config, "ODOO_URL", "http://localhost:8069")
        ODOO_DB = getattr(config, "ODOO_DB", "your_database_name")
        ODOO_USERNAME = getattr(config, "ODOO_USERNAME", "admin")
        ODOO_PASSWORD = getattr(config, "ODOO_PASSWORD", "admin")
        EXCEL_FILE = getattr(config, "MRP_EXCEL_FILE", "output.xlsx")
        SHEET_NAME = getattr(config, "MRP_SHEET_NAME", None)
        DRY_RUN = getattr(config, "MRP_DRY_RUN", True)
    except ImportError:
        logger.error(f"Failed to import config from {config_path}")
        logger.error("Please ensure config.py exists in the BOM directory")
        # Fallback defaults
        ODOO_URL = "http://localhost:8099"
        ODOO_DB = "lingjack-test"
        ODOO_USERNAME = "admin"
        ODOO_PASSWORD = "admin"
        EXCEL_FILE = "output.xlsx"
        SHEET_NAME = "Manufacturing Order (mrp.produc"
        DRY_RUN = True

    # CLI overrides
    if len(sys.argv) > 1 and sys.argv[1] not in ("--execute", "--dry-run"):
        EXCEL_FILE = sys.argv[1]

    if "--execute" in sys.argv:
        DRY_RUN = False
    if "--dry-run" in sys.argv:
        DRY_RUN = True

    # Resolve Excel path relative to this script directory if not absolute,
    # so it works both when run directly and via run_all_imports.py.
    excel_path = Path(EXCEL_FILE)
    if not excel_path.is_absolute():
        excel_path = script_dir / excel_path
        EXCEL_FILE = str(excel_path)

    logger.info("Excel file: %s", EXCEL_FILE)
    logger.info("Sheet name: %s", SHEET_NAME)
    logger.info("Dry run: %s", DRY_RUN)

    importer = OdooMRPProductionImporter(ODOO_URL, ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD)
    importer.import_mrp_productions(
        excel_path=EXCEL_FILE,
        sheet_name=SHEET_NAME,
        dry_run=DRY_RUN,
    )


if __name__ == "__main__":
    main()

