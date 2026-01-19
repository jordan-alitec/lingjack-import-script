#!/usr/bin/env python3
"""
Odoo 18 Sale Work Order Import Script

Reads sale work orders from an Excel file (output.xlsx) and
creates sale.work.order and sale.work.order.line records via XML-RPC.

Each row in the Excel creates a separate SWO with one line.
"""

import sys
import logging
from typing import Dict, List, Optional
from datetime import datetime, date
import os

import xmlrpc.client
from openpyxl import load_workbook

# Set up logging to both console and file
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Remove existing handlers to avoid duplicates
logger.handlers = []

# Console handler
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)

# File handler for errors and detailed logs
log_file_path = os.path.join(os.path.dirname(__file__), 'import_swo_errors.log')
file_handler = logging.FileHandler(log_file_path, mode='a', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_formatter = logging.Formatter(
    "%(asctime)s - %(levelname)s - %(name)s - %(message)s",
    datefmt='%Y-%m-%d %H:%M:%S'
)
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

logger.info("=" * 60)
logger.info("SWO Import Script Started")
logger.info(f"Log file: {log_file_path}")
logger.info("=" * 60)


class OdooSaleWorkOrderImporter:
    """Import Sale Work Orders from Excel to Odoo 18"""

    def __init__(self, url: str, db: str, username: str, password: str):
        """
        Initialize Odoo connection

        Args:
            url: Odoo server URL (e.g., 'http://localhost:8069')
            db: Database name
            username: Odoo username
            password: Odoo password
        """
        self.url = url
        self.db = db
        self.username = username
        self.password = password

        # Authenticate
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
        self.uid = common.authenticate(db, username, password, {})
        if not self.uid:
            raise Exception(
                f"Authentication failed for user '{username}' "
                f"on database '{db}'."
            )

        self.models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")
        logger.info("Connected to Odoo DB '%s' as '%s'", db, username)

    # ---------------- Generic helpers ------------------

    def _search(self, model: str, domain: list, limit: int = 1) -> List[int]:
        """Search records in Odoo"""
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'search',
            [domain],
            {'limit': limit}
        )

    def _create(self, model: str, vals: dict) -> int:
        """Create a record in Odoo"""
        # Filter out None values - XML-RPC cannot marshal None
        filtered_vals = {k: v for k, v in vals.items() if v is not None}
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'create',
            [filtered_vals]
        )

    # ---------------- Lookups ------------------

    def find_product_by_default_code(self, default_code: str) -> Optional[int]:
        """Find product.product by default_code"""
        if not default_code:
            return None
        default_code = str(default_code).strip()
        if not default_code:
            return None
        product_ids = self._search('product.product', [('default_code', '=', default_code)], limit=1)
        return product_ids[0] if product_ids else None

    def find_customer_by_name(self, name: str) -> Optional[int]:
        """Find res.partner (customer) by name"""
        if not name:
            return None
        name = str(name).strip()
        if not name:
            return None
        partner_ids = self._search('res.partner', [('name', '=', name)], limit=1)
        return partner_ids[0] if partner_ids else None

    def find_user_by_name(self, name: str) -> Optional[int]:
        """Find res.users by name"""
        if not name:
            return None
        name = str(name).strip()
        if not name:
            return None
        user_ids = self._search('res.users', [('name', '=ilike', name)], limit=1)
        return user_ids[0] if user_ids else None

    def find_or_create_user_by_name(self, name: str, dry_run: bool = False) -> Optional[int]:
        """
        Find res.users by name, or create a new internal user if not found.
        Created users are archived (active=False).
        
        Args:
            name: User name
            dry_run: If True, only log without creating
            
        Returns:
            User ID or None
        """
        if not name:
            return None
        name = str(name).strip()
        if not name:
            return None
        
        # Try to find existing user
        user_id = self.find_user_by_name(name)
        if user_id:
            return user_id
        
        # User not found - create new one
        if dry_run:
            logger.info(f"[DRY RUN] Would create archived internal user: {name}")
            return None
        
        try:
            # Generate login from name (lowercase, replace spaces with dots)
            login = name.lower().replace(' ', '.').replace("'", "").replace("-", ".")
            # Ensure login is unique
            existing_logins = self._search('res.users', [('login', '=', login)])
            if existing_logins:
                # Add a number suffix
                counter = 1
                while self._search('res.users', [('login', '=', f"{login}{counter}")]):
                    counter += 1
                login = f"{login}{counter}"
            
            # Get internal user group (base.group_user)
            # Try to find by external ID first, then by name
            try:
                # Try to get by XML ID
                internal_group_id = self.models.execute_kw(
                    self.db, self.uid, self.password,
                    'ir.model.data', 'xmlid_to_res_id',
                    ['base.group_user']
                )
            except:
                # Fallback: search by name
                internal_group_ids = self._search('res.groups', [
                    ('category_id.name', '=', 'User Types'),
                    ('name', '=', 'Internal User')
                ], limit=1)
                internal_group_id = internal_group_ids[0] if internal_group_ids else None
            
            user_vals = {
                'name': name,
                'login': login,
                'active': False,  # Archive the user
            }
            
            if internal_group_id:
                user_vals['groups_id'] = [(4, internal_group_id)]  # Add to internal user group
            
            user_id = self._create('res.users', user_vals)
            logger.info(f"Created archived internal user '{name}' (ID: {user_id}, login: {login})")
            return user_id
            
        except Exception as e:
            logger.error(f"Error creating user '{name}': {e}", exc_info=True)
            return None

    def find_uom_by_name(self, name: str) -> Optional[int]:
        """Find uom.uom by name (default: Units)"""
        if not name:
            name = 'Units'
        name = str(name).strip()
        if not name:
            name = 'Units'
        uom_ids = self._search('uom.uom', [('name', '=', name)], limit=1)
        if uom_ids:
            return uom_ids[0]
        # Fallback to Units
        uom_ids = self._search('uom.uom', [('name', '=', 'Units')], limit=1)
        return uom_ids[0] if uom_ids else None

    # ---------------- Parsing helpers ------------------

    def _to_datetime_str(self, value) -> Optional[str]:
        """Convert Excel date/datetime to ISO string (YYYY-MM-DD HH:MM:SS)"""
        if not value:
            return None
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d 00:00:00')
        if isinstance(value, str):
            # Try to parse common date formats
            value = value.strip()
            try:
                # Try DD/MM/YYYY HH:MM:SS format (from Excel)
                dt = datetime.strptime(value, '%d/%m/%Y %H:%M:%S')
                return dt.strftime('%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    # Try DD/MM/YYYY format
                    dt = datetime.strptime(value, '%d/%m/%Y')
                    return dt.strftime('%Y-%m-%d 00:00:00')
                except ValueError:
                    try:
                        # Try YYYY-MM-DD HH:MM:SS format
                        dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                        return dt.strftime('%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        try:
                            # Try YYYY-MM-DD format
                            dt = datetime.strptime(value, '%Y-%m-%d')
                            return dt.strftime('%Y-%m-%d 00:00:00')
                        except ValueError:
                            # Return as-is if already in correct format or log warning
                            logger.warning(f"Could not parse date format: {value}")
                            return value
        return None

    def _map_pwo_status_to_state(self, status: str) -> str:
        """Map PWO Status to sale.work.order state"""
        if not status:
            return 'draft'
        status = str(status).strip().lower()
        mapping = {
            'in progress': 'in_production',
            'planning': 'confirmed',
            'completed': 'produced',
            'cancelled': 'cancelled',
        }
        return mapping.get(status, 'draft')

    def _combine_remarks(self, cs_remarks: Optional[str], prod_remarks: Optional[str]) -> Optional[str]:
        """Combine CS and Production remarks with sections"""
        parts = []
        if cs_remarks and str(cs_remarks).strip():
            parts.append("Remarks (CS)\n" + str(cs_remarks).strip())
        if prod_remarks and str(prod_remarks).strip():
            parts.append("Remarks (Production)\n" + str(prod_remarks).strip())
        return "\n\n".join(parts) if parts else None

    # ---------------- Excel parsing ------------------

    def parse_excel(self, excel_path: str, sheet_name: Optional[str] = None, header_row: int = 1) -> List[Dict]:
        """
        Parse Excel file and return list of SWO records

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name (default: active sheet)
            header_row: Row number of header (1-based)

        Returns:
            List of dicts with SWO data
        """
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        start_row = header_row + 1
        max_row = ws.max_row

        records = []

        for row_idx in range(start_row, max_row + 1):
            row_cells = [cell.value for cell in ws[row_idx]]

            # Extract values by column index (0-based)
            # Column A (0): S/N - ignore
            # Column B (1): SWO Number
            # Column C (2): SO Number
            # Column D (3): Contact Code - ignore
            # Column E (4): Contact Name
            # Column F (5): SWO Issue Date
            # Column G (6): SWO Issue By
            # Column H (7): Completion Date
            # Column I (8): PWO Number(s)
            # Column J (9): Item Code
            # Column K (10): Item Description - ignore
            # Column L (11): PWO Status
            # Column M (12): SWO Committed Qty
            # Column N (13): PWO Finished Qty
            # Column O (14): Outstanding Qty - ignore
            # Column P (15): Remarks (CS)
            # Column Q (16): Remarks (Production)

            if len(row_cells) < 17:
                continue

            swo_number = row_cells[1] if len(row_cells) > 1 else None
            so_number = row_cells[2] if len(row_cells) > 2 else None
            contact_name = row_cells[4] if len(row_cells) > 4 else None
            swo_issue_date = row_cells[5] if len(row_cells) > 5 else None
            swo_issue_by = row_cells[6] if len(row_cells) > 6 else None
            completion_date = row_cells[7] if len(row_cells) > 7 else None
            pwo_number = row_cells[8] if len(row_cells) > 8 else None
            item_code = row_cells[9] if len(row_cells) > 9 else None
            pwo_status = row_cells[11] if len(row_cells) > 11 else None
            committed_qty = row_cells[12] if len(row_cells) > 12 else None
            finished_qty = row_cells[13] if len(row_cells) > 13 else None
            cs_remarks = row_cells[15] if len(row_cells) > 15 else None
            prod_remarks = row_cells[16] if len(row_cells) > 16 else None

            # Skip rows without essential data
            if not swo_number or not item_code:
                continue

            record = {
                'swo_number': str(swo_number).strip() if swo_number else None,
                'so_number': str(so_number).strip() if so_number else None,
                'contact_name': str(contact_name).strip() if contact_name else None,
                'swo_issue_date': swo_issue_date,
                'swo_issue_by': str(swo_issue_by).strip() if swo_issue_by else None,
                'completion_date': completion_date,
                'pwo_number': str(pwo_number).strip() if pwo_number else None,
                'item_code': str(item_code).strip() if item_code else None,
                'pwo_status': str(pwo_status).strip() if pwo_status else None,
                'committed_qty': float(committed_qty) if committed_qty else 0.0,
                'finished_qty': float(finished_qty) if finished_qty else 0.0,
                'cs_remarks': str(cs_remarks).strip() if cs_remarks else None,
                'prod_remarks': str(prod_remarks).strip() if prod_remarks else None,
                'row_index': row_idx,
            }
            records.append(record)

        wb.close()
        logger.info("Parsed %d SWO records from Excel", len(records))
        return records

    # ---------------- Import Logic ------------------

    def import_sale_work_orders(self, excel_path: str, sheet_name: Optional[str] = None, dry_run: bool = True):
        """
        Import sale work orders from Excel

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name (default: active sheet)
            dry_run: If True, only log operations without creating records
        """
        records = self.parse_excel(excel_path, sheet_name)

        if not records:
            logger.warning("No records found in Excel file")
            return

        stats = {
            'total_rows': len(records),
            'created_swo': 0,
            'created_lines': 0,
            'errors': [],
        }

        if dry_run:
            logger.info("DRY RUN MODE - No records will be created in Odoo")

        # Note: sale_order_id is required by model, but the model's create() method
        # will automatically create a dummy sale.order if old_so_number is provided
        # We just store old_so_number - no need to create sale orders here

        for rec in records:
            try:
                # Find product by default_code
                product_id = self.find_product_by_default_code(rec['item_code'])
                if not product_id:
                    error_msg = f"Row {rec['row_index']}: Product not found for Item Code '{rec['item_code']}'"
                    logger.error(error_msg)
                    logger.debug(f"Row {rec['row_index']} full record: {rec}")
                    stats['errors'].append(error_msg)
                    continue

                # Find customer by name
                customer_id = None
                if rec['contact_name']:
                    customer_id = self.find_customer_by_name(rec['contact_name'])
                    if not customer_id:
                        logger.warning(
                            f"Row {rec['row_index']}: Customer not found for '{rec['contact_name']}'"
                        )

                # Find or create user by name (create archived internal user if not found)
                user_id = None
                if rec['swo_issue_by']:
                    user_id = self.find_or_create_user_by_name(rec['swo_issue_by'], dry_run=dry_run)
                    if not user_id and not dry_run:
                        logger.warning(
                            f"Row {rec['row_index']}: Failed to find or create user for '{rec['swo_issue_by']}'"
                        )

                # Get product UOM
                product_data = self.models.execute_kw(
                    self.db, self.uid, self.password,
                    'product.product', 'read',
                    [[product_id]],
                    {'fields': ['uom_id']}
                )
                uom_id = product_data[0]['uom_id'][0] if product_data and product_data[0].get('uom_id') else None

                # Prepare SWO values - store old_so_number as char field (not linking to sale.order)
                # The model's create() method will automatically handle the required sale_order_id
                # by creating a dummy sale.order when old_so_number is provided
                print(f"\n\n\n{customer_id}")
                swo_vals = {
                    'customer_id': customer_id,
                    'old_swo_number': rec['swo_number'],
                    'name': rec['swo_number'],
                    'old_so_number': rec['so_number'],  # Store SO number as char field only
                    'old_pwo_number': rec['pwo_number'],
                    'request_date': self._to_datetime_str(rec['swo_issue_date']),
                    'completion_date': self._to_datetime_str(rec['completion_date']),
                    'state': self._map_pwo_status_to_state(rec['pwo_status']),
                    'remarks': self._combine_remarks(rec['cs_remarks'], rec['prod_remarks']),
                    
                }
                # Don't set sale_order_id - let the model's create() method handle it

                # Add cs_in_charge_id if user found
                if user_id:
                    swo_vals['cs_in_charge_id'] = user_id

                if not dry_run:
                    # Create SWO
                    swo_id = self._create('sale.work.order', swo_vals)
                    stats['created_swo'] += 1
                    logger.info(
                        f"Row {rec['row_index']}: Created SWO '{rec['swo_number']}' (ID: {swo_id})"
                    )

                    # Create SWO line
                    line_vals = {
                        'work_order_id': swo_id,
                        'product_id': product_id,
                        'product_qty': rec['committed_qty'],
                        'product_uom_id': uom_id,
                        'qty_produced': rec['finished_qty'],
                        'state': self._map_pwo_status_to_state(rec['pwo_status']),
                        'remarks': self._combine_remarks(rec['cs_remarks'], rec['prod_remarks']),
                    }
                    line_id = self._create('sale.work.order.line', line_vals)
                    stats['created_lines'] += 1
                    logger.info(
                        f"Row {rec['row_index']}: Created SWO line (ID: {line_id}) for product '{rec['item_code']}'"
                    )
                else:
                    logger.info(
                        f"[DRY RUN] Row {rec['row_index']}: Would create SWO '{rec['swo_number']}' "
                        f"with line for product '{rec['item_code']}'"
                    )
                    stats['created_swo'] += 1
                    stats['created_lines'] += 1

            except Exception as e:
                error_msg = f"Row {rec['row_index']}: Error processing SWO '{rec.get('swo_number', 'N/A')}': {e}"
                # Log full exception details to file
                logger.error(error_msg, exc_info=True)
                # Also log to console without traceback
                logger.error(error_msg)
                stats['errors'].append(error_msg)

        # Summary
        logger.info("=" * 60)
        logger.info("Import Summary:")
        logger.info("  Total rows processed: %d", stats['total_rows'])
        logger.info("  SWOs created: %d", stats['created_swo'])
        logger.info("  Lines created: %d", stats['created_lines'])
        logger.info("  Errors: %d", len(stats['errors']))
        logger.info("=" * 60)

        if stats['errors']:
            logger.error("=" * 60)
            logger.error("ERRORS ENCOUNTERED DURING IMPORT:")
            logger.error("=" * 60)
            for idx, err in enumerate(stats['errors'], 1):
                logger.error(f"Error #{idx}: {err}")
            logger.error("=" * 60)
            logger.error(f"Total errors: {len(stats['errors'])}")
            logger.error("=" * 60)
            
            # Also log to console (first 10)
            logger.info("\nErrors (first 10):")
            for err in stats['errors'][:10]:
                logger.info(f"  - {err}")
            if len(stats['errors']) > 10:
                logger.info(f"  ... and {len(stats['errors']) - 10} more (see log file for all errors)")
        
        logger.info("=" * 60)
        logger.info("SWO Import Script Completed")
        logger.info(f"Full log saved to: {log_file_path}")
        logger.info("=" * 60)


def main():
    """
    CLI entry point.

    Configuration is taken from the central config.py in the BOM directory.
    """
    # Load configuration from central config.py in BOM directory
    import os
    from pathlib import Path
    
    # Get the BOM directory (parent of this script's directory)
    script_dir = Path(__file__).parent
    bom_dir = script_dir.parent
    config_path = bom_dir / 'config.py'
    
    # Add BOM directory to path to import config
    if str(bom_dir) not in sys.path:
        sys.path.insert(0, str(bom_dir))
    
    try:
        import config  # type: ignore

        ODOO_URL = getattr(config, "ODOO_URL", "http://localhost:8069")
        ODOO_DB = getattr(config, "ODOO_DB", "your_database_name")
        ODOO_USERNAME = getattr(config, "ODOO_USERNAME", "admin")
        ODOO_PASSWORD = getattr(config, "ODOO_PASSWORD", "admin")
        EXCEL_FILE = getattr(config, "SWO_EXCEL_FILE", "output.xlsx")
        SHEET_NAME = getattr(config, "SWO_SHEET_NAME", "Outstanding SWO Listing")
        DRY_RUN = getattr(config, "SWO_DRY_RUN", True)
    except ImportError:
        logger.error(f"Failed to import config from {config_path}")
        logger.error("Please ensure config.py exists in the BOM directory")
        # Fallback defaults
        ODOO_URL = "http://localhost:8099"
        ODOO_DB = "lingjack-test"
        ODOO_USERNAME = "admin"
        ODOO_PASSWORD = "admin"
        EXCEL_FILE = "output.xlsx"
        SHEET_NAME = "Outstanding SWO Listing"
        DRY_RUN = False

    # CLI overrides
    if len(sys.argv) > 1 and sys.argv[1] not in ("--execute", "--dry-run"):
        EXCEL_FILE = sys.argv[1]

    if "--execute" in sys.argv:
        DRY_RUN = False
    if "--dry-run" in sys.argv:
        DRY_RUN = True

    logger.info("Excel file: %s", EXCEL_FILE)
    logger.info("Sheet name: %s", SHEET_NAME)
    logger.info("Dry run: %s", DRY_RUN)

    importer = OdooSaleWorkOrderImporter(ODOO_URL, ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD)
    importer.import_sale_work_orders(
        excel_path=EXCEL_FILE,
        sheet_name=SHEET_NAME,
        dry_run=DRY_RUN,
    )


if __name__ == "__main__":
    main()

