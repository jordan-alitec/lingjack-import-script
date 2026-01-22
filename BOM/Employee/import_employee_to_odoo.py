#!/usr/bin/env python3
"""
Odoo 18 Employee Import Script

- Reads employees from an Excel file (employee.xlsx)
- Creates/updates hr.employee records
- Handles parent/manager (parent_id) in a second pass

Columns are mapped by header names in the first row of each sheet.
You can optionally override default mappings in a config_employee.py file
in the same directory.
"""

import xmlrpc.client
from openpyxl import load_workbook
import sys
from typing import List, Dict, Optional
import logging
import socket
from urllib.parse import urlparse
from datetime import date, datetime
from pathlib import Path

# ---------------------------------------------------------
# CONFIGURATION (can be overridden by config_employee.py)
# ---------------------------------------------------------

# Load configuration from central config.py in BOM directory

# Get the BOM directory (parent of this script's directory)
script_dir = Path(__file__).parent
bom_dir = script_dir.parent
config_path = bom_dir / 'config.py'

# Add BOM directory to path to import config
if str(bom_dir) not in sys.path:
    sys.path.insert(0, str(bom_dir))

# Try to load configuration from central config.py
try:
    import config
    ODOO_URL = getattr(config, 'ODOO_URL', 'http://localhost:8099')
    ODOO_DB = getattr(config, 'ODOO_DB', 'lingjack-main')
    ODOO_USERNAME = getattr(config, 'ODOO_USERNAME', 'admin')
    ODOO_PASSWORD = getattr(config, 'ODOO_PASSWORD', 'admin')
    EMPLOYEE_EXCEL_FILE = getattr(config, 'EMPLOYEE_EXCEL_FILE', 'employee.xlsx')
    DRY_RUN = getattr(config, 'EMPLOYEE_DRY_RUN', False)
    # Get header mapping from config
    HEADER_MAPPING = getattr(config, 'EMPLOYEE_HEADER_MAPPING', {})
except ImportError:
    print(f"Warning: Failed to import config from {config_path}")
    print("Please ensure config.py exists in the BOM directory")
    # Fallback defaults
    # ODOO_URL = 'https://lingjack-data-migration-script-27115365.dev.odoo.com'
    # ODOO_DB = 'lingjack-data-migration-script-27115365'
    ODOO_URL = 'http://localhost:8099'
    ODOO_DB = 'lingjack-main'
    ODOO_USERNAME = 'admin'
    ODOO_PASSWORD = 'admin'
    EMPLOYEE_EXCEL_FILE = 'employee.xlsx'
    DRY_RUN = False

    # Default header mapping – tuned to your screenshot of employee.xlsx.
    # If any header text differs slightly, update the value to match
    # EXACTLY the text in row 1 of the sheet.
    HEADER_MAPPING = {
        # Identification / core info
        'employee_code': None,                  # no dedicated code column
        'employee_name': 'Employee Name',
        # 'employee_type': 'Employee Type',
        'department_name': 'Department',
        'job_title': 'Job Position',
        'manager_code': None,
        'manager_name': 'Manager',

        # Work contacts
        'work_email': 'Work Email',
        'work_phone': 'Work Phone',

        # Private contacts & address
        'private_email': 'Private Email',
        'private_phone': 'Private Phone',
        'private_street': 'Private Street',
        'private_street2': 'Private Street2',
        'private_zip': 'Private Zip',
        'private_city': 'Private City',
        'private_country': 'Private Country',

        # Identity / personal data
        'nationality': 'Nationality',
        'identification_no': 'Identification No',
        'ssn_no': 'SSN No',
        'passport_no': 'Passport No',
        'gender': 'Gender',
        'birthday': 'Date of Birth',
        'place_of_birth': 'Place of Birth',
        'country_of_birth': 'Country of Birth',
        'study_field': 'Field of Study',
        'visa_no': 'Visa No',
        'visa_expire': 'Visa Expiration Date',
        'work_permit_no': 'Work Permit No',
        'work_permit_expiration': 'Work Permit Expiration Date',
        'marital_status': 'Marital Status',
        'children': 'Number of Children',

        # HR / cost / badge
        'hourly_cost': 'Hourly Cost',
        'badge_id': 'Badge ID',

        # Emergency contact
        'emergency_contact_name': 'Contact Name',
        'emergency_contact_phone': 'Contact Phone',
    }

# ---------------------------------------------------------
# LOGGING
# ---------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------
# Odoo Connection / Helper
# ---------------------------------------------------------

class OdooEmployeeImporter:
    """Import Employee data from Excel to Odoo 18 (hr.employee)"""

    def __init__(self, url: str, db: str, username: str, password: str):
        self.url = url
        self.db = db
        self.username = username
        self.password = password

        self._test_connection(url)

        try:
            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common')
            self.uid = common.authenticate(db, username, password, {})
        except ConnectionRefusedError:
            raise ConnectionError(
                f"Cannot connect to Odoo server at {url}.\n"
                f"Please check:\n"
                f"  1. Is Odoo server running?\n"
                f"  2. Is the URL correct? (current: {url})\n"
                f"  3. Is the port correct?\n"
                f"  4. Is there a firewall blocking the connection?"
            )
        except Exception as e:
            raise ConnectionError(
                f"Failed to connect to Odoo: {str(e)}\n"
                f"URL: {url}\n"
                f"Please verify the server is running and accessible."
            )

        if not self.uid:
            raise Exception(
                f"Authentication failed for user '{username}'.\n"
                f"Check username/password/db and access rights."
            )

        self.models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object')
        logger.info(f"Successfully connected to Odoo database: {db}")

    def _test_connection(self, url: str):
        """Light socket test before XML-RPC"""
        try:
            parsed = urlparse(url)
            host = parsed.hostname or 'localhost'
            port = parsed.port or (8069 if parsed.scheme == 'http' else 443)

            logger.info(f"Testing connection to {host}:{port}...")
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(5)
            result = sock.connect_ex((host, port))
            sock.close()

            if result != 0:
                raise ConnectionError(
                    f"Cannot reach Odoo server at {host}:{port}.\n"
                    f"Ensure server is running and accessible."
                )
        except Exception as e:
            logger.warning(f"Connection test warning: {e}")

    # ---------------- Generic helpers ------------------

    def _search(self, model: str, domain: list, limit: int = 1) -> List[int]:
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'search',
            [domain],
            {'limit': limit}
        )

    def _create(self, model: str, vals: dict) -> int:
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'create',
            [vals]
        )

    def _write(self, model: str, ids: List[int], vals: dict) -> bool:
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'write',
            [ids, vals]
        )

    # ---------------- Lookups ------------------

    def find_department_by_name(self, name: str, temp_company_id: Optional[str] = None) -> Optional[int]:
        """
        Find department by name and optionally by temp_company_id.
        
        Args:
            name: Department name
            temp_company_id: Optional temp company ID to match
            
        Returns:
            Department ID or None
        """
        if not name:
            return None
        name = str(name).strip()
        if not name:
            return None

        # Search by name and temp_company_id if provided
        domain = [('name', '=', name)]
        if temp_company_id:
            domain.append(('temp_company_id', '=', temp_company_id))

        dept_ids = self._search('hr.department', domain, limit=1)
        return dept_ids[0] if dept_ids else None

    def find_or_create_department(self, name: str, temp_company_id: Optional[str] = None) -> Optional[int]:
        """
        Find or create department by name and temp_company_id.
        
        Args:
            name: Department name
            temp_company_id: Optional temp company ID (sheet name) to store
            
        Returns:
            Department ID or None
        """
        if not name:
            return None
        name = str(name).strip()
        if not name:
            return None

        # Search for existing department with same name and temp_company_id
        dept_id = self.find_department_by_name(name, temp_company_id)
        if dept_id:
            return dept_id

        # Create new department with temp_company_id
        try:
            dept_vals = {'name': name}
            if temp_company_id:
                dept_vals['temp_company_id'] = temp_company_id
            dept_id = self._create('hr.department', dept_vals)
            logger.info(f"Created department: {name} (ID: {dept_id}, temp_company_id: {temp_company_id})")
            return dept_id
        except Exception as e:
            logger.error(f"Error creating department '{name}': {e}")
            return None

    def find_country_by_name(self, name: str) -> Optional[int]:
        """Find country by exact name (used for nationality and addresses)."""
        if not name:
            return None
        name = str(name).strip()
        if not name:
            return None
        country_ids = self._search('res.country', [('name', '=', name)], limit=1)
        return country_ids[0] if country_ids else None

    def find_employee_by_code(self, code: str, company_id: Optional[int] = None, temp_company_id: Optional[str] = None) -> Optional[int]:
        """
        Lookup by identification_id (assuming employee_code -> identification_id).
        
        Args:
            code: Employee identification code
            company_id: Optional company ID to filter by
            temp_company_id: Optional temp company ID (sheet name) to filter by
            
        Returns:
            Employee ID or None
        """
        if not code:
            return None
        code = str(code).strip()
        if not code:
            return None
        domain = [('identification_id', '=', code)]
        if company_id:
            domain.append(('company_id', '=', company_id))
        if temp_company_id:
            domain.append(('temp_company_id', '=', temp_company_id))
        emp_ids = self._search('hr.employee', domain, limit=1)
        return emp_ids[0] if emp_ids else None

    def find_employee_by_email(self, email: str, company_id: Optional[int] = None, temp_company_id: Optional[str] = None) -> Optional[int]:
        """
        Find employee by work email.
        
        Args:
            email: Work email address
            company_id: Optional company ID to filter by
            temp_company_id: Optional temp company ID (sheet name) to filter by
            
        Returns:
            Employee ID or None
        """
        if not email:
            return None
        email = str(email).strip()
        if not email:
            return None
        domain = [('work_email', '=', email)]
        if company_id:
            domain.append(('company_id', '=', company_id))
        if temp_company_id:
            domain.append(('temp_company_id', '=', temp_company_id))
        emp_ids = self._search('hr.employee', domain, limit=1)
        return emp_ids[0] if emp_ids else None

    def find_employee_by_name(self, name: str, company_id: Optional[int] = None, temp_company_id: Optional[str] = None) -> Optional[int]:
        """
        Find employee by name.
        
        Args:
            name: Employee name
            company_id: Optional company ID to filter by
            temp_company_id: Optional temp company ID (sheet name) to filter by
            
        Returns:
            Employee ID or None
        """
        if not name:
            return None
        name = str(name).strip()
        if not name:
            return None
        domain = [('name', '=', name)]
        if company_id:
            domain.append(('company_id', '=', company_id))
        if temp_company_id:
            domain.append(('temp_company_id', '=', temp_company_id))
        emp_ids = self._search('hr.employee', domain, limit=1)
        return emp_ids[0] if emp_ids else None

    # ---------------- Parsing Excel ------------------

    def _build_header_index(self, header_row: List[Optional[str]]) -> Dict[str, int]:
        """Return dict: normalized header text -> column index."""
        index = {}
        for i, val in enumerate(header_row):
            if val is None:
                continue
            text = str(val).strip()
            if not text:
                continue
            index[text.lower()] = i
        return index

    def _get_by_header(self, row: tuple, header_index: Dict[str, int], header_name: str) -> Optional[str]:
        if not header_name:
            return None
        col_idx = header_index.get(header_name.lower())
        if col_idx is None or col_idx >= len(row):
            return None
        value = row[col_idx]
        return value

    def parse_workbook(self, excel_path: str) -> List[Dict]:
        """
        Parse all sheets in the workbook and return a list of employee dicts.

        Each dict contains:
        {
            'employee_code': str or None,
            'employee_name': str,
            'work_email': str or None,
            'mobile_phone': str or None,
            'job_title': str or None,
            'department_name': str or None,
            'manager_code': str or None,
            'manager_name': str or None,
            'sheet_title': str,
            'row_index': int,
        }
        """
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        employees: List[Dict] = []

        for ws in wb.worksheets:
            # Some sheets may be empty (no rows) or non-standard; safely skip them
            if ws.max_row is None or ws.max_row < 1 or ws.title == 'lookup':
                logger.info(f"Sheet '{ws.title}' has no rows, skipping.")
                continue

            try:
                header_row = ws[1]
            except IndexError:
                logger.info(f"Sheet '{ws.title}' header row not found, skipping.")
                continue

            header = [cell.value for cell in header_row]
            header_index = self._build_header_index(header)
            logger.info(f"Sheet '{ws.title}' columns: {header}")

            # Helpers to get header name for each internal key
            def h(key: str) -> Optional[str]:
                return HEADER_MAPPING.get(key)

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                emp_name_val = self._get_by_header(row, header_index, h('employee_name'))
                if not emp_name_val or not str(emp_name_val).strip():

                    continue

                emp_code_val = self._get_by_header(row, header_index, h('employee_code'))
                # emp_type_val = self._get_by_header(row, header_index, h('employee_type'))
                work_email_val = self._get_by_header(row, header_index, h('work_email'))
                work_phone_val = self._get_by_header(row, header_index, h('work_phone'))
                private_email_val = self._get_by_header(row, header_index, h('private_email'))
                private_phone_val = self._get_by_header(row, header_index, h('private_phone'))
                private_street_val = self._get_by_header(row, header_index, h('private_street'))
                private_street2_val = self._get_by_header(row, header_index, h('private_street2'))
                private_zip_val = self._get_by_header(row, header_index, h('private_zip'))
                private_city_val = self._get_by_header(row, header_index, h('private_city'))
                private_country_val = self._get_by_header(row, header_index, h('private_country'))

                nationality_val = self._get_by_header(row, header_index, h('nationality'))
                identification_val = self._get_by_header(row, header_index, h('identification_no'))
                ssn_val = self._get_by_header(row, header_index, h('ssn_no'))
                passport_val = self._get_by_header(row, header_index, h('passport_no'))
                gender_val = self._get_by_header(row, header_index, h('gender'))
                birthday_val = self._get_by_header(row, header_index, h('birthday'))
                pob_val = self._get_by_header(row, header_index, h('place_of_birth'))
                cob_val = self._get_by_header(row, header_index, h('country_of_birth'))
                study_field_val = self._get_by_header(row, header_index, h('study_field'))
                visa_no_val = self._get_by_header(row, header_index, h('visa_no'))
                visa_exp_val = self._get_by_header(row, header_index, h('visa_expire'))
                wp_no_val = self._get_by_header(row, header_index, h('work_permit_no'))
                wp_exp_val = self._get_by_header(row, header_index, h('work_permit_expiration'))
                marital_val = self._get_by_header(row, header_index, h('marital_status'))
                children_val = self._get_by_header(row, header_index, h('children'))

                hourly_cost_val = self._get_by_header(row, header_index, h('hourly_cost'))
                badge_id_val = self._get_by_header(row, header_index, h('badge_id'))
                emergency_name_val = self._get_by_header(row, header_index, h('emergency_contact_name'))
                emergency_phone_val = self._get_by_header(row, header_index, h('emergency_contact_phone'))

                job_title_val = self._get_by_header(row, header_index, h('job_title'))
                dept_val = self._get_by_header(row, header_index, h('department_name'))
                mgr_code_val = self._get_by_header(row, header_index, h('manager_code'))
                mgr_name_val = self._get_by_header(row, header_index, h('manager_name'))

                rec = {
                    'employee_code': str(emp_code_val).strip() if emp_code_val else None,
                    'employee_name': str(emp_name_val).strip(),
                    # 'employee_type': str(emp_type_val).strip() if emp_type_val else None,
                    'work_email': str(work_email_val).strip() if work_email_val else None,
                    'work_phone': str(work_phone_val).strip() if work_phone_val else None,
                    'private_email': str(private_email_val).strip() if private_email_val else None,
                    'private_phone': str(private_phone_val).strip() if private_phone_val else None,
                    'private_street': str(private_street_val).strip() if private_street_val else None,
                    'private_street2': str(private_street2_val).strip() if private_street2_val else None,
                    'private_zip': str(private_zip_val).strip() if private_zip_val else None,
                    'private_city': str(private_city_val).strip() if private_city_val else None,
                    'private_country': str(private_country_val).strip() if private_country_val else None,
                    'nationality': str(nationality_val).strip() if nationality_val else None,
                    'identification_no': str(identification_val).strip() if identification_val else None,
                    'ssn_no': str(ssn_val).strip() if ssn_val else None,
                    'passport_no': str(passport_val).strip() if passport_val else None,
                    'gender': str(gender_val).strip() if gender_val else None,
                    'birthday': birthday_val,
                    'place_of_birth': str(pob_val).strip() if pob_val else None,
                    'country_of_birth': str(cob_val).strip() if cob_val else None,
                    'study_field': str(study_field_val).strip() if study_field_val else None,
                    'visa_no': str(visa_no_val).strip() if visa_no_val else None,
                    'visa_expire': visa_exp_val,
                    'work_permit_no': str(wp_no_val).strip() if wp_no_val else None,
                    'work_permit_expiration': wp_exp_val,
                    'marital_status': str(marital_val).strip() if marital_val else None,
                    'children': children_val,
                    'hourly_cost': hourly_cost_val,
                    'badge_id': str(badge_id_val).strip() if badge_id_val else None,
                    'emergency_contact_name': str(emergency_name_val).strip() if emergency_name_val else None,
                    'emergency_contact_phone': str(emergency_phone_val).strip() if emergency_phone_val else None,
                    'job_title': str(job_title_val).strip() if job_title_val else None,
                    'department_name': str(dept_val).strip() if dept_val else None,
                    'manager_code': str(mgr_code_val).strip() if mgr_code_val else None,
                    'manager_name': str(mgr_name_val).strip() if mgr_name_val else None,
                    'sheet_title': ws.title,
                    'row_index': row_idx,
                }
                employees.append(rec)

        wb.close()
        logger.info(f"Parsed {len(employees)} employees from workbook")
        return employees

    # ---------------- Import Logic ------------------

    def _build_employee_key(self, rec: Dict) -> str:
        """Build a stable key for mapping employees in this import."""
        # if rec.get('employee_code'):
        #     return f"code:{rec['employee_code']}"
        if rec.get('work_email'):
            return f"email:{rec['work_email']}company:{rec['sheet_title']}"
        return f"name:{rec['employee_name']}"

    def _build_manager_key(self, rec: Dict) -> Optional[str]:
        if rec.get('manager_code'):
            return f"code:{rec['manager_code']}"
        if rec.get('manager_name'):
            return f"name:{rec['manager_name']}"
        return None

    def _get_company_id_from_temp(self, temp_company_id: Optional[str]) -> Optional[int]:
        """
        Get company_id from temp_company_id (sheet name) using the company mapping.
        
        Args:
            temp_company_id: Sheet name (e.g., "LJE", "LJD", etc.)
            
        Returns:
            Company ID or None
        """
        if not temp_company_id:
            return None
        # Company mapping from import_employees method
        company_vals = {
            "LJE": 1,
            "LJD": 2,
            "LJ Fire": 10,
            "LJ Malaysia": 6,
            "LJDT": 11,
            'LJ Property':4,
            'LJ GreenTech':3,
            'PT Lingjack':1,

        }
        return company_vals.get(temp_company_id)

    def _find_employee_in_odoo(self, key: str, company_id: Optional[int] = None, temp_company_id: Optional[str] = None, allow_cross_company: bool = False, use_if_found_in_other_company: bool = False) -> Optional[int]:
        
        """
        Find employee in Odoo by key (code, email, or name) with company filtering.
        
        First searches within the specified company context. If not found and allow_cross_company
        is True, searches across all companies. This allows creating separate employee records
        for the same person in different companies.
        
        Args:
            key: Employee key in format "code:xxx", "email:xxx", or "name:xxx"
            company_id: Optional company ID to filter by (preferred search)
            temp_company_id: Optional temp company ID (sheet name) to filter by (preferred search)
            allow_cross_company: If True and not found in company context, search across all companies
            use_if_found_in_other_company: If True and employee found in another company, return that ID.
                                          If False, return None to allow creating new record for current company.
            
        Returns:
            Employee ID or None. 
            - If found in same company: returns the ID
            - If found in another company and use_if_found_in_other_company=True: returns that ID
            - If found in another company and use_if_found_in_other_company=False: returns None (to create new record)
        """
        value = key.split(':', 1)[1] if ':' in key else None
        if not value:
            return None
            
        # First, try to find within the specified company context (preferred)
        if key.startswith('code:'):
            emp_id = self.find_employee_by_code(value, company_id=company_id, temp_company_id=temp_company_id)
        elif key.startswith('email:'):
            emp_id = self.find_employee_by_email(value, company_id=company_id, temp_company_id=temp_company_id)
        elif key.startswith('name:'):
            emp_id = self.find_employee_by_name(value, company_id=company_id, temp_company_id=temp_company_id)
        else:
            return None
            
        # If found in the same company, return it
        if emp_id:
            return emp_id
            
        # If not found and allow_cross_company is True, check if employee exists in another company
        if allow_cross_company:
            if key.startswith('code:'):
                emp_id_other = self.find_employee_by_code(value)
            elif key.startswith('email:'):
                emp_id_other = self.find_employee_by_email(value)
            elif key.startswith('name:'):
                emp_id_other = self.find_employee_by_name(value)
            else:
                return None
                
            if emp_id_other:
                if use_if_found_in_other_company:
                    # Use the employee from another company (e.g., for managers)
                    logger.info(
                        f"Employee '{value}' found in another company (ID: {emp_id_other}). "
                        f"Using existing record for company {temp_company_id or company_id}."
                    )
                    return emp_id_other
                else:
                    # Employee exists in another company - return None to allow creating new record for current company
                    logger.info(
                        f"Employee '{value}' found in another company (ID: {emp_id_other}). "
                        f"Will create new employee record for company {temp_company_id or company_id}."
                    )
                    return None
                
        return None

    def _to_date_str(self, value) -> Optional[str]:
        """
        Normalize Excel/JSON date values to ISO string (YYYY-MM-DD).
        Skips integers (Excel serials) that we cannot safely convert without
        workbook date mode information.
        """
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, str):
            return value  # assume already a valid date string
        # For now, ignore pure numeric values to avoid TypeError in Odoo
        return None

    def _prepare_employee_vals(self, rec: Dict) -> dict:
        vals = {
            'name': rec['employee_name'],
        }
        # Identification / classification
        if rec.get('employee_code'):
            vals['identification_id'] = rec['employee_code']
        # if rec.get('employee_type'):
        #     vals['employee_type'] = rec['employee_type']

        # Work contacts
        if rec.get('work_email'):
            vals['work_email'] = rec['work_email']
        if rec.get('work_phone'):
            vals['work_phone'] = rec['work_phone']

        # Private contact / address (on hr.employee itself)
        if rec.get('private_email'):
            vals['private_email'] = rec['private_email']
        if rec.get('private_phone'):
            vals['private_phone'] = rec['private_phone']
        if rec.get('private_street'):
            vals['private_street'] = rec['private_street']
        if rec.get('private_street2'):
            vals['private_street2'] = rec['private_street2']
        if rec.get('private_zip'):
            vals['private_zip'] = rec['private_zip']
        if rec.get('private_city'):
            vals['private_city'] = rec['private_city']
        if rec.get('private_country'):
            country_id = self.find_country_by_name(rec['private_country'])
            if country_id:
                vals['private_country_id'] = country_id

        # Personal identity
        if rec.get('identification_no'):
            vals['identification_id'] = rec['identification_no']
        if rec.get('ssn_no'):
            vals['ssnid'] = rec['ssn_no']
        if rec.get('passport_no'):
            vals['passport_id'] = rec['passport_no']
        if rec.get('gender'):
            g = str(rec['gender']).strip().lower()
            # Normalize common short codes to Odoo's selection values
            if g in ('m', 'male'):
                vals['gender'] = 'male'
            elif g in ('f', 'female'):
                vals['gender'] = 'female'
            elif g in ('o', 'other'):
                vals['gender'] = 'other'
            else:
                # Unknown/invalid value: skip setting gender to avoid errors
                pass
        if rec.get('birthday'):
            d = self._to_date_str(rec['birthday'])
            if d:
                vals['birthday'] = d
        if rec.get('place_of_birth'):
            vals['place_of_birth'] = rec['place_of_birth']
        if rec.get('country_of_birth'):
            country_id = self.find_country_by_name(rec['country_of_birth'])
            if country_id:
                vals['country_of_birth'] = country_id
        if rec.get('study_field'):
            vals['study_field'] = rec['study_field']
        if rec.get('visa_no'):
            vals['visa_no'] = rec['visa_no']
        # Map both explicit Visa Expiration and Work Permit Expiration
        # into the same hr.employee field 'visa_expire'.
        visa_exp_val = self._to_date_str(rec.get('visa_expire'))
        if not visa_exp_val and rec.get('work_permit_expiration'):
            visa_exp_val = self._to_date_str(rec.get('work_permit_expiration'))
        # if visa_exp_val:
        #     vals['visa_expire'] = visa_exp_val if visa_exp_val != 'Null' else False
        # In Odoo the field is 'permit_no', not 'work_permit_no'
        if rec.get('work_permit_no'):
            vals['permit_no'] = rec['work_permit_no']
        if rec.get('marital_status'):
            vals['marital'] = rec['marital_status'].lower()
        if rec.get('children') is not None:
            try:
                vals['children'] = int(rec['children'])
            except (TypeError, ValueError):
                pass

        # Job / department
        if rec.get('job_title'):
            vals['job_title'] = rec['job_title']

        if rec.get('department_name'):
            # Pass sheet_title as temp_company_id to distinguish departments across companies
            temp_company_id = rec.get('sheet_title')
            dept_id = self.find_or_create_department(rec['department_name'], temp_company_id=temp_company_id)
            if dept_id:
                vals['department_id'] = dept_id

        # Nationality (country)
        if rec.get('nationality'):
            nat_id = self.find_country_by_name(rec['nationality'])
            if nat_id:
                vals['nationality_id'] = nat_id

        # Cost / badge
        if rec.get('hourly_cost') is not None:
            vals['hourly_cost'] = rec['hourly_cost']
        if rec.get('badge_id'):
            vals['barcode'] = rec['badge_id']

        # Emergency contact
        if rec.get('emergency_contact_name'):
            vals['emergency_contact'] = rec['emergency_contact_name']
        if rec.get('emergency_contact_phone'):
            vals['emergency_phone'] = rec['emergency_contact_phone']

        # Custom field: store the sheet name as company indicator
        if rec.get('sheet_title'):
            vals['temp_company_id'] = rec['sheet_title']

        # parent_id handled in second pass
        return vals

    def import_employees(self, excel_path: str, dry_run: bool = False) -> Dict:
        """Main import flow (2-pass: employees then parents)."""
        stats = {
            'total_rows': 0,
            'created_employees': 0,
            'updated_employees': 0,
            'parent_links_set': 0,
            'parent_links_skipped': 0,
            'errors': [],
        }

        records = self.parse_workbook(excel_path)
        stats['total_rows'] = len(records)

        if dry_run:
            logger.info("DRY RUN MODE - No records will be created/updated")

        # Map: internal key -> hr.employee.id (or negative fake id in dry_run)
        key_to_emp_id: Dict[str, int] = {}
        # ---------- PASS 1: Create/update employees w/o parent_id ----------
        for idx, rec in enumerate(records, start=1):
            key = self._build_employee_key(rec)
            try:
                # if key in key_to_emp_id:
                #     continue

                # Get company context from record
                temp_company_id = rec.get('sheet_title')
                company_id = self._get_company_id_from_temp(temp_company_id)
                # Allow cross-company search to detect if employee exists elsewhere,
                # but still create new record for current company if needed
                emp_id = self._find_employee_in_odoo(key, company_id=company_id, temp_company_id=temp_company_id, allow_cross_company=True)
                vals = self._prepare_employee_vals(rec)

                if not dry_run:
                    if emp_id:
                        self._write('hr.employee', [emp_id], vals)
                        stats['updated_employees'] += 1
                        logger.info(
                            f"Sheet {rec['sheet_title']} Row {rec['row_index']}: Updated employee "
                            f"'{rec['employee_name']}' (ID: {emp_id})"
                        )
                    else:
                        try:
                            emp_id = self._create('hr.employee', vals)
                        except Exception as e:
                            # If creation fails, try again without barcode
                            vals['barcode'] = False
                            try:
                                emp_id = self._create('hr.employee', vals)
                            except Exception as e2:
                                # If it still fails, raise the error
                                raise e2
    

                        stats['created_employees'] += 1
                        logger.info(
                            f"Sheet {rec['sheet_title']} Row {rec['row_index']}: Created employee "
                            f"'{rec['employee_name']}' (ID: {emp_id})"
                        )
                else:
                    # DRY RUN - simulate an ID so we can still link parents
                    if not emp_id:
                        emp_id = -idx
                    logger.info(
                        f"[DRY RUN] Sheet {rec['sheet_title']} Row {rec['row_index']}: Would "
                        f"{'update' if emp_id and emp_id > 0 else 'create'} employee "
                        f"'{rec['employee_name']}'"
                    )

                key_to_emp_id[key] = emp_id

            except Exception as e:
                msg = (
                    f"Sheet {rec['sheet_title']} Row {rec['row_index']}: Error creating/updating "
                    f"employee '{rec['employee_name']}': {e}"
                )
                logger.error(msg, exc_info=True)
                stats['errors'].append(msg)

        # ---------- PASS 2: Set parent/manager ----------
        for rec in records:
            emp_key = self._build_employee_key(rec)
            mgr_key = self._build_manager_key(rec)

            if not mgr_key:
                continue

            # Get company context from record
            temp_company_id = rec.get('sheet_title')
            company_id = self._get_company_id_from_temp(temp_company_id)

            emp_id = key_to_emp_id.get(emp_key)
            if not emp_id or (dry_run and emp_id < 0):
                if not dry_run:
                    # Search with cross-company enabled to find employee in current company
                    emp_id = self._find_employee_in_odoo(emp_key, company_id=company_id, temp_company_id=temp_company_id, allow_cross_company=True)
                    if not emp_id:
                        msg = (
                            f"Sheet {rec['sheet_title']} Row {rec['row_index']}: Cannot set manager for "
                            f"'{rec['employee_name']}' – employee not found."
                        )
                        logger.warning(msg)
                        stats['errors'].append(msg)
                        stats['parent_links_skipped'] += 1
                        continue
                else:
                    stats['parent_links_skipped'] += 1
                    continue

            mgr_id = key_to_emp_id.get(mgr_key)
            if (not mgr_id) or (dry_run and mgr_id < 0):
                if not dry_run:
                    # For managers, allow cross-company search and use manager even if found in another company
                    # Managers can be shared across companies
                    mgr_id = self._find_employee_in_odoo(
                        mgr_key, 
                        company_id=company_id, 
                        temp_company_id=temp_company_id, 
                        allow_cross_company=True,
                        use_if_found_in_other_company=True
                    )
                else:
                    mgr_id = None

            if not mgr_id:
                msg = (
                    f"Sheet {rec['sheet_title']} Row {rec['row_index']}: Manager not found for "
                    f"employee '{rec['employee_name']}'. Manager key: {mgr_key}"
                )
                logger.warning(msg)
                stats['errors'].append(msg)
                stats['parent_links_skipped'] += 1
                continue

            if not dry_run:
                try:
                    self._write('hr.employee', [emp_id], {'parent_id': mgr_id})
                    stats['parent_links_set'] += 1
                    logger.info(
                        f"Sheet {rec['sheet_title']} Row {rec['row_index']}: Set manager for "
                        f"employee '{rec['employee_name']}' (Employee ID: {emp_id}, Manager ID: {mgr_id})"
                    )
                except Exception as e:
                    msg = (
                        f"Sheet {rec['sheet_title']} Row {rec['row_index']}: Error setting manager for "
                        f"'{rec['employee_name']}': {e}"
                    )
                    logger.error(msg, exc_info=True)
                    stats['errors'].append(msg)
                    stats['parent_links_skipped'] += 1
            else:
                logger.info(
                    f"[DRY RUN] Sheet {rec['sheet_title']} Row {rec['row_index']}: Would set manager for "
                    f"'{rec['employee_name']}' to '{mgr_key}'"
                )
                stats['parent_links_set'] += 1

        # Match employees and departments to companies after import
        if not dry_run:
            company_vals = {
                "LJE": 1,
                "LJD": 2,
                "LJ Fire": 10,
                "LJ Malaysia": 6,
                "LJDT": 11,
                'LJ Property':4,
                'LJ GreenTech':3,
                'PT Lingjack':1,

            }

            # Match employees to companies
            try:
                self.models.execute_kw(
                    self.db, self.uid, self.password,
                    'hr.employee',
                    'match_employee_to_company',
                    [company_vals]
                )
                logger.info("Matched employees to companies")
            except Exception as e:
                logger.error(f"Error matching employees to companies: {e}", exc_info=True)

            # Match departments to companies
            try:
                self.models.execute_kw(
                    self.db, self.uid, self.password,
                    'hr.department',
                    'match_department_to_company',
                    [company_vals]
                )
                logger.info("Matched departments to companies")
            except Exception as e:
                logger.error(f"Error matching departments to companies: {e}", exc_info=True)
        else:
            logger.info("[DRY RUN] Would match employees and departments to companies")

        return stats


# ---------------------------------------------------------
# Utility to test connection only
# ---------------------------------------------------------

def test_connection(url: str, db: str, username: str, password: str):
    print("Testing Odoo connection...")
    print(f"URL: {url}")
    print(f"Database: {db}")
    print(f"Username: {username}")
    print()

    try:
        importer = OdooEmployeeImporter(url, db, username, password)
        print("\u2713 Connection successful!")
        print(f"\u2713 Authenticated as user ID: {importer.uid}")
        return True
    except Exception as e:
        print(f"\u2717 Connection failed: {e}")
        return False


# ---------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------

def main():
    url = ODOO_URL
    db = ODOO_DB
    username = ODOO_USERNAME
    password = ODOO_PASSWORD
    excel_file = EMPLOYEE_EXCEL_FILE
    dry_run = DRY_RUN

    if '--test' in sys.argv:
        test_connection(url, db, username, password)
        return

    # First positional arg can override Excel path
    if len(sys.argv) > 1 and sys.argv[1] not in ('--execute', '--test'):
        excel_file = sys.argv[1]

    if '--execute' in sys.argv:
        dry_run = True

    # Resolve Excel path relative to this script directory if not absolute,
    # so it works both when run directly and via run_all_imports.py.
    excel_path = Path(excel_file)
    if not excel_path.is_absolute():
        excel_path = Path(__file__).parent / excel_path
        excel_file = str(excel_path)

    try:
        importer = OdooEmployeeImporter(url, db, username, password)
        stats = importer.import_employees(excel_file, dry_run=dry_run)

        print("\n" + "=" * 60)
        print("EMPLOYEE IMPORT STATISTICS")
        print("=" * 60)
        print(f"Rows processed:        {stats['total_rows']}")
        print(f"Employees created:     {stats['created_employees']}")
        print(f"Employees updated:     {stats['updated_employees']}")
        print(f"Manager links created: {stats['parent_links_set']}")
        print(f"Manager links skipped: {stats['parent_links_skipped']}")
        print(f"Errors:                {len(stats['errors'])}")

        if stats['errors']:
            print("\nErrors (first 10):")
            for err in stats['errors'][:30]:
                print(f"  - {err}")
            if len(stats['errors']) > 10:
                print(f"  ... and {len(stats['errors']) - 10} more")

        print("=" * 60)

        if dry_run:
            print("\nNOTE: This was a DRY RUN. No records were created/updated.")
            print("Run with --execute flag to actually import data.")
    except ConnectionError as e:
        print("\n" + "=" * 60)
        print("CONNECTION ERROR")
        print("=" * 60)
        print(str(e))
        print("=" * 60)
        sys.exit(1)
    except Exception as e:
        logger.error(f"Import failed: {e}", exc_info=True)
        print("\n" + "=" * 60)
        print("ERROR")
        print("=" * 60)
        print(str(e))
        print("=" * 60)
        sys.exit(1)


if __name__ == '__main__':
    main()
