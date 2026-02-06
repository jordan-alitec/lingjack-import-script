#!/usr/bin/env python3
"""
Odoo 18 BoM Import Script
Imports Bill of Materials from Excel file to Odoo mrp.bom model
"""

import xmlrpc.client
from openpyxl import load_workbook
import sys
from typing import List, Dict, Optional
import logging
import socket
from urllib.parse import urlparse

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class OdooBoMImporter:
    """Import BoM data from Excel to Odoo 18"""
    
    def __init__(self, url: str, db: str, username: str, password: str):
        """
        Initialize Odoo connection
        
        Args:
            url: Odoo server URL (e.g., 'http://localhost:8069')
            db: Database name
            username: Odoo username
            password: Odoo password
        """
        self.url = url
        self.db = db
        self.username = username
        self.password = password
        
        # Test connection first
        self._test_connection(url)
        
        # Connect to Odoo
        try:
            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common')
            self.uid = common.authenticate(db, username, password, {})
        except ConnectionRefusedError:
            raise ConnectionError(
                f"Cannot connect to Odoo server at {url}.\n"
                f"Please check:\n"
                f"  1. Is Odoo server running?\n"
                f"  2. Is the URL correct? (current: {url})\n"
                f"  3. Is the port correct? (check Odoo configuration)\n"
                f"  4. Is there a firewall blocking the connection?"
            )
        except Exception as e:
            raise ConnectionError(
                f"Failed to connect to Odoo: {str(e)}\n"
                f"URL: {url}\n"
                f"Please verify the server is running and accessible."
            )
        
        if not self.uid:
            raise Exception(
                f"Authentication failed for user '{username}'.\n"
                f"Please check:\n"
                f"  1. Username is correct\n"
                f"  2. Password is correct\n"
                f"  3. Database name '{db}' is correct\n"
                f"  4. User has proper permissions"
            )
        
        self.models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object')
        logger.info(f"Successfully connected to Odoo database: {db}")
    
    def _test_connection(self, url: str):
        """Test if we can reach the Odoo server"""
        try:
            parsed = urlparse(url)
            host = parsed.hostname or 'localhost'
            port = parsed.port or (8069 if parsed.scheme == 'http' else 443)
            
            logger.info(f"Testing connection to {host}:{port}...")
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(5)
            result = sock.connect_ex((host, port))
            sock.close()
            
            if result != 0:
                raise ConnectionError(
                    f"Cannot reach Odoo server at {host}:{port}.\n"
                    f"Connection test failed. Please ensure:\n"
                    f"  1. Odoo server is running\n"
                    f"  2. Port {port} is correct\n"
                    f"  3. Server is accessible from this machine"
                )
        except socket.gaierror as e:
            raise ConnectionError(
                f"Cannot resolve hostname '{host}'.\n"
                f"Please check the URL: {url}"
            )
        except Exception as e:
            # If socket test fails, we'll let XML-RPC handle it with better error
            logger.warning(f"Connection test warning: {e}")
    
    def find_product_by_reference(self, reference: str) -> Optional[int]:
        """
        Find product by reference (default_code)
        
        Args:
            reference: Product reference code
            
        Returns:
            Product ID or None if not found
        """
        if not reference or not str(reference).strip():
            return None
        
        product_ids = self.models.execute_kw(
            self.db, self.uid, self.password,
            'product.product',
            'search',
            [[('default_code', '=', str(reference).strip())]],
            {'limit': 1}
        )
        
        if product_ids:
            return product_ids[0]
        return None
    
    def create_product(
        self,
        reference: str,
        name: str = None,
        auto_create: bool = True,
        uom_name: Optional[str] = None,
    ) -> Optional[int]:
        """
        Create a product in Odoo
        
        Args:
            reference: Product reference code (default_code)
            name: Product name (if None, uses reference as name)
            auto_create: If True, creates product if not found
            
        Returns:
            Product ID or None if creation failed
        """
        if not reference or not str(reference).strip():
            return None
        
        reference = str(reference).strip()
        
        # Check if product already exists
        product_id = self.find_product_by_reference(reference)
        if product_id:
            return product_id
        
        if not auto_create:
            return None
        
        # Prepare product name
        if not name or not str(name).strip():
            # Extract name from reference if it contains brackets
            # Example: [0307418100000] Product Name -> Product Name
            if ']' in reference:
                name = reference.split(']', 1)[-1].strip()
            else:
                name = reference
        
        # Clean up product name - remove brackets and reference code if present
        name = str(name).strip()
        if name.startswith('[') and ']' in name:
            name = name.split(']', 1)[-1].strip()
        
        # Get default product category (All / Product category)
        try:
            # Try to find "All" category first
            categ_ids = self.models.execute_kw(
                self.db, self.uid, self.password,
                'product.category',
                'search',
                [[('name', '=', 'All')]],
                {'limit': 1}
            )
            if not categ_ids:
                # Try to find any category
                categ_ids = self.models.execute_kw(
                    self.db, self.uid, self.password,
                    'product.category',
                    'search',
                    [[]],
                    {'limit': 1}
                )
            categ_id = categ_ids[0] if categ_ids else False  # False instead of None for Odoo
        except Exception as e:
            logger.warning(f"Could not find product category: {e}")
            categ_id = False
        
        # Get product UOM, prioritizing UOM from Excel column if provided
        uom_id = False
        # 1) Try UOM from Excel column (via helper) if given
        if uom_name:
            try:
                uom_from_column = self.find_or_create_uom(uom_name)
                if uom_from_column:
                    uom_id = uom_from_column
            except Exception as e:
                logger.warning(f"Could not resolve UOM '{uom_name}' from column: {e}")

        # 2) Fallback to default UOM (Units) if nothing found from column
        if not uom_id:
            try:
                uom_ids = self.models.execute_kw(
                    self.db,
                    self.uid,
                    self.password,
                    'uom.uom',
                    'search',
                    [[('name', '=', 'Units')]],
                    {'limit': 1},
                )
                if not uom_ids:
                    # Try to find any UOM
                    uom_ids = self.models.execute_kw(
                        self.db,
                        self.uid,
                        self.password,
                        'uom.uom',
                        'search',
                        [[]],
                        {'limit': 1},
                    )
                uom_id = uom_ids[0] if uom_ids else False
            except Exception as e:
                logger.warning(f"Could not find default UOM: {e}")
                uom_id = False
        
        # Create product template first
        product_vals = {
            'name': name,
            'default_code': reference,
            'type': 'consu',  # 'product', 'consu', or 'service'
            'import_newly_created': True,
        }
        
        # Only add categ_id if we found one
        if categ_id:
            product_vals['categ_id'] = categ_id
        
        # Only add UOM if we found one
        if uom_id:
            product_vals['uom_id'] = uom_id
            product_vals['uom_po_id'] = uom_id
        
        logger.debug(f"Creating product with values: {product_vals}")
        
        try:
            # Create product template
            template_id = self.models.execute_kw(
                self.db, self.uid, self.password,
                'product.template',
                'create',
                [product_vals]
            )
            
            # Get the product variant (product.product)
            product_ids = self.models.execute_kw(
                self.db, self.uid, self.password,
                'product.product',
                'search',
                [[('product_tmpl_id', '=', template_id)]],
                {'limit': 1}
            )
            
            if product_ids:
                product_id = product_ids[0]
                logger.info(f"Created product: {name} (Reference: {reference}, ID: {product_id})")
                return product_id
            else:
                logger.error(f"Failed to create product variant for {reference}")
                return None
                
        except xmlrpc.client.Fault as e:
            # Odoo-specific error
            error_msg = str(e)
            logger.error(f"Odoo error creating product '{reference}': {error_msg}")
            # Try to extract more details from the fault
            if hasattr(e, 'faultString'):
                logger.error(f"Fault details: {e.faultString}")
            return None
        except Exception as e:
            # General error
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"Error creating product '{reference}': {str(e)}")
            logger.debug(f"Full traceback: {error_details}")
            return None
    
    def find_or_create_uom(self, uom_name: str) -> Optional[int]:
        """
        Find or create Unit of Measure by name
        
        Args:
            uom_name: UOM name (e.g., 'nos.', 'pcs')
            
        Returns:
            UOM ID or None
        """
        if not uom_name or not str(uom_name).strip():
            return None
        
        uom_name = str(uom_name).strip().lower()
        
        # Try to find existing UOM
        uom_ids = self.models.execute_kw(
            self.db, self.uid, self.password,
            'uom.uom',
            'search',
            [[('name', 'ilike', uom_name)]],
            {'limit': 1}
        )
        
        if uom_ids:
            return uom_ids[0]
        
        # Try common UOM codes
        uom_mapping = {
            'nos.': 'Units',
            'pcs': 'Units',
            'unit': 'Units',
            'piece': 'Units',
        }
        
        search_name = uom_mapping.get(uom_name, uom_name.title())
        uom_ids = self.models.execute_kw(
            self.db, self.uid, self.password,
            'uom.uom',
            'search',
            [[('name', '=', search_name)]],
            {'limit': 1}
        )
        
        if uom_ids:
            return uom_ids[0]
        
        logger.warning(f"UOM '{uom_name}' not found. Using default UOM.")
        return None
    
    def create_bom(self, product_id: int, product_name: str = None) -> int:
        """
        Create a BoM record
        
        Args:
            product_id: Product ID for the BoM
            product_name: Optional product name for BoM name
            
        Returns:
            BoM ID
        """
        bom_vals = {
            'product_id': product_id,
            'product_tmpl_id': self.models.execute_kw(
                self.db, self.uid, self.password,
                'product.product',
                'read',
                [[product_id]],
                {'fields': ['product_tmpl_id']}
            )[0]['product_tmpl_id'][0],
            'type': 'normal',  # normal, phantom, or subcontract
            'empty_cabinet_bom': True,
        }
        
        if product_name:
            bom_vals['display_name'] = product_name
        
        bom_id = self.models.execute_kw(
            self.db, self.uid, self.password,
            'mrp.bom',
            'create',
            [bom_vals]
        )
        
        logger.info(f"Created BoM ID {bom_id} for product ID {product_id}")
        return bom_id
    
    def create_bom_line(self, bom_id: int, component_product_id: int, 
                       quantity: float, uom_id: Optional[int] = None) -> int:
        """
        Create a BoM line (component)
        
        Args:
            bom_id: BoM ID
            component_product_id: Component product ID
            quantity: Quantity needed
            uom_id: Optional UOM ID
            
        Returns:
            BoM line ID
        """
        line_vals = {
            'bom_id': bom_id,
            'product_id': component_product_id,
            'product_uom_id': uom_id,  # Will use product's default UOM if None
            'product_qty': quantity,
        }
        
        line_id = self.models.execute_kw(
            self.db, self.uid, self.password,
            'mrp.bom.line',
            'create',
            [line_vals]
        )
        
        return line_id
    
    def parse_excel(self, excel_path: str) -> List[Dict]:
        """
        Parse Excel file and extract BoM data
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            List of BoM dictionaries with components
        """
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        
        boms = []
        current_bom = None
        
        # Read header row
        header = [cell.value for cell in ws[1]]
        logger.info(f"Excel columns: {header}")
        
        # Process data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            product_name = row[0]  # Column 1: Product
            reference = row[4] if len(row) > 4 else None  # Column 5: Reference
            component_ref = row[5] if len(row) > 5 else None  # Column 6: Component No
            quantity = row[6] if len(row) > 6 else None  # Column 7: Quantity
            uom_name = row[7] if len(row) > 7 else None  # Column 8: UOM
            
            # If product_name exists, it's a new BoM
            if product_name and str(product_name).strip():
                # Save previous BoM if exists
                if current_bom:
                    boms.append(current_bom)
                
                # Clean product name - remove reference code if present at the start
                product_name_clean = str(product_name).strip()
                if product_name_clean.startswith('[') and ']' in product_name_clean:
                    # Extract name after the bracket: [CODE] Name -> Name
                    product_name_clean = product_name_clean.split(']', 1)[-1].strip()
                
                # Start new BoM
                current_bom = {
                    'product_name': product_name_clean,
                    'reference': str(reference).strip() if reference else None,
                    'components': []
                }
                logger.debug(f"Found new BoM: {current_bom['product_name']}")
            
            # Add component to current BoM
            if current_bom and component_ref and quantity:
                try:
                    qty = float(quantity) if quantity else 0.0
                    if qty > 0:
                        current_bom['components'].append({
                            'component_ref': str(component_ref).strip(),
                            'quantity': qty,
                            'uom': str(uom_name).strip() if uom_name else None
                        })
                except (ValueError, TypeError) as e:
                    logger.warning(f"Row {row_idx}: Invalid quantity '{quantity}': {e}")
        
        # Add last BoM
        if current_bom:
            boms.append(current_bom)
        
        wb.close()
        logger.info(f"Parsed {len(boms)} BoMs from Excel file")
        return boms
    
    def import_boms(self, excel_path: str, dry_run: bool = False) -> Dict:
        """
        Import BoMs from Excel to Odoo
        
        Args:
            excel_path: Path to Excel file
            dry_run: If True, only validate without creating records
            
        Returns:
            Dictionary with import statistics
        """
        stats = {
            'total_boms': 0,
            'created_boms': 0,
            'skipped_boms': 0,
            'total_lines': 0,
            'created_lines': 0,
            'skipped_lines': 0,
            'created_products': 0,
            'errors': []
        }
        
        # Parse Excel
        boms = self.parse_excel(excel_path)
        stats['total_boms'] = len(boms)
        
        if dry_run:
            logger.info("DRY RUN MODE - No records will be created")
        
        # Process each BoM
        for bom_idx, bom_data in enumerate(boms, 1):
            try:
                # Find or create product
                if not bom_data['reference']:
                    error_msg = f"BoM {bom_idx}: No reference code found"
                    logger.error(error_msg)
                    stats['errors'].append(error_msg)
                    stats['skipped_boms'] += 1
                    continue
                
                product_id = self.find_product_by_reference(bom_data['reference'])
                if not product_id:
                    # Create product if not found
                    logger.info(f"BoM {bom_idx}: Product '{bom_data['reference']}' not found, creating...")
                    try:
                        # Use the first component's UOM as the product UOM when available
                        uom_name = None
                        if bom_data['components']:
                            uom_name = bom_data['components'][0].get('uom')

                        product_id = self.create_product(
                            bom_data['reference'],
                            bom_data['product_name'],
                            auto_create=True,
                            uom_name=uom_name,
                        )
                        if not product_id:
                            error_msg = f"BoM {bom_idx}: Failed to create product with reference '{bom_data['reference']}'. Check logs for details."
                            logger.error(error_msg)
                            stats['errors'].append(error_msg)
                            stats['skipped_boms'] += 1
                            continue
                        else:
                            stats['created_products'] += 1
                    except Exception as e:
                        error_msg = f"BoM {bom_idx}: Exception creating product '{bom_data['reference']}': {str(e)}"
                        logger.error(error_msg, exc_info=True)
                        stats['errors'].append(error_msg)
                        stats['skipped_boms'] += 1
                        continue
                
                # Create BoM
                if not dry_run:
                    bom_id = self.create_bom(product_id, bom_data['product_name'])
                else:
                    bom_id = None
                    logger.info(f"[DRY RUN] Would create BoM for product: {bom_data['reference']}")
                
                stats['created_boms'] += 1
                stats['total_lines'] += len(bom_data['components'])
                
                # Create BoM lines
                for comp_idx, component in enumerate(bom_data['components'], 1):
                    try:
                        comp_product_id = self.find_product_by_reference(component['component_ref'])
                        if not comp_product_id:
                            # Create component product if not found
                            logger.info(f"BoM {bom_idx}, Component {comp_idx}: Product '{component['component_ref']}' not found, creating...")
                            comp_product_id = self.create_product(
                                component['component_ref'],
                                name=None,  # Component name not available in Excel
                                auto_create=True,
                                uom_name=component.get('uom'),
                            )
                            if not comp_product_id:
                                error_msg = f"BoM {bom_idx}, Component {comp_idx}: Failed to create product '{component['component_ref']}'"
                                logger.warning(error_msg)
                                stats['errors'].append(error_msg)
                                stats['skipped_lines'] += 1
                                continue
                            else:
                                stats['created_products'] += 1
                        
                        uom_id = None
                        if component['uom']:
                            uom_id = self.find_or_create_uom(component['uom'])
                        
                        if not dry_run:
                            self.create_bom_line(
                                bom_id, 
                                comp_product_id, 
                                component['quantity'],
                                uom_id
                            )
                        else:
                            logger.info(f"[DRY RUN] Would create line: {component['component_ref']} x {component['quantity']}")
                        
                        stats['created_lines'] += 1
                        
                    except Exception as e:
                        error_msg = f"BoM {bom_idx}, Component {comp_idx}: {str(e)}"
                        logger.error(error_msg)
                        stats['errors'].append(error_msg)
                        stats['skipped_lines'] += 1
                
            except Exception as e:
                error_msg = f"BoM {bom_idx}: {str(e)}"
                logger.error(error_msg)
                stats['errors'].append(error_msg)
                stats['skipped_boms'] += 1
        
        return stats


def test_connection(url: str, db: str, username: str, password: str):
    """Test Odoo connection without importing"""
    print("Testing Odoo connection...")
    print(f"URL: {url}")
    print(f"Database: {db}")
    print(f"Username: {username}")
    print()
    
    try:
        importer = OdooBoMImporter(url, db, username, password)
        print("✓ Connection successful!")
        print(f"✓ Authenticated as user ID: {importer.uid}")
        return True
    except Exception as e:
        print(f"✗ Connection failed: {e}")
        return False


def main():
    """Main function"""
    # Try to load configuration from the central BOM config.py in the parent directory.
    # This ensures all BOM-related scripts share a single config source.
    try:
        from pathlib import Path
        script_dir = Path(__file__).parent
        bom_dir = script_dir.parent

        # Prefer the BOM root directory on sys.path so we pick /Bom/BOM/config.py
        import sys as _sys  # local import to avoid polluting module namespace
        if str(bom_dir) not in _sys.path:
            _sys.path.insert(0, str(bom_dir))

        import config  # now resolves to BOM root config.py

        ODOO_URL = getattr(config, "ODOO_URL", "https://lingjack-data-migration-script-27889585.dev.odoo.com")
        ODOO_DB = getattr(config, "ODOO_DB", "lingjack-data-migration-script-27889585")
        ODOO_USERNAME = getattr(config, "ODOO_USERNAME", "DataMigration1")
        ODOO_PASSWORD = getattr(config, "ODOO_PASSWORD", "Alitec!@#456789")
        # Prefer BOM_* keys from central config, fall back to legacy names if present
        EXCEL_FILE = getattr(
            config,
            "BOM_EXCEL_FILE",
            getattr(config, "EXCEL_FILE", "output.xlsx"),
        )
        DRY_RUN = getattr(
            config,
            "BOM_DRY_RUN",
            getattr(config, "DRY_RUN", True),
        )
    except ImportError:
        # Default configuration - UPDATE THESE VALUES
        # Odoo Connection Settings
        ODOO_URL = 'https://lingjack-data-migration-script-27889585.dev.odoo.com'  # Your Odoo server URL
        ODOO_DB = 'lingjack-data-migration-script-27889585'            # Your Odoo database name
        ODOO_USERNAME = 'DataMigration1'              # Your Odoo username
        ODOO_PASSWORD = 'Alitec!@#456789'              # Your Odoo password

        EXCEL_FILE = "output.xlsx"
        DRY_RUN = True  # Set to False to actually import
    
    # Handle command line arguments
    if '--test' in sys.argv:
        # Test connection only
        test_connection(ODOO_URL, ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD)
        return
    
    if len(sys.argv) > 1 and sys.argv[1] != '--execute' and sys.argv[1] != '--test':
        EXCEL_FILE = sys.argv[1]
    
    if '--execute' in sys.argv:
        DRY_RUN = False
    
    # Resolve Excel path relative to this script directory if not absolute,
    # so it works both when run directly and via run_all_imports.py.
    excel_path = Path(EXCEL_FILE)
    if not excel_path.is_absolute():
        excel_path = script_dir / excel_path
        EXCEL_FILE = str(excel_path)
    
    try:
        importer = OdooBoMImporter(ODOO_URL, ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD)
        stats = importer.import_boms(EXCEL_FILE, dry_run=DRY_RUN)
        
        # Print statistics
        print("\n" + "="*60)
        print("IMPORT STATISTICS")
        print("="*60)
        print(f"Total BoMs found: {stats['total_boms']}")
        print(f"BoMs created: {stats['created_boms']}")
        print(f"BoMs skipped: {stats['skipped_boms']}")
        print(f"Total lines found: {stats['total_lines']}")
        print(f"Lines created: {stats['created_lines']}")
        print(f"Lines skipped: {stats['skipped_lines']}")
        print(f"Products created: {stats['created_products']}")
        print(f"Errors: {len(stats['errors'])}")
        
        if stats['errors']:
            print("\nErrors encountered:")
            for error in stats['errors'][:10]:  # Show first 10 errors
                print(f"  - {error}")
            if len(stats['errors']) > 10:
                print(f"  ... and {len(stats['errors']) - 10} more errors")
        
        print("="*60)
        
        if DRY_RUN:
            print("\nNOTE: This was a DRY RUN. No records were created.")
            print("Run with --execute flag to actually import data.")
        
    except ConnectionError as e:
        print("\n" + "="*60)
        print("CONNECTION ERROR")
        print("="*60)
        print(str(e))
        print("\nTroubleshooting steps:")
        print("  1. Verify Odoo server is running:")
        print("     - Check if Odoo process is active")
        print("     - Look for Odoo in running services")
        print("  2. Test the connection manually:")
        print(f"     - Open browser: {ODOO_URL}")
        print(f"     - Or test with: curl {ODOO_URL}/xmlrpc/2/common")
        print("  3. Check your configuration in config.py:")
        print(f"     - URL: {ODOO_URL}")
        print(f"     - Database: {ODOO_DB}")
        print(f"     - Username: {ODOO_USERNAME}")
        print("="*60)
        sys.exit(1)
    except Exception as e:
        logger.error(f"Import failed: {e}")
        print("\n" + "="*60)
        print("ERROR")
        print("="*60)
        print(str(e))
        print("="*60)
        sys.exit(1)


if __name__ == '__main__':
    main()


