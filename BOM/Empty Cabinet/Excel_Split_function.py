from openpyxl import load_workbook, Workbook

def is_col_a_empty(row):
    return not row[0] or str(row[0]).strip() == ""


def split_excel_to_excel(
    input_xlsx,
    output_xlsx,
    batch_size=2000,
    has_header=True,
):
    src_wb = load_workbook(input_xlsx, read_only=True, data_only=True)
    src_ws = src_wb.active

    out_wb = Workbook()
    out_ws = out_wb.active
    out_wb.remove(out_ws)

    rows = list(src_ws.iter_rows(values_only=True))

    if not rows:
        print("Input Excel is empty.")
        return

    header = rows[0] if has_header else None
    data_rows = rows[1:] if has_header else rows

    sheet_index = 1
    current_rows = []

    def write_row_as_text(ws, row):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=ws.max_row + 1 if col_idx == 1 else ws.max_row,
                            column=col_idx)
            cell.value = "" if value is None else str(value)
            cell.number_format = '@'

    def flush(rows, idx):
        ws = out_wb.create_sheet(f"Sheet_{idx:03d}")

        if header:
            for col_idx, value in enumerate(header, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = str(value)
                cell.number_format = '@'

        for r in rows:
            ws.append([""] * len(r))  # create empty row
            for col_idx, value in enumerate(r, start=1):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.value = "" if value is None else str(value)
                cell.number_format = '@'

        print(f"Wrote {len(rows)} rows -> Sheet_{idx:03d}")

    i = 0
    while i < len(data_rows):
        current_rows.append(data_rows[i])

        if len(current_rows) >= batch_size:
            j = i + 1
            while j < len(data_rows) and is_col_a_empty(data_rows[j]):
                current_rows.append(data_rows[j])
                j += 1

            flush(current_rows, sheet_index)
            sheet_index += 1
            current_rows = []
            i = j
        else:
            i += 1

    if current_rows:
        flush(current_rows, sheet_index)

    out_wb.save(output_xlsx)
    print(f"\nExcel file created: {output_xlsx}")



split_excel_to_excel(
    input_xlsx="output.xlsx",
    output_xlsx="split_output.xlsx",
    batch_size=1000,
    has_header=True,
)
