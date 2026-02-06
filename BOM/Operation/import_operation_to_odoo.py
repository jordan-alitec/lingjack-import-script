#!/usr/bin/env python3
"""
Odoo 18 BoM Operation Import Script

Reads operations from Book1.xlsx and creates:
- Operation templates (mrp.operation.template)
- Work centers (mrp.workcenter)
- Routing workcenters (mrp.routing.workcenter) linked to BOMs

Column A: Domain filter for product_id in mrp.bom
Column B: Work Center name
Column C: Operation Template name
"""

import sys
import logging
import ast
import re
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
from datetime import datetime

import xmlrpc.client
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


class OdooBoMOperationUpdater:
    """Import BoM operations in Odoo 18 from Excel"""

    def __init__(self, url: str, db: str, username: str, password: str):
        """
        Initialize Odoo connection

        Args:
            url: Odoo server URL (e.g., 'http://localhost:8069')
            db: Database name
            username: Odoo username
            password: Odoo password
        """
        self.url = url
        self.db = db
        self.username = username
        self.password = password

        # Authenticate
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
        self.uid = common.authenticate(db, username, password, {})
        if not self.uid:
            raise Exception(
                f"Authentication failed for user '{username}' "
                f"on database '{db}'."
            )

        self.models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")
        logger.info("Connected to Odoo DB '%s' as '%s'", db, username)

    def _parse_domain(self, domain_str: str) -> Optional[List]:
        """
        Parse domain string from Excel into Odoo domain format.

        Args:
            domain_str: Domain string from Excel (e.g., "[('default_code', '=ilike', '70%')]")

        Returns:
            Parsed domain list or None if invalid
        """
        if not domain_str or not isinstance(domain_str, str):
            return None

        # Clean up the domain string - remove extra whitespace but preserve structure
        domain_str = domain_str.strip()
        if not domain_str:
            return None

        # Normalize newlines and whitespace for better parsing
        normalized = re.sub(r'\s+', ' ', domain_str)

        try:
            # Try to evaluate as Python literal (Odoo supports =ilike as-is)
            domain = ast.literal_eval(normalized)
            if isinstance(domain, list):
                return domain
        except (ValueError, SyntaxError) as e:
            logger.warning("Failed to parse domain as literal: %s - %s", domain_str[:100], e)
            # Try original string if normalized version failed
            try:
                domain = ast.literal_eval(domain_str)
                if isinstance(domain, list):
                    return domain
            except Exception as e2:
                logger.error("Failed to parse domain after all attempts: %s - %s", domain_str[:100], e2)

        return None

    def _find_products_by_domain(self, domain: List) -> List[int]:
        """
        Find product.product IDs matching the domain.

        Args:
            domain: Odoo domain filter

        Returns:
            List of product.product IDs
        """
        print(f"\n\n_find_products_by_domain: {domain}")
        try:
            product_ids = self.models.execute_kw(
                self.db,
                self.uid,
                self.password,
                'product.product',
                'search',
                [domain],
            )
            return product_ids
        except Exception as e:
            logger.error("Error searching products with domain %s: %s", domain, e)
            return []

    def _find_boms_by_products(self, product_ids: List[int]) -> List[Dict[str, Any]]:
        """
        Find mrp.bom records for given product IDs.

        Args:
            product_ids: List of product.product IDs

        Returns:
            List of BOM records with fields: id, product_id, product_name, product_code
        """
        if not product_ids:
            return []

        try:
            bom_ids = self.models.execute_kw(
                self.db,
                self.uid,
                self.password,
                'mrp.bom',
                'search',
                [[('product_id', 'in', product_ids)]],
            )

            if not bom_ids:
                return []

            # Read BOM details
            boms = self.models.execute_kw(
                self.db,
                self.uid,
                self.password,
                'mrp.bom',
                'read',
                [bom_ids],
                {'fields': ['id', 'product_id', 'product_tmpl_id']},
            )

            # Get product details
            product_ids_list = [bom['product_id'][0] if bom.get('product_id') else None for bom in boms]
            product_ids_list = [pid for pid in product_ids_list if pid]

            if product_ids_list:
                products = self.models.execute_kw(
                    self.db,
                    self.uid,
                    self.password,
                    'product.product',
                    'read',
                    [product_ids_list],
                    {'fields': ['id', 'name', 'default_code']},
                )
                product_dict = {p['id']: p for p in products}

                # Enrich BOM data
                for bom in boms:
                    product_id = bom.get('product_id')
                    if product_id and product_id[0] in product_dict:
                        product = product_dict[product_id[0]]
                        bom['product_name'] = product.get('name', '')
                        bom['product_code'] = product.get('default_code', '')
                    else:
                        bom['product_name'] = ''
                        bom['product_code'] = ''

            return boms
        except Exception as e:
            logger.error("Error searching BOMs for products %s: %s", product_ids, e)
            return []

    def _get_or_create_operation_template(self, template_name: str) -> Optional[int]:
        """
        Search for or create mrp.operation.template.

        Args:
            template_name: Operation template name

        Returns:
            Template ID or None if error
        """
        if not template_name or not isinstance(template_name, str):
            return None

        template_name = template_name.strip()
        if not template_name:
            return None

        try:
            # Search for existing template
            template_ids = self.models.execute_kw(
                self.db,
                self.uid,
                self.password,
                'mrp.operation.template',
                'search',
                [[('name', '=', template_name)]],
            )

            if template_ids:
                return template_ids[0]

            # Create new template
            template_ids = self.models.execute_kw(
                self.db,
                self.uid,
                self.password,
                'mrp.operation.template',
                'create',
                [[{'name': template_name}]],
            )
            template_id = template_ids[0] if template_ids else None
            if template_id:
                logger.info("Created operation template: %s (ID: %s)", template_name, template_id)
            return template_id
        except Exception as e:
            logger.error("Error getting/creating operation template '%s': %s", template_name, e)
            return None

    def _get_or_create_workcenter(self, workcenter_name: str) -> Optional[int]:
        """
        Search for or create mrp.workcenter.

        Args:
            workcenter_name: Work center name

        Returns:
            Work center ID or None if error
        """
        if not workcenter_name or not isinstance(workcenter_name, str):
            return None

        workcenter_name = workcenter_name.strip()
        if not workcenter_name:
            return None

        try:
            # Search for existing work center
            workcenter_ids = self.models.execute_kw(
                self.db,
                self.uid,
                self.password,
                'mrp.workcenter',
                'search',
                [[('name', '=', workcenter_name)]],
            )

            if workcenter_ids:
                return workcenter_ids[0]

            # Create new work center
            workcenter_ids = self.models.execute_kw(
                self.db,
                self.uid,
                self.password,
                'mrp.workcenter',
                'create',
                [[{'name': workcenter_name}]],
            )
            workcenter_id = workcenter_ids[0] if workcenter_ids else None
            if workcenter_id:
                logger.info("Created work center: %s (ID: %s)", workcenter_name, workcenter_id)
            return workcenter_id
        except Exception as e:
            logger.error("Error getting/creating work center '%s': %s", workcenter_name, e)
            return None

    def _link_workcenter_to_template(self, template_id: int, workcenter_id: int) -> bool:
        """
        Link work center to operation template.

        Args:
            template_id: Operation template ID
            workcenter_id: Work center ID

        Returns:
            True if successful, False otherwise
        """
        try:
            self.models.execute_kw(
                self.db,
                self.uid,
                self.password,
                'mrp.operation.template',
                'write',
                [[template_id], {'work_center_id': workcenter_id}],
            )
            return True
        except Exception as e:
            logger.error("Error linking work center %s to template %s: %s", workcenter_id, template_id, e)
            return False

    def _create_routing_workcenter(
        self,
        bom_id: int,
        template_id: int,
        workcenter_id: int,
        sequence: int = 100,
    ) -> Optional[int]:
        """
        Create mrp.routing.workcenter for a BOM.

        Args:
            bom_id: BOM ID
            template_id: Operation template ID
            workcenter_id: Work center ID
            sequence: Sequence number (default 100)

        Returns:
            Routing workcenter ID or None if error
        """
        try:
            # Get template name for the operation name
            template_data = self.models.execute_kw(
                self.db,
                self.uid,
                self.password,
                'mrp.operation.template',
                'read',
                [template_id],
                {'fields': ['name']},
            )
            operation_name = template_data[0].get('name', '') if template_data else ''

            routing_ids = self.models.execute_kw(
                self.db,
                self.uid,
                self.password,
                'mrp.routing.workcenter',
                'create',
                [[{
                    'bom_id': bom_id,
                    'mrp_operation_temp_id': template_id,
                    'workcenter_id': workcenter_id,
                    'name': operation_name,
                    'sequence': sequence,
                }]],
            )
            routing_id = routing_ids[0] if routing_ids else None
            return routing_id
        except Exception as e:
            logger.error(
                "Error creating routing workcenter for BOM %s, template %s, workcenter %s: %s",
                bom_id, template_id, workcenter_id, e
            )
            return None

    def _sanitize_sheet_name(self, name: str, max_length: int = 31) -> str:
        """
        Sanitize string for use as Excel sheet name.

        Args:
            name: Original name
            max_length: Maximum sheet name length (Excel limit is 31)

        Returns:
            Sanitized name
        """
        # Remove invalid characters: \ / ? * [ ]
        sanitized = re.sub(r'[\\/?*\[\]]', '_', str(name))
        # Truncate to max length
        if len(sanitized) > max_length:
            sanitized = sanitized[:max_length]
        return sanitized

    def process_operations_from_excel(
        self,
        excel_path: str,
        sheet_name: Optional[str] = None,
        header_row: int = 1,
        domain_col: int = 1,
        workcenter_col: int = 2,
        template_col: int = 3,
        retrieve_mode: bool = False,
        output_path: Optional[str] = None,
    ):
        """
        Read Excel and create operation templates, work centers, and routing workcenters.

        Args:
            excel_path: Path to Excel file (e.g. 'Book1.xlsx')
            sheet_name: Sheet name (default: active sheet)
            header_row: Row number of the header (data starts at header_row+1)
            domain_col: 1-based column index of domain filter (Column A)
            workcenter_col: 1-based column index of work center name (Column B)
            template_col: 1-based column index of operation template name (Column C)
            retrieve_mode: If True, only retrieve BOMs without creating records
            output_path: Path for output Excel file in retrieve mode
        """
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        start_row = header_row + 1
        max_row = ws.max_row

        # Statistics
        stats = {
            'total_rows': 0,
            'processed_rows': 0,
            'errors': 0,
            'boms_found': 0,
            'templates_created': 0,
            'workcenters_created': 0,
            'routings_created': 0,
        }

        # Error log
        errors = []

        # Current domain (for handling empty cells in Column A)
        current_domain: Optional[List] = None
        current_domain_str: Optional[str] = None

        # For retrieve mode
        retrieve_data: Dict[str, List[Dict]] = {}

        logger.info("=" * 60)
        logger.info("Starting operation import from Excel")
        logger.info("=" * 60)
        logger.info("Excel file: %s", excel_path)
        logger.info("Retrieve mode: %s", retrieve_mode)

        # Process each row
        for row_idx in range(start_row, max_row + 1):
            row_cells = [cell.value for cell in ws[row_idx]]
            stats['total_rows'] += 1

            try:
                # Extract domain from Column A
                domain_str = None
                if len(row_cells) >= domain_col:
                    domain_value = row_cells[domain_col - 1]
                    if domain_value is not None:
                        domain_str = str(domain_value).strip()
                        if domain_str:
                            current_domain = self._parse_domain(domain_str)
                            current_domain_str = domain_str
                     
                    elif current_domain is not None:
                        # Use previous row's domain
                        domain_str = current_domain_str

                # Extract work center from Column B
                workcenter_name = None
                if len(row_cells) >= workcenter_col:
                    workcenter_value = row_cells[workcenter_col - 1]
                    if workcenter_value is not None:
                        workcenter_name = str(workcenter_value).strip()

                # Extract template from Column C
                template_name = None
                if len(row_cells) >= template_col:
                    template_value = row_cells[template_col - 1]
                    if template_value is not None:
                        template_name = str(template_value).strip()

                # Skip if no domain or missing required fields
                if not current_domain:
                    if domain_str:
                        logger.warning("Row %s: Could not parse domain: %s", row_idx, domain_str)
                        errors.append({
                            'row': row_idx,
                            'domain': domain_str,
                            'workcenter': workcenter_name,
                            'template': template_name,
                            'error': f"Could not parse domain: {domain_str}",
                        })
                    continue

                if not workcenter_name or not template_name:
                    logger.warning("Row %s: Missing work center or template name", row_idx)
                    continue

                # Find products matching domain
                product_ids = self._find_products_by_domain(current_domain)
                if not product_ids:
                    logger.warning("Row %s: No products found for domain: %s", row_idx, current_domain_str)
                    errors.append({
                        'row': row_idx,
                        'domain': current_domain_str,
                        'workcenter': workcenter_name,
                        'template': template_name,
                        'error': "No products found for domain",
                    })
                    continue

                # Find BOMs for these products
                boms = self._find_boms_by_products(product_ids)
                if not boms:
                    logger.warning("Row %s: No BOMs found for products matching domain: %s", row_idx, current_domain_str)
                    errors.append({
                        'row': row_idx,
                        'domain': current_domain_str,
                        'workcenter': workcenter_name,
                        'template': template_name,
                        'error': "No BOMs found for products",
                    })
                    continue

                stats['boms_found'] += len(boms)

                if retrieve_mode:
                    # Store BOM data for export
                    sheet_key = current_domain_str or f"Row_{row_idx}"
                    if sheet_key not in retrieve_data:
                        retrieve_data[sheet_key] = []
                    for bom in boms:
                        retrieve_data[sheet_key].append({
                            'bom_id': bom.get('id'),
                            'product_name': bom.get('product_name', ''),
                            'product_code': bom.get('product_code', ''),
                        })
                    logger.info(
                        "Row %s: Found %d BOMs for domain (retrieve mode)",
                        row_idx, len(boms)
                    )
                else:
                    # Get or create operation template
                    template_id = self._get_or_create_operation_template(template_name)
                    if not template_id:
                        error_msg = f"Failed to get/create operation template: {template_name}"
                        logger.error("Row %s: %s", row_idx, error_msg)
                        errors.append({
                            'row': row_idx,
                            'domain': current_domain_str,
                            'workcenter': workcenter_name,
                            'template': template_name,
                            'error': error_msg,
                        })
                        continue

                    # Get or create work center
                    workcenter_id = self._get_or_create_workcenter(workcenter_name)
                    if not workcenter_id:
                        error_msg = f"Failed to get/create work center: {workcenter_name}"
                        logger.error("Row %s: %s", row_idx, error_msg)
                        errors.append({
                            'row': row_idx,
                            'domain': current_domain_str,
                            'workcenter': workcenter_name,
                            'template': template_name,
                            'error': error_msg,
                        })
                        continue

                    # Link work center to template
                    self._link_workcenter_to_template(template_id, workcenter_id)

                    # Create routing workcenter for each BOM
                    sequence = 100
                    routing_count = 0
                    for bom in boms:
                        routing_id = self._create_routing_workcenter(
                            bom_id=bom['id'],
                            template_id=template_id,
                            workcenter_id=workcenter_id,
                            sequence=sequence,
                        )
                        if routing_id:
                            routing_count += 1
                            sequence += 10  # Increment sequence for next operation
                        else:
                            errors.append({
                                'row': row_idx,
                                'domain': current_domain_str,
                                'workcenter': workcenter_name,
                                'template': template_name,
                                'bom_id': bom['id'],
                                'error': "Failed to create routing workcenter",
                            })

                    stats['routings_created'] += routing_count
                    logger.info(
                        "Row %s: Created %d routing workcenters for %d BOMs (template: %s, workcenter: %s)",
                        row_idx, routing_count, len(boms), template_name, workcenter_name
                    )

                stats['processed_rows'] += 1

            except Exception as e:
                stats['errors'] += 1
                error_msg = f"Unexpected error: {str(e)}"
                logger.error("Row %s: %s", row_idx, error_msg, exc_info=True)
                errors.append({
                    'row': row_idx,
                    'domain': current_domain_str,
                    'workcenter': workcenter_name if 'workcenter_name' in locals() else None,
                    'template': template_name if 'template_name' in locals() else None,
                    'error': error_msg,
                })

        wb.close()

        # Handle retrieve mode output
        if retrieve_mode and retrieve_data:
            if not output_path:
                script_dir = Path(excel_path).parent
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = str(script_dir / f"retrieved_boms_{timestamp}.xlsx")

            output_wb = Workbook()
            output_wb.remove(output_wb.active)  # Remove default sheet

            for sheet_key, bom_list in retrieve_data.items():
                sheet_name = self._sanitize_sheet_name(sheet_key)
                ws_output = output_wb.create_sheet(title=sheet_name)

                # Headers
                headers = ['BOM ID', 'Product Name', 'Product Code']
                ws_output.append(headers)

                # Style headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws_output[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                # Data rows
                for bom in bom_list:
                    ws_output.append([
                        bom.get('bom_id', ''),
                        bom.get('product_name', ''),
                        bom.get('product_code', ''),
                    ])

            output_wb.save(output_path)
            logger.info("Retrieved BOMs exported to: %s", output_path)
            logger.info("Total sheets created: %d", len(retrieve_data))

        # Summary
        logger.info("=" * 60)
        logger.info("Import Summary:")
        logger.info("  Total rows processed: %d", stats['processed_rows'])
        logger.info("  Total rows with data: %d", stats['total_rows'])
        logger.info("  BOMs found: %d", stats['boms_found'])
        if not retrieve_mode:
            logger.info("  Templates created/found: %d", stats.get('templates_created', 0))
            logger.info("  Workcenters created/found: %d", stats.get('workcenters_created', 0))
            logger.info("  Routing workcenters created: %d", stats['routings_created'])
        logger.info("  Errors: %d", stats['errors'])
        logger.info("=" * 60)

        # Log errors
        if errors:
            logger.error("=" * 60)
            logger.error("Error Details:")
            logger.error("=" * 60)
            for error in errors:
                logger.error(
                    "Row %s - Domain: %s, Work Center: %s, Template: %s, BOM ID: %s - Error: %s",
                    error.get('row'),
                    error.get('domain'),
                    error.get('workcenter'),
                    error.get('template'),
                    error.get('bom_id'),
                    error.get('error'),
                )


def main():
    """
    CLI entry point.

    Configuration is taken from the central config.py in the BOM directory.
    """
    import os
    from pathlib import Path

    # Get the BOM directory (parent of this script's directory)
    script_dir = Path(__file__).parent
    bom_dir = script_dir.parent
    config_path = bom_dir / 'config.py'

    # Add BOM directory to path to import config
    if str(bom_dir) not in sys.path:
        sys.path.insert(0, str(bom_dir))

    try:
        import config  # type: ignore


        # ODOO_URL = getattr(config, "ODOO_URL", "https://lingjack-data-migration-script-27889585.dev.odoo.com")
        # ODOO_DB = getattr(config, "ODOO_DB", "lingjack-data-migration-script-27889585")
        # ODOO_USERNAME = getattr(config, "ODOO_USERNAME", "DataMigration1")
        # ODOO_PASSWORD = getattr(config, "ODOO_PASSWORD", "Alitec!@#456789")

        ODOO_URL = getattr(config, "ODOO_URL", "http://localhost:8099")
        ODOO_DB = getattr(config, "ODOO_DB", "lingjack-rerun")
        ODOO_USERNAME = getattr(config, "ODOO_USERNAME", "admin")
        ODOO_PASSWORD = getattr(config, "ODOO_PASSWORD", "admin")
        EXCEL_FILE = getattr(config, "OPERATION_EXCEL_FILE", "Book1.xlsx")
        SHEET_NAME = getattr(config, "OPERATION_SHEET_NAME", None)
        DRY_RUN = getattr(config, "OPERATION_DRY_RUN", True)
        RETRIEVE = getattr(config, "OPERATION_RETRIEVE", False)
    except ImportError:
        logger.error(f"Failed to import config from {config_path}")
        logger.error("Please ensure config.py exists in the BOM directory")
        # Fallback defaults
        ODOO_URL = "https://lingjack-data-migration-script-27889585.dev.odoo.com"
        ODOO_DB = "lingjack-data-migration-script-27889585"
        ODOO_USERNAME = "DataMigration1"
        ODOO_PASSWORD = "Alitec!@#456789"
        EXCEL_FILE = "Book1.xlsx"
        SHEET_NAME = None
        DRY_RUN = False
        RETRIEVE = False

    ODOO_URL = 'https://lingjack.odoo.com/'
    ODOO_DB = 'alitecpteltd-lingjack-main-21976694'
    ODOO_USERNAME = 'dataimport'
    ODOO_PASSWORD = 'Admin@123456'

    # CLI overrides
    if len(sys.argv) > 1 and sys.argv[1] not in ("--execute", "--dry-run", "--retrieve"):
        EXCEL_FILE = sys.argv[1]

    if "--execute" in sys.argv:
        DRY_RUN = False
        RETRIEVE = False
    if "--dry-run" in sys.argv:
        DRY_RUN = True
        RETRIEVE = False
    if "--retrieve" in sys.argv:
        RETRIEVE = True
        DRY_RUN = False

    # Resolve Excel path relative to this script directory if not absolute
    excel_path = Path(EXCEL_FILE)
    if not excel_path.is_absolute():
        excel_path = script_dir / excel_path
        EXCEL_FILE = str(excel_path)

    logger.info("Excel file: %s", EXCEL_FILE)
    logger.info("Dry run: %s", DRY_RUN)
    logger.info("Retrieve mode: %s", RETRIEVE)

    if DRY_RUN and not RETRIEVE:
        logger.info("DRY RUN MODE - No records will be created in Odoo")
        return

    updater = OdooBoMOperationUpdater(ODOO_URL, ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD)
    updater.process_operations_from_excel(
        excel_path=EXCEL_FILE,
        sheet_name=SHEET_NAME,
        header_row=1,
        domain_col=1,  # Column A: Domain filter
        workcenter_col=2,  # Column B: Work Center name
        template_col=3,  # Column C: Operation Template name
        retrieve_mode=RETRIEVE,
    )


if __name__ == "__main__":
    main()
