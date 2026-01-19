from openpyxl import load_workbook

# --- CONFIG ---
FILE_PATH = "operation-template (1).xlsx"   # your excel file
SHEET_NAME = "Sheet1"      # sheet to work on
COL_E = 5                  # column E index
COL_A = 1                  # column A index
# --------------

# Load workbook and sheet
wb = load_workbook(FILE_PATH)
ws = wb[SHEET_NAME]

previous_e = None

for row in range(1, ws.max_row + 1):
    current_e = ws.cell(row=row, column=COL_E).value

    if previous_e is None:
        previous_e = current_e
        continue

    # If current row E equals previous row E → clear A & E
    if current_e == previous_e:
        ws.cell(row=row, column=COL_E).value = None
        ws.cell(row=row, column=COL_A).value = None

    previous_e = current_e

# Save output
wb.save("output.xlsx")
print("Done. Saved to output.xlsx")
