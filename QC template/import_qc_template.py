#!/usr/bin/env python3
"""
Odoo 18 QC Template to BOM Link Import Script

- Reads rows from an Excel file (QC-Template.xlsx): column B = QC template name,
  column C = product default code (Com No).
- Finds quality.spreadsheet.template by name (column B).
- Finds mrp.bom by product_id.default_code = column C.
- Links the BOM to the template: bom_id.qc_template_id = template_id.

Uses the same RPC pattern and config as import_employee_to_odoo.py (BOM/config.py).
"""

import xmlrpc.client
from openpyxl import load_workbook
import sys
from typing import List, Dict, Optional
import logging
import socket
from urllib.parse import urlparse
from pathlib import Path

# ---------------------------------------------------------
# CONFIGURATION (reuse BOM config when available)
# ---------------------------------------------------------

script_dir = Path(__file__).resolve().parent
# Config lives in sibling BOM folder
bom_dir = script_dir.parent / 'BOM'
config_path = bom_dir / 'config.py'

if str(bom_dir) not in sys.path:
    sys.path.insert(0, str(bom_dir))

ODOO_URL = 'https://lingjack.odoo.com/'
ODOO_DB = 'alitecpteltd-lingjack-main-21976694'
ODOO_USERNAME = 'admin'
ODOO_PASSWORD = 'Admin@123456'
QC_TEMPLATE_EXCEL_FILE = 'QC-Template.xlsx'
QC_TEMPLATE_DRY_RUN = False

# ---------------------------------------------------------
# LOGGING
# ---------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------
# Odoo connection and helpers
# ---------------------------------------------------------

class OdooQCTemplateImporter:
    """Import QC template links from Excel to Odoo 18 (mrp.bom.qc_template_id)."""

    def __init__(self, url: str, db: str, username: str, password: str):
        self.url = url
        self.db = db
        self.username = username
        self.password = password

        self._test_connection(url)

        try:
            common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common')
            self.uid = common.authenticate(db, username, password, {})
        except ConnectionRefusedError:
            raise ConnectionError(
                f"Cannot connect to Odoo server at {url}. "
                "Check server is running and URL/port are correct."
            )
        except Exception as e:
            raise ConnectionError(f"Failed to connect to Odoo: {e}")

        if not self.uid:
            raise Exception(
                f"Authentication failed for user '{username}'. "
                "Check username/password/db and access rights."
            )

        self.models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object')
        logger.info(f"Successfully connected to Odoo database: {db}")

    def _test_connection(self, url: str) -> None:
        try:
            parsed = urlparse(url)
            host = parsed.hostname or 'localhost'
            port = parsed.port or (8069 if parsed.scheme == 'http' else 443)
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(5)
            result = sock.connect_ex((host, port))
            sock.close()
            if result != 0:
                raise ConnectionError(f"Cannot reach Odoo server at {host}:{port}.")
        except Exception as e:
            logger.warning(f"Connection test warning: {e}")

    def _search(self, model: str, domain: list, limit: Optional[int] = None) -> List[int]:
        kwargs = {}
        if limit is not None:
            kwargs['limit'] = limit
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'search',
            [domain],
            kwargs
        )

    def _write(self, model: str, ids: List[int], vals: dict) -> bool:
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'write',
            [ids, vals]
        )

    def find_template_by_name(self, name: str) -> Optional[int]:
        """Search quality.spreadsheet.template by name. Returns template id or None."""
        if not name or not str(name).strip():
            return None
        name = str(name).strip()
        ids = self._search('quality.spreadsheet.template', [('name', '=', name)], limit=1)
        return ids[0] if ids else None

    def find_bom_by_product_default_code(self, default_code: str) -> Optional[int]:
        """Search mrp.bom by product_id.default_code. Returns first BOM id or None."""
        if default_code is None:
            return None
        default_code = str(default_code).strip()
        if not default_code:
            return None
        # Search BOMs whose product has this default_code
        ids = self._search(
            'mrp.bom',
            [('product_id.default_code', '=', default_code)],
            limit=1
        )
        return ids[0] if ids else None

    def parse_workbook(self, excel_path: str) -> List[Dict]:
        """
        Parse Excel: row 1 = headers, from row 2 data.
        Column B (index 1) = template name, Column C (index 2) = product default code (Com No).
        Returns list of dicts: template_name, product_code, row_index.
        """
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        rows: List[Dict] = []
        ws = wb.active
        if not ws or (ws.max_row or 0) < 2:
            wb.close()
            return rows

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            template_name = row[1] if len(row) > 1 else None
            product_code = row[2] if len(row) > 2 else None
            if not template_name or not product_code:
                continue
            template_name = str(template_name).strip()
            product_code = str(product_code).strip()
            if not template_name or not product_code:
                continue
            rows.append({
                'template_name': template_name,
                'product_code': product_code,
                'row_index': row_idx,
            })

        wb.close()
        logger.info(f"Parsed {len(rows)} data rows from workbook")
        return rows

    def import_qc_links(self, excel_path: str, dry_run: bool = False) -> Dict:
        """Main import: for each row, find template by name, find BOM by product code, set qc_template_id."""
        stats = {
            'total_rows': 0,
            'linked': 0,
            'skipped_no_template': 0,
            'skipped_no_bom': 0,
            'errors': [],
        }

        records = self.parse_workbook(excel_path)
        stats['total_rows'] = len(records)

        if dry_run:
            logger.info("DRY RUN MODE - No records will be updated")

        for rec in records:
            template_name = rec['template_name']
            product_code = rec['product_code']
            row_idx = rec['row_index']
            try:
                template_id = self.find_template_by_name(template_name)
                if not template_id:
                    logger.warning(f"Row {row_idx}: QC template not found: {template_name!r}")
                    stats['skipped_no_template'] += 1
                    continue

                bom_id = self.find_bom_by_product_default_code(product_code)
                if not bom_id:
                    logger.warning(f"Row {row_idx}: BOM not found for product default_code: {product_code!r}")
                    stats['skipped_no_bom'] += 1
                    continue

                if not dry_run:
                    self._write('mrp.bom', [bom_id], {'qc_template_id': template_id})
                    stats['linked'] += 1
                    logger.info(
                        f"Row {row_idx}: Linked BOM id={bom_id} (product {product_code!r}) "
                        f"to QC template id={template_id} ({template_name!r})"
                    )
                else:
                    stats['linked'] += 1
                    logger.info(
                        f"[DRY RUN] Row {row_idx}: Would link BOM (product {product_code!r}) "
                        f"to QC template {template_name!r}"
                    )
            except Exception as e:
                msg = f"Row {row_idx}: Error linking {product_code!r} -> {template_name!r}: {e}"
                logger.error(msg, exc_info=True)
                stats['errors'].append(msg)

        return stats


# ---------------------------------------------------------
# CLI
# ---------------------------------------------------------

def main():
    url = ODOO_URL
    db = ODOO_DB
    username = ODOO_USERNAME
    password = ODOO_PASSWORD
    excel_file = QC_TEMPLATE_EXCEL_FILE
    dry_run = QC_TEMPLATE_DRY_RUN

    if '--test' in sys.argv:
        try:
            OdooQCTemplateImporter(url, db, username, password)
            print("Connection successful!")
        except Exception as e:
            print(f"Connection failed: {e}")
            sys.exit(1)
        return

    if len(sys.argv) > 1 and sys.argv[1] not in ('--execute', '--test'):
        excel_file = sys.argv[1]
    if '--execute' in sys.argv:
        dry_run = False

    excel_path = Path(excel_file)
    if not excel_path.is_absolute():
        excel_path = script_dir / excel_path

    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        sys.exit(1)

    try:
        importer = OdooQCTemplateImporter(url, db, username, password)
        stats = importer.import_qc_links(str(excel_path), dry_run=dry_run)

        print("\n" + "=" * 60)
        print("QC TEMPLATE IMPORT STATISTICS")
        print("=" * 60)
        print(f"Rows processed:           {stats['total_rows']}")
        print(f"BOMs linked to template:  {stats['linked']}")
        print(f"Skipped (no template):    {stats['skipped_no_template']}")
        print(f"Skipped (no BOM):         {stats['skipped_no_bom']}")
        print(f"Errors:                   {len(stats['errors'])}")
        if stats['errors']:
            for err in stats['errors'][:20]:
                print(f"  - {err}")
            if len(stats['errors']) > 20:
                print(f"  ... and {len(stats['errors']) - 20} more")
        print("=" * 60)
        if dry_run:
            print("DRY RUN - no changes written. Use --execute to apply.")
    except ConnectionError as e:
        print(f"Connection error: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Import failed: {e}", exc_info=True)
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
