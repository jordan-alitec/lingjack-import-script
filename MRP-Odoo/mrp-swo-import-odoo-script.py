#!/usr/bin/env python3
"""
Odoo 18 MRP + Sale Work Order Import Script

Imports Outstanding-MO.xlsx (mrp.production) and SWO template (e.g. 260201_swo-template.xlsx)
(sale.work.order + sale.work.order.line) via XML-RPC in a strict 7-step sequence:

1. Create all mrp.production from Outstanding-MO.xlsx (draft).
2. Create sale.work.order (grouped by SWO number + SO) from SWO template, only where MO exists by PWO.
3. action_confirm on all created SWOs.
4. Link SWO/SWO lines to MO (sale_work_order_ids, sale_work_order_line_ids, production_ids).
5. action_confirm on all created MOs.
6. action_start on MOs where Excel state is "in progress".
7. Write date_start/date_finished from Excel; swap_old_name on MO and SWO.
"""

import sys
import logging
import os
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime, date
from collections import defaultdict

import xmlrpc.client
from openpyxl import load_workbook

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
logger.handlers = []

console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
logger.addHandler(console_handler)

log_file_path = os.path.join(os.path.dirname(__file__), 'mrp_swo_import_errors.log')
file_handler = logging.FileHandler(log_file_path, mode='a', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(logging.Formatter(
    "%(asctime)s - %(levelname)s - %(name)s - %(message)s",
    datefmt='%Y-%m-%d %H:%M:%S'
))
logger.addHandler(file_handler)

logger.info("=" * 60)
logger.info("MRP / SWO Import Script")
logger.info("Log file: %s", log_file_path)
logger.info("=" * 60)

# Product ID to use when Excel product code is "Non-Stock" (product not stored in stock)
NON_STOCK_PRODUCT_ID = 37341

# When sale.order not found by name: temporarily link to this ID for testing (set False to skip SWO instead)
USE_FALLBACK_SALE_ORDER_WHEN_NOT_FOUND = False
FALLBACK_SALE_ORDER_ID = 500


class OdooMRPSWOImporter:
    """Import MRP productions and Sale Work Orders from Excel to Odoo 18 via RPC."""

    def __init__(self, url: str, db: str, username: str, password: str):
        self.url = url
        self.db = db
        self.username = username
        self.password = password
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
        self.uid = common.authenticate(db, username, password, {})
        if not self.uid:
            raise Exception(f"Authentication failed for '{username}' on database '{db}'.")
        self.models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")
        logger.info("Connected to Odoo DB '%s' as '%s'", db, username)

    def _search(self, model: str, domain: list, limit: Optional[int] = 1) -> List[int]:
        kwargs = {} if limit is None else {'limit': limit}
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'search',
            [domain],
            kwargs
        )

    def _create(self, model: str, vals: dict) -> int:
        filtered = {k: v for k, v in vals.items() if v is not None}
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'create',
            [filtered]
        )

    def _read(self, model: str, ids: List[int], fields: List[str]) -> List[dict]:
        if not ids:
            return []
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'read',
            [ids],
            {'fields': fields}
        )

    def _write(self, model: str, ids: List[int], vals: dict) -> bool:
        filtered = {k: v for k, v in vals.items() if v is not None}
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, 'write',
            [ids, filtered]
        )

    def _call(self, model: str, method: str, ids: List[int], *args, **kwargs) -> Any:
        return self.models.execute_kw(
            self.db, self.uid, self.password,
            model, method,
            [ids] + list(args),
            kwargs or {}
        )

    def find_product_by_default_code(self, default_code: str) -> Optional[int]:
        if not default_code or not str(default_code).strip():
            return None
        code = str(default_code).strip()
        if code.lower() == 'non-stock':
            return NON_STOCK_PRODUCT_ID
        ids = self._search('product.product', [('default_code', '=', code)], limit=1)
        return ids[0] if ids else None

    def create_product(self, reference: str, name: Optional[str] = None) -> Optional[int]:
        """
        Create product.template (and return product.product id) with BOM-import style vals.
        Uses name, default_code, type='consu', import_newly_created=True; optional categ_id (All), uom_id (Units).
        """
        if not reference or not str(reference).strip():
            return None
        reference = str(reference).strip()
        name = (name or reference).strip() or reference
        try:
            categ_ids = self._search('product.category', [('name', '=', 'All')], limit=1)
            categ_id = categ_ids[0] if categ_ids else None
        except Exception as e:
            logger.debug("Could not find product category: %s", e)
            categ_id = None
        uom_id = self.find_uom_by_name('Units')
        product_vals = {
            'name': name,
            'default_code': reference,
            'type': 'consu',
            'import_newly_created': True,
        }
        if categ_id:
            product_vals['categ_id'] = categ_id
        if uom_id:
            product_vals['uom_id'] = uom_id
            product_vals['uom_po_id'] = uom_id
        product_vals = {k: v for k, v in product_vals.items() if v is not None}
        try:
            template_id = self._create('product.template', product_vals)
            product_ids = self._search('product.product', [('product_tmpl_id', '=', template_id)], limit=1)
            if product_ids:
                logger.info("Created product: %s (reference=%s, id=%s)", name, reference, product_ids[0])
                return product_ids[0]
            return None
        except xmlrpc.client.Fault as e:
            logger.error("Odoo error creating product '%s': %s", reference, e)
            return None
        except Exception as e:
            logger.error("Error creating product '%s': %s", reference, e, exc_info=True)
            return None

    def find_or_create_product(self, default_code: str, name: Optional[str] = None) -> Optional[int]:
        """Find product by default_code, or create with import_newly_created=True. 'Non-Stock' returns NON_STOCK_PRODUCT_ID."""
        if not default_code or not str(default_code).strip():
            return None
        code = str(default_code).strip()
        if code.lower() == 'non-stock':
            return NON_STOCK_PRODUCT_ID
        product_id = self.find_product_by_default_code(code)
        if product_id:
            return product_id
        return self.create_product(code, name=name or code)

    def find_sale_order_by_name(self, name: str) -> Optional[int]:
        if not name or not str(name).strip():
            return None
        ids = self._search('sale.order', [('name', '=', str(name).strip())], limit=1)
        return ids[0] if ids else None

    def find_user_by_name(self, name: str) -> Optional[int]:
        if not name or not str(name).strip():
            return None
        ids = self._search('res.users', [('name', '=ilike', str(name).strip())], limit=1)
        return ids[0] if ids else None

    def find_uom_by_name(self, name: str = 'Units') -> Optional[int]:
        name = (name or 'Units').strip() or 'Units'
        ids = self._search('uom.uom', [('name', '=', name)], limit=1)
        if ids:
            return ids[0]
        ids = self._search('uom.uom', [('name', '=', 'Units')], limit=1)
        return ids[0] if ids else None

    def find_first_bom_for_product(self, product_id: int) -> Optional[int]:
        """Return first mrp.bom id for this product (by product_tmpl_id)."""
        if not product_id:
            return None
        try:
            prod = self._read('product.product', [product_id], ['product_tmpl_id'])
            if not prod or not prod[0].get('product_tmpl_id'):
                return None
            pt_id = prod[0]['product_tmpl_id'][0]
            bom_ids = self._search('mrp.bom', [('product_tmpl_id', '=', pt_id)], limit=1)
            return bom_ids[0] if bom_ids else None
        except Exception as e:
            logger.debug("find_first_bom_for_product %s: %s", product_id, e)
            return None

    def find_manufacture_picking_type_id(self) -> Optional[int]:
        """Return the manufacture picking type (manu_type_id) from the default warehouse so MO has picking_type_id/warehouse for pick component creation."""
        try:
            wh_ids = self._search('stock.warehouse', [], limit=1)
            if not wh_ids:
                return None
            wh = self._read('stock.warehouse', [wh_ids[0]], ['manu_type_id'])
            if not wh or not wh[0].get('manu_type_id'):
                return None
            return wh[0]['manu_type_id'][0]
        except Exception as e:
            logger.debug("find_manufacture_picking_type_id: %s", e)
            return None

    def _to_datetime_str(self, value) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d 00:00:00')
        if isinstance(value, str):
            value = value.strip()
            for fmt in ['%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                try:
                    dt = datetime.strptime(value, fmt)
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                except ValueError:
                    continue
        return None

    # --------------- Outstanding-MO.xlsx parsing ---------------
    # Col A=0: old_pwo_id, B=1: old_pwo_number, C=2: date_start, D=3: date_finished,
    # E=4: product (main), F=5: Quantity To Produce, G=6: state (ignore or in progress), H=7: component product, J=9: component qty,
    # L=11: old_so_id, M=12: old_so_number, N=13: old_swo_id, O=14: old_swo_number, O=14: old_qty_produced (quantity already produced)
    # MO product_qty = F - O (Quantity To Produce - quantity_produced); qty_producing is not set

    def parse_outstanding_mo(self, excel_path: str, sheet_name: Optional[str] = None, header_row: int = 1) -> Dict[str, Dict]:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        start_row = header_row + 1
        mo_data = defaultdict(lambda: {'mo_data': None, 'components': []})

        for row_idx in range(start_row, ws.max_row + 1):
            row_cells = [cell.value for cell in ws[row_idx]]
            # Require at least 10 columns (A–J) so component rows with H=code, J=qty are not skipped
            if len(row_cells) < 10:
                continue
            pwo_id = row_cells[0]
            pwo_number = row_cells[1] if len(row_cells) > 1 else None
            start_date = row_cells[2] if len(row_cells) > 2 else None
            end_date = row_cells[3] if len(row_cells) > 3 else None
            product_code = row_cells[4] if len(row_cells) > 4 else None
            product_qty = row_cells[5] if len(row_cells) > 5 else None
            state = row_cells[6] if len(row_cells) > 6 else None
            # Component: H=7 or I=8 for code, J=9 for qty (try both in case layout differs)
            component_code = (row_cells[7] if len(row_cells) > 7 else None) or (row_cells[8] if len(row_cells) > 8 else None)
            component_qty = row_cells[9] if len(row_cells) > 9 else None
            old_so_id = row_cells[11] if len(row_cells) > 11 else None
            old_so_number = row_cells[12] if len(row_cells) > 12 else None
            old_swo_id = row_cells[13] if len(row_cells) > 13 else None
            old_swo_number = row_cells[14] if len(row_cells) > 14 else None
            # Column O (index 14): old_qty_produced for mrp.production
            raw_o = row_cells[14] if len(row_cells) > 14 else None
            try:
                old_qty_produced = float(raw_o) if raw_o is not None else 0.0
            except (TypeError, ValueError):
                old_qty_produced = 0.0

            if not pwo_id:
                continue
            pwo_id_str = str(pwo_id).strip()

            # Main MO row: has product in column E
            if product_code:
                mo_data[pwo_id_str]['mo_data'] = {
                    'pwo_id': pwo_id_str,
                    'pwo_number': str(pwo_number).strip() if pwo_number else None,
                    'start_date': start_date,
                    'end_date': end_date,
                    'product_code': str(product_code).strip() if product_code else None,
                    'product_qty': float(product_qty) if product_qty else 0.0,
                    'state': str(state).strip() if state else None,
                    'old_so_id': str(old_so_id).strip() if old_so_id else None,
                    'old_so_number': str(old_so_number).strip() if old_so_number else None,
                    'old_swo_id': str(old_swo_id).strip() if old_swo_id else None,
                    'old_swo_number': str(old_swo_number).strip() if old_swo_number else None,
                    'old_qty_produced': old_qty_produced,
                    'row_index': row_idx,
                }
            # Component row: has component in column H (same row can also be main row)
            if component_code:
                mo_data[pwo_id_str]['components'].append({
                    'component_code': str(component_code).strip() if component_code else None,
                    'component_qty': float(component_qty) if component_qty else 0.0,
                    'row_index': row_idx,
                })

        wb.close()
        total_components = sum(len(d['components']) for d in mo_data.values())
        logger.info("Parsed %d PWO groups from Outstanding-MO (%d total components)", len(mo_data), total_components)
        return dict(mo_data)

    # --------------- SWO template (e.g. 260201_swo-template.xlsx) parsing ---------------
    # B=1: old_swo_number, C=2: SO number, F=5: old_issue_date, G=6: issue_by, H=7: request_date,
    # I=8: line old_pwo_number, J=9: product (default_code), M=12: committed qty, N=13: finished qty (also used for old_qty_produced when PWO not found), P=15, Q=16: remarks
    # product_qty for line = max(M, N) else M

    def parse_swo_template(self, excel_path: str, sheet_name: Optional[str] = None, header_row: int = 1) -> List[Dict]:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        start_row = header_row + 1
        records = []
        for row_idx in range(start_row, ws.max_row + 1):
            row_cells = [cell.value for cell in ws[row_idx]]
            if len(row_cells) < 17:
                continue
            swo_number = row_cells[1] if len(row_cells) > 1 else None
            so_number = row_cells[2] if len(row_cells) > 2 else None
            old_issue_date = row_cells[5] if len(row_cells) > 5 else None
            issue_by = row_cells[6] if len(row_cells) > 6 else None
            request_date = row_cells[7] if len(row_cells) > 7 else None
            old_pwo_number = row_cells[8] if len(row_cells) > 8 else None
            item_code = row_cells[9] if len(row_cells) > 9 else None
            col_m = row_cells[12] if len(row_cells) > 12 else None
            col_n = row_cells[13] if len(row_cells) > 13 else None
            try:
                m_val = float(col_m) if col_m is not None else None
            except (TypeError, ValueError):
                m_val = None
            try:
                n_val = float(col_n) if col_n is not None else None
            except (TypeError, ValueError):
                n_val = None
            product_qty = max(m_val, n_val) if (m_val is not None and n_val is not None) else (m_val if m_val is not None else n_val if n_val is not None else 0.0)
            col_p = row_cells[15] if len(row_cells) > 15 else None
            col_q = row_cells[16] if len(row_cells) > 16 else None
            if not swo_number or not item_code:
                continue
            remarks = f"Remarks (CS)\n{col_p or ''}\n\n{col_q or ''}".strip() if (col_p or col_q) else None
            # Column N value: used for sale.work.order.line.old_qty_produced when PWO is not found
            col_n_val = n_val if n_val is not None else 0.0
            records.append({
                'swo_number': str(swo_number).strip() if swo_number else None,
                'so_number': str(so_number).strip() if so_number else None,
                'old_issue_date': old_issue_date,
                'issue_by': str(issue_by).strip() if issue_by else None,
                'request_date': request_date,
                'old_pwo_number': str(old_pwo_number).strip() if old_pwo_number else None,
                'item_code': str(item_code).strip() if item_code else None,
                'product_qty': product_qty if isinstance(product_qty, (int, float)) else 0.0,
                'remarks': remarks,
                'row_index': row_idx,
                'col_n_val': col_n_val,
            })
        wb.close()
        logger.info("Parsed %d SWO rows from template", len(records))
        return records

    def _is_in_progress_state(self, state: Optional[str]) -> bool:
        if not state:
            return False
        return str(state).strip().lower() in ('in progress', 'in_progress', 'progress')

    # --------------- Step 1: Create all MOs ---------------
    def import_mrp_all(
        self,
        excel_path: str,
        sheet_name: Optional[str] = None,
        dry_run: bool = True,
    ) -> Tuple[Dict[str, int], Dict[str, bool], Dict[str, Dict], Dict[str, int], set]:  # mo_ids_with_non_stock: set of mo_id
        """Create all mrp.production from Outstanding-MO. Returns (mo_map, in_progress_map, mo_data_by_pwo, mo_product_map, mo_ids_with_non_stock). MOs with Non-Stock component are created but not confirmed (ids in mo_ids_with_non_stock)."""
        mo_data = self.parse_outstanding_mo(excel_path, sheet_name)
        mo_map = {}
        in_progress_map = {}
        mo_data_by_pwo = {}
        mo_product_map = {}
        mo_ids_with_non_stock = set()

        for pwo_id, data in mo_data.items():
            mo_info = data.get('mo_data')
            if not mo_info:
                logger.warning("PWO %s: no main MO data (missing product column)", pwo_id)
                continue
            product_id = self.find_or_create_product(mo_info['product_code'], name=mo_info.get('product_code'))
            if not product_id:
                logger.error("PWO %s: product not found/created for '%s'", pwo_id, mo_info['product_code'])
                continue
            product_data = self._read('product.product', [product_id], ['uom_id'])
            uom_id = product_data[0]['uom_id'][0] if product_data and product_data[0].get('uom_id') else None

            move_raw_vals = []
            has_non_stock_component = False
            components = data.get('components', [])
            for comp in components:
                comp_product_id = self.find_or_create_product(comp['component_code'])
                if not comp_product_id:
                    logger.warning("PWO %s: component product not found/created '%s'", pwo_id, comp['component_code'])
                    continue
                if comp_product_id == NON_STOCK_PRODUCT_ID:
                    has_non_stock_component = True
                comp_uom_id = None
                try:
                    comp_data = self._read('product.product', [comp_product_id], ['uom_id'])
                    if comp_data and comp_data[0].get('uom_id'):
                        comp_uom_id = comp_data[0]['uom_id'][0]
                except Exception:
                    pass
                raw_vals = {
                    'product_id': comp_product_id,
                    'product_uom_qty': comp['component_qty'],
                    # 'quantity': comp['component_qty'],
                }
                if comp_uom_id:
                    raw_vals['product_uom'] = comp_uom_id
                move_raw_vals.append((0, 0, raw_vals))
            if components and not move_raw_vals:
                logger.warning("PWO %s: parsed %d component row(s) but none added to move_raw_ids", pwo_id, len(components))

            # product_qty = Quantity To Produce - quantity_produced (column O); do not set qty_producing
            qty_to_produce = float(mo_info['product_qty'])
            qty_produced = float(mo_info.get('old_qty_produced', 0.0) or 0.0)
            mo_product_qty = max(0.0, qty_to_produce - qty_produced)

            if mo_product_qty == 0:
                logger.info(
                    "Skipped PWO %s (old_pwo_number=%s, product=%s): product_qty would be 0 (Quantity To Produce %.2f - quantity_produced %.2f)",
                    pwo_id, mo_info.get('pwo_number'), mo_info.get('product_code'), qty_to_produce, qty_produced
                )
                continue

            picking_type_id = self.find_manufacture_picking_type_id()
            mo_vals = {
                'product_id': product_id,
                'product_qty': mo_product_qty,
                'product_uom_id': uom_id,
                'date_start': self._to_datetime_str(mo_info['start_date']),
                'date_finished': self._to_datetime_str(mo_info['end_date']),
                'old_pwo_id': mo_info['pwo_id'],
                'old_pwo_number': mo_info['pwo_number'],
                'old_so_id': mo_info['old_so_id'],
                'old_so_number': mo_info['old_so_number'],
                'old_swo_id': mo_info['old_swo_id'],
                'old_swo_number': mo_info['old_swo_number'],
                'old_qty_produced': float(mo_info.get('old_qty_produced', 0.0) or 0.0),
                'bom_id': False,
                'warehouse_id': 1, # hardcode here for auto pick component
            }
            if picking_type_id:
                mo_vals['picking_type_id'] = picking_type_id
            if move_raw_vals:
                mo_vals['move_raw_ids'] = move_raw_vals

            if not dry_run:
                mo_id = self._create('mrp.production', mo_vals)
                pwo_key = mo_info['pwo_number'] or pwo_id
                mo_map[pwo_key] = mo_id
                in_progress_map[pwo_key] = self._is_in_progress_state(mo_info.get('state'))
                mo_data_by_pwo[pwo_key] = mo_info
                mo_product_map[pwo_key] = product_id
                if has_non_stock_component:
                    mo_ids_with_non_stock.add(mo_id)
                    logger.info(
                        "Created MO %s (old_pwo_number=%s) with %d component(s) [Non-Stock: will not confirm]",
                        mo_id, mo_info['pwo_number'], len(move_raw_vals)
                    )
                else:
                    logger.info(
                        "Created MO %s (old_pwo_number=%s) with %d component(s)",
                        mo_id, mo_info['pwo_number'], len(move_raw_vals)
                    )
            else:
                logger.info("[DRY RUN] Would create MO for PWO %s", pwo_id)

        return mo_map, in_progress_map, mo_data_by_pwo, mo_product_map, mo_ids_with_non_stock

    # --------------- Step 2: Create SWOs (grouped; create all groups even if column I / PWO is empty) ---------------
    def import_swo_where_mo_exists(
        self,
        swo_excel_path: str,
        mo_map: Dict[str, int],
        sheet_name: Optional[str] = None,
        dry_run: bool = True,
    ) -> Tuple[List[int], List[Dict]]:
        """Create sale.work.order for all groups (B, C); create even when column I (old_pwo_number) is empty. Returns (swo_ids, lines_with_pwo) with product_id on each line for linkage."""
        records = self.parse_swo_template(swo_excel_path, sheet_name)
        groups = defaultdict(list)
        for rec in records:
            key = (rec['swo_number'], rec['so_number'])
            groups[key].append(rec)

        swo_ids = []
        lines_with_pwo = []

        for (swo_number, so_number), rows in groups.items():
            sale_order_id = self.find_sale_order_by_name(so_number)
            company_id = None
            if not sale_order_id:
                if USE_FALLBACK_SALE_ORDER_WHEN_NOT_FOUND:
                    sale_order_id = FALLBACK_SALE_ORDER_ID
                    logger.warning(
                        "SWO %s: sale.order not found for name '%s', using fallback sale_order_id=%s (set USE_FALLBACK_SALE_ORDER_WHEN_NOT_FOUND=False to create without sale order)",
                        swo_number, so_number, sale_order_id,
                    )
                else:
                    # Create SWO even when no sale order is linked
                    logger.warning(
                        "SWO %s: sale.order not found for name '%s', creating SWO without linked sale.order",
                        swo_number, so_number,
                    )
                    sale_order_id = False
            if sale_order_id:
                so_data = self._read('sale.order', [sale_order_id], ['company_id'])
                company_id = so_data[0]['company_id'][0] if so_data and so_data[0].get('company_id') else None

            request_date = self._to_datetime_str(rows[0].get('request_date'))
            old_issue_date = self._to_datetime_str(rows[0].get('old_issue_date'))
            remarks = rows[0].get('remarks') if rows else None

            swo_vals = {
                'sale_order_id': sale_order_id or False,
                'old_swo_number': swo_number,
                'old_so_number': so_number,
                'old_issue_date': old_issue_date,
                'request_date': request_date or datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'remarks': remarks,
            }
            if company_id:
                swo_vals['company_id'] = company_id
            if rows[0].get('issue_by'):
                swo_vals['issue_by'] = rows[0].get('issue_by')

            if dry_run:
                logger.info("[DRY RUN] Would create SWO %s with %d lines", swo_number, len(rows))
                continue

            swo_id = self._create('sale.work.order', swo_vals)
            swo_ids.append(swo_id)

            for r in rows:
                item_code = r.get('item_code')
                if not item_code or not str(item_code).strip():
                    logger.debug("Row %s: skip line (product/item_code is empty)", r['row_index'])
                    continue
                product_id = self.find_or_create_product(item_code, name=item_code)
                if not product_id:
                    logger.warning("Row %s: product not found/created '%s', skip line", r['row_index'], item_code)
                    continue
                product_data = self._read('product.product', [product_id], ['uom_id'])
                uom_id = product_data[0]['uom_id'][0] if product_data and product_data[0].get('uom_id') else None
                line_vals = {
                    'work_order_id': swo_id,
                    'product_id': product_id,
                    'product_qty': r['product_qty'],
                    'product_uom_id': uom_id,
                    'old_pwo_number': r.get('old_pwo_number'),
                    'remarks': r.get('remarks'),
                }
                line_id = self._create('sale.work.order.line', line_vals)
                lines_with_pwo.append({
                    'line_id': line_id,
                    'work_order_id': swo_id,
                    'old_pwo_number': r.get('old_pwo_number'),
                    'product_id': product_id,
                    'sale_order_id': sale_order_id,
                    'col_n_val': r.get('col_n_val', 0.0),
                })

        return swo_ids, lines_with_pwo

    # --------------- Step 3: Confirm all SWOs ---------------
    def confirm_swo_all(self, swo_ids: List[int], dry_run: bool = True) -> None:
        if dry_run or not swo_ids:
            return
        for sid in swo_ids:
            try:
                self._call('sale.work.order', 'action_confirm', [sid])
                logger.info("Confirmed SWO id=%s", sid)
            except Exception as e:
                logger.error("Failed to confirm SWO id=%s: %s", sid, e)

    # --------------- Step 4: Link SWO to MO (match by old_pwo_number and product_id) ---------------
    def link_swo_to_mo(
        self,
        mo_map: Dict[str, int],
        mo_product_map: Dict[str, int],
        lines_with_pwo: List[Dict],
        dry_run: bool = True,
    ) -> None:
        if dry_run or not lines_with_pwo:
            return
        linked_line_ids = set()
        # Link MO only to SWO lines that match both old_pwo_number (PWO name) and product_id
        for old_pwo_number, mo_id in mo_map.items():
            mo_product_id = mo_product_map.get(old_pwo_number)
            if not mo_product_id:
                continue
            lines_for_mo = [
                L for L in lines_with_pwo
                if L.get('old_pwo_number') == old_pwo_number and L.get('product_id') == mo_product_id
            ]
            line_ids = [L['line_id'] for L in lines_for_mo]
            if not line_ids:
                continue
            line_data = self._read('sale.work.order.line', line_ids, ['work_order_id'])
            work_order_ids = list({r['work_order_id'][0] for r in line_data if r.get('work_order_id')})
            swo_data = self._read('sale.work.order', work_order_ids, ['sale_order_id'])
            sale_order_ids = list({r['sale_order_id'][0] for r in swo_data if r.get('sale_order_id')})

            try:
                self._write('mrp.production', [mo_id], {
                    'sale_work_order_ids': [(6, 0, work_order_ids)],
                    'sale_work_order_line_ids': [(6, 0, line_ids)],
                    'sale_order_ids': [(6, 0, sale_order_ids)],
                })
                # Link MO and set old_qty_produced from column N for all lines with this PWO
                for L in lines_for_mo:
                    self._write('sale.work.order.line', [L['line_id']], {
                        'production_ids': [(4, mo_id)],
                        'old_qty_produced': float(L.get('col_n_val', 0.0)),
                    })
                    linked_line_ids.add(L['line_id'])
                logger.info("Linked MO %s (PWO=%s, product_id=%s) to SWO lines %s", mo_id, old_pwo_number, mo_product_id, line_ids)
            except Exception as e:
                logger.error("Failed to link MO %s: %s", mo_id, e)

        # When PWO is not found, set sale.work.order.line.old_qty_produced from column N (linked lines already set above)
        for L in lines_with_pwo:
            if not L.get('old_pwo_number'):
                continue
            if L['line_id'] in linked_line_ids:
                continue
            col_n_val = L.get('col_n_val')
            if col_n_val is None:
                continue
            try:
                self._write('sale.work.order.line', [L['line_id']], {'old_qty_produced': float(col_n_val)})
                logger.info("Set old_qty_produced=%.2f on SWO line %s (PWO %s not found)", float(col_n_val), L['line_id'], L.get('old_pwo_number'))
            except Exception as e:
                logger.warning("Failed to set old_qty_produced on line %s: %s", L['line_id'], e)

    # --------------- Step 5: Confirm all MOs (skip Non-Stock); write dates before button_plan ---------------
    def confirm_mo_all(
        self,
        mo_map: Dict[str, int],
        mo_data_by_pwo: Dict[str, Dict],
        mo_ids_with_non_stock: set,
        dry_run: bool = True,
    ) -> None:
        if dry_run or not mo_map:
            return
        for old_pwo_number, mo_id in mo_map.items():
            if mo_id in mo_ids_with_non_stock:
                logger.info("Skip confirm MO id=%s (PWO=%s, has Non-Stock component)", mo_id, old_pwo_number)
                continue
            try:
                self._call('mrp.production', 'action_import_confirm', [mo_id])
                logger.info("Confirmed MO id=%s", mo_id)
            except Exception as e:
                logger.error("Failed to confirm MO id=%s: %s", mo_id, e)
                continue
            mo_info = mo_data_by_pwo.get(old_pwo_number)
            if mo_info:
                try:
                    self._write('mrp.production', [mo_id], {
                        'date_start': self._to_datetime_str(mo_info.get('start_date')),
                        'date_finished': self._to_datetime_str(mo_info.get('end_date')),
                    })
                    logger.debug("Wrote dates on MO id=%s (before button_plan)", mo_id)
                except Exception as e:
                    logger.warning("Failed to write dates on MO id=%s: %s", mo_id, e)
            try:
                self._call('mrp.production', 'button_plan', [mo_id])
                logger.info("Planned MO id=%s (pick component created if warehouse manufacture_steps=pbm)", mo_id)
            except Exception as e:
                logger.warning("button_plan MO id=%s: %s", mo_id, e)

    # --------------- Step 6: action_start for in-progress MOs (exclude Non-Stock MOs) ---------------
    def action_start_in_progress_mos(
        self,
        mo_ids_in_progress: List[int],
        mo_ids_with_non_stock: set,
        dry_run: bool = True,
    ) -> None:
        if dry_run or not mo_ids_in_progress:
            return
        for mo_id in mo_ids_in_progress:
            if mo_id in mo_ids_with_non_stock:
                continue
            try:
                self._call('mrp.production', 'action_start', [mo_id])
                logger.info("Started MO id=%s", mo_id)
            except Exception as e:
                logger.error("Failed action_start MO id=%s: %s", mo_id, e)

    # --------------- Step 7: Dates (draft/Non-Stock MOs only) and swap_old_name ---------------
    def apply_dates_and_swap_names(
        self,
        mo_map: Dict[str, int],
        mo_data_by_pwo: Dict[str, Dict],
        mo_ids_with_non_stock: set,
        swo_ids: List[int],
        dry_run: bool = True,
    ) -> None:
        if dry_run:
            return
        for old_pwo_number, mo_id in mo_map.items():
            if mo_id in mo_ids_with_non_stock:
                mo_info = mo_data_by_pwo.get(old_pwo_number)
                if mo_info:
                    try:
                        self._write('mrp.production', [mo_id], {
                            'date_start': self._to_datetime_str(mo_info.get('start_date')),
                            'date_finished': self._to_datetime_str(mo_info.get('end_date')),
                        })
                        logger.debug("Wrote dates on draft MO id=%s (Non-Stock)", mo_id)
                    except Exception as e:
                        logger.warning("Failed to write dates on MO %s: %s", mo_id, e)
            try:
                self._call('mrp.production', 'swap_old_name', [mo_id])
            except Exception as e:
                logger.error("Failed swap_old_name MO %s: %s", mo_id, e)
        for swo_id in swo_ids:
            try:
                self._call('sale.work.order', 'swap_old_name', [swo_id])
            except Exception as e:
                logger.error("Failed swap_old_name SWO %s: %s", swo_id, e)

    # --------------- Full run ---------------
    def run_import(
        self,
        mo_excel_path: str,
        swo_excel_path: str,
        mo_sheet_name: Optional[str] = None,
        swo_sheet_name: Optional[str] = None,
        dry_run: bool = True,
    ) -> None:
        logger.info("Step 1: Import all MO from Outstanding-MO")
        mo_map, in_progress_map, mo_data_by_pwo, mo_product_map, mo_ids_with_non_stock = self.import_mrp_all(
            mo_excel_path, mo_sheet_name, dry_run
        )
        if not mo_map:
            logger.warning("No MOs created; aborting rest of import.")
            return

        logger.info("Step 2: Import SWO (all groups; column I may be empty)")
        swo_ids, lines_with_pwo = self.import_swo_where_mo_exists(swo_excel_path, mo_map, swo_sheet_name, dry_run)

        logger.info("Step 3: Confirm all SWO")
        self.confirm_swo_all(swo_ids, dry_run)

        logger.info("Step 4: Link SWO to MO (by PWO name + product_id)")
        self.link_swo_to_mo(mo_map, mo_product_map, lines_with_pwo, dry_run)

        logger.info("Step 5: Confirm MO (skip Non-Stock); write dates before button_plan")
        self.confirm_mo_all(mo_map, mo_data_by_pwo, mo_ids_with_non_stock, dry_run)

        logger.info("Step 6: action_start for in-progress MOs (exclude Non-Stock)")
        in_progress_ids = [mo_id for pwo, mo_id in mo_map.items() if in_progress_map.get(pwo)]
        self.action_start_in_progress_mos(in_progress_ids, mo_ids_with_non_stock, dry_run)

        logger.info("Step 7: Apply dates (draft/Non-Stock MOs only) and swap_old_name")
        self.apply_dates_and_swap_names(mo_map, mo_data_by_pwo, mo_ids_with_non_stock, swo_ids, dry_run)

        logger.info("Import complete. MOs: %d, SWOs: %d", len(mo_map), len(swo_ids))


def main():
    from pathlib import Path
    script_dir = Path(__file__).resolve().parent
    repo_root = script_dir.parent
    default_mo_path = repo_root / 'MRP' / 'Raw Script' / 'Outstanding-MO.xlsx'
    default_swo_path = repo_root / 'SWO' / 'Raw Script' / '260204_swo-template.xlsx'

    # ODOO_URL = os.environ.get('ODOO_URL', 'http://localhost:8069')
    # ODOO_DB = os.environ.get('ODOO_DB', 'lingjack-test')
    # ODOO_USERNAME = os.environ.get('ODOO_USERNAME', 'admin')
    # ODOO_PASSWORD = os.environ.get('ODOO_PASSWORD', 'admin')

    # ODOO_URL = 'https://lingjack.odoo.com/'
    # ODOO_DB = 'alitecpteltd-lingjack-main-21976694'
    # ODOO_USERNAME = 'dataimport'
    # ODOO_PASSWORD = 'Admin@123456'

    # ODOO_URL = 'http://localhost:8099'
    # ODOO_DB = 'lingjack-test4'
    # ODOO_USERNAME = 'dataimport'
    # ODOO_PASSWORD = 'Admin@123456'

    ODOO_URL = 'https://lingjack-data-migration-script-28135253.dev.odoo.com'
    ODOO_DB = 'lingjack-data-migration-script-28135253'
    ODOO_USERNAME = 'dataimport'
    ODOO_PASSWORD = 'Admin@123456'

    MO_EXCEL = os.environ.get('MRP_MO_EXCEL', str(default_mo_path))
    SWO_EXCEL = os.environ.get('MRP_SWO_EXCEL', str(default_swo_path))
    DRY_RUN = os.environ.get('MRP_SWO_DRY_RUN', '1').lower() in ('1', 'true', 'yes')

    if '--execute' in sys.argv:
        DRY_RUN = False
    if '--dry-run' in sys.argv:
        DRY_RUN = True

    if not Path(MO_EXCEL).is_absolute():
        candidate = repo_root / MO_EXCEL
        MO_EXCEL = str(candidate if candidate.exists() else script_dir / MO_EXCEL)
    if not Path(SWO_EXCEL).is_absolute():
        candidate = repo_root / SWO_EXCEL
        SWO_EXCEL = str(candidate if candidate.exists() else script_dir / SWO_EXCEL)

    logger.info("MO Excel: %s", MO_EXCEL)
    logger.info("SWO Excel: %s", SWO_EXCEL)
    logger.info("Dry run: %s", DRY_RUN)

    importer = OdooMRPSWOImporter(ODOO_URL, ODOO_DB, ODOO_USERNAME, ODOO_PASSWORD)
    importer.run_import(
        mo_excel_path=MO_EXCEL,
        swo_excel_path=SWO_EXCEL,
        dry_run=DRY_RUN,
    )


if __name__ == '__main__':
    main()
