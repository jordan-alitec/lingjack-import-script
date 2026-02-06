#!/usr/bin/env python3
"""
Filter Manufacturing Order Excel: keep only rows where column G (State) is not "complete".
Output: Outstanding-MO.xlsx in the same folder.
"""

from pathlib import Path
from openpyxl import load_workbook, Workbook

# --- CONFIG ---
SCRIPT_DIR = Path(__file__).parent.resolve()
INPUT_EXCEL = SCRIPT_DIR / "250131_Manufacturing Order-template.xlsx"
OUTPUT_EXCEL = SCRIPT_DIR / "Outstanding-MO.xlsx"
COL_G = 7  # Column G = State (1-based)
HEADER_ROW = 1
# --------------
VALUE_COMPLETE = "complete"


def _cell_value_str(ws, row: int, col: int):
    """Get cell value as stripped lowercase string for comparison."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip().lower()


def _copy_row(ws_src, row_src: int, ws_dst, row_dst: int, max_col: int):
    """Copy one row from source sheet to destination sheet (values only)."""
    for col in range(1, max_col + 1):
        dst_cell = ws_dst.cell(row=row_dst, column=col)
        dst_cell.value = ws_src.cell(row=row_src, column=col).value


def main():
    if not INPUT_EXCEL.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_EXCEL}")

    wb = load_workbook(INPUT_EXCEL, data_only=False)
    ws = wb.active

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title

    max_col = ws.max_column
    # Copy header row
    _copy_row(ws, HEADER_ROW, out_ws, HEADER_ROW, max_col)
    out_row = HEADER_ROW + 1

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        state_val = _cell_value_str(ws, row, COL_G)
        if state_val == VALUE_COMPLETE:
            continue
        _copy_row(ws, row, out_ws, out_row, max_col)
        out_row += 1

    out_wb.save(OUTPUT_EXCEL)
    kept = out_row - 1 - HEADER_ROW
    total_data = ws.max_row - HEADER_ROW
    print(f"Done. Kept {kept} rows (excluding 'complete'). Total data rows was {total_data}.")
    print(f"Saved to: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()
